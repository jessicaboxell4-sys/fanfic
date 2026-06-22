"""``routes/books.py`` — core book lifecycle, upload pipeline, metadata.

This file is the heart of Shelfsort. As the app grew, the more
specialised endpoint clusters got peeled off into their own modules. The
table below is the current map so future maintainers can find things
fast (line numbers are approximate; sections move as code is added).

================ STILL IN THIS FILE ================
Section / lines
  Upload & ingestion           : single-file + bulk POST /api/books/upload
  EPUB metadata extraction     : extract_epub_metadata, extract_chapters,
                                 diff_chapters, _normalize_title_for_match,
                                 _updated_shelf_name, OLD_STORIES_SHELF
  Fanfic detection             : detect_source_from_text, find_duplicate_candidates,
                                 normalize_fanfic_url, fanfic-URL canonicalisation
  Single-book CRUD             : GET/PATCH/DELETE /api/books/{book_id}
  Reader assets                : GET /api/books/{book_id}/download
                                 GET /api/books/{book_id}/cover/{filename}
                                 GET /api/books/{book_id}/download-original
                                 GET /api/books/{book_id}/diff (vs previous version)
  Library listings & filters   : GET /api/library/all and friends (by category,
                                 status, fandom, pairing, tag)
  Categories / shelves         : POST/DELETE /api/categories,
                                 POST /api/books/{book_id}/category
  Cover regeneration           : POST /api/books/{book_id}/cover/regenerate
  Manual status mutator        : PATCH /api/books/{book_id}/status
  Trash                        : (extracted to routes/trash.py — pre Phase 5)
  Relationships / pairings     : last block in the file, /api/relationships*

================ EXTRACTED MODULES ================
routes/refresh.py             : POST /api/books/{book_id}/refresh (Phase 4)
routes/duplicates.py          : auto-pending-duplicate badge helpers (Phase 4)
routes/duplicate_resolution.py: POST /api/books/{book_id}/resolve-duplicate
                                POST /api/books/resolve-group
                                GET  /api/library/duplicates(/count) (Phase 5D)
routes/library_views.py       : GET /api/library/trends, status-counts,
                                complete, ongoing, linkless, unreadable
                                + _status_query / _list_status_shelf (Phase 5E)
routes/reading_activity.py    : POST /api/books/{id}/mark, /heartbeat,
                                /progress, /touch + _log_activity (Phase 5F)
routes/url_lists.py           : POST /api/url-lists/scan + helpers (Phase 5A)
routes/fandoms.py             : GET /api/fandoms (community list) (Phase 5B)
routes/exports.py             : GET /api/library/download(?kind=xlsx)
                                + ZIP/XLSX builders (Phase 5C)
routes/conversions.py         : POST /api/library/originals/{id}/convert
                                + bulk convert (Phase 1)
routes/trash.py               : GET /api/library/trash, restore, empty (Phase 2)

The shared helpers (extract_chapters, diff_chapters, OLD_STORIES_SHELF,
_normalize_title_for_match, etc.) live HERE because the upload + refresh
pipelines depend on them.  Extracted modules import them by name from
``routes.books``; that import is one-way (no cycles).
"""
from fastapi import (
    APIRouter, UploadFile, File, HTTPException, Request, Response,
    Depends, Form, Query,
)
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from starlette.background import BackgroundTask
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta, date
from pathlib import Path
import os
import io
import re
import json
import uuid
import base64
import zipfile
import asyncio
import tempfile
import secrets
import bcrypt
import resend
import requests as http_requests

import ebooklib
from ebooklib import epub
from bs4 import BeautifulSoup

from deps import (
    db, app, api_router, logger, ROOT_DIR, STORAGE_DIR,
    EMERGENT_LLM_KEY, RESET_TOKEN_TTL_HOURS, RESEND_API_KEY,
    SENDER_EMAIL, FRONTEND_URL,
)
from models import User, BookOut
from auth_dep import get_current_user, get_current_user_or_none, require_admin
from utils.admin_audit import record_admin_action


from emergentintegrations.llm.chat import LlmChat, UserMessage

from utils.cover_notifications import (
    notify_vote_milestone,
    notify_import_milestone,
    notify_friends_of_new_share,
)


# Heuristic fandom detection. Keys are the canonical shelf name (AO3-style
# canonicals where reasonable — see https://archiveofourown.org/wrangling for
# AO3's fandom-tag convention. When adding NEW fandoms, prefer AO3's exact
# canonical form, e.g. `Stargate SG-1`, `Stargate Atlantis`, `Stargate
# (Movies)` rather than colloquial short names. The umbrella term
# `Stargate - All Media Types` is intentionally NOT used as a default — we
# bucket into the specific sub-fandom so the user can find SG-1 vs Atlantis
# works at a glance, with a cross-listing shelf already auto-built when a
# work spans multiple sub-fandoms.
FANDOM_KEYWORDS = {
    "Harry Potter": ["harry potter", "hogwarts", "hermione", "voldemort", "dumbledore", "weasley", "snape", "draco malfoy", "ron weasley"],
    "Twilight": ["twilight saga", "bella swan", "edward cullen", "stephenie meyer", "forks washington", "jacob black", "cullen family"],
    "Marvel": ["avengers", "iron man", "tony stark", "spider-man", "spider man", "captain america", "marvel comics", "x-men", "wolverine"],
    "DC Comics": ["batman", "superman", "wonder woman", "gotham", "bruce wayne", "clark kent", "dc comics"],
    "Star Wars": ["star wars", "jedi", "sith", "skywalker", "darth vader", "obi-wan", "the force"],
    "Lord of the Rings": ["lord of the rings", "frodo", "gandalf", "middle-earth", "middle earth", "hobbit", "tolkien"],
    "Sherlock Holmes": ["sherlock holmes", "221b baker", "john watson", "moriarty"],
    "Percy Jackson and the Olympians": [
        "percy jackson", "camp half-blood", "rick riordan",
        "annabeth chase", "olympians", "lightning thief", "pjo",
        "percy jackson and the olympians", "son of poseidon",
        "grover underwood", "luke castellan",
        # Direct-sequel novels — they file under PJO on AO3, so we use
        # title keywords rather than spawning new fandoms.
        "chalice of the gods", "wrath of the triple goddess",
        # Mythology compendium books (Greek Gods/Heroes/etc.) and
        # companion guides also file under PJO.
        "percy jackson's greek gods", "percy jackson's greek heroes",
        "camp half-blood confidential", "the demigod files",
        "demigods and magicians",
    ],
    "Percy Jackson and the Olympians (TV)": [
        "percy jackson and the olympians tv", "pjo tv", "disney+ percy jackson",
        "walker scobell",
    ],
    "Heroes of Olympus": [
        "heroes of olympus", "lost hero", "son of neptune",
        "mark of athena", "house of hades", "blood of olympus",
        "jason grace", "piper mclean", "leo valdez", "frank zhang",
        "hazel levesque", "nico di angelo",
    ],
    "Trials of Apollo": [
        "trials of apollo", "lester papadopoulos", "the hidden oracle",
        "dark prophecy", "the burning maze", "tyrant's tomb",
        "tower of nero",
    ],
    "The Sun and the Star": [
        "the sun and the star", "nico di angelo and will solace",
        # Riordan + Mark Oshiro co-authored Nico/Will spinoff novel
        "from the world of percy jackson the sun and the star",
    ],
    "Magnus Chase and the Gods of Asgard": [
        "magnus chase", "gods of asgard", "sword of summer",
        "hammer of thor", "ship of the dead",
    ],
    "The Kane Chronicles": [
        "kane chronicles", "carter kane", "sadie kane",
        "red pyramid", "throne of fire", "serpent's shadow",
    ],
    "Daughter of the Deep": [
        "daughter of the deep", "ana dakkar", "house of nemo",
        "harding-pencroft academy",
    ],
    # ----- Cassandra Clare's Shadowhunters universe sub-series.
    # The umbrella "Shadowhunter Chronicles - Cassandra Clare" already
    # exists in the AO3 seed; we add the canonical sub-series so books
    # from each get filed under the correct shelf instead of falling
    # back to the umbrella.
    "The Mortal Instruments": [
        "mortal instruments", "city of bones", "city of ashes",
        "city of glass", "city of fallen angels", "city of lost souls",
        "city of heavenly fire", "clary fray", "jace wayland",
        "jace herondale", "alec lightwood", "magnus bane",
    ],
    "The Infernal Devices": [
        "infernal devices", "clockwork angel", "clockwork prince",
        "clockwork princess", "tessa gray", "will herondale",
        "jem carstairs",
    ],
    "The Dark Artifices": [
        "dark artifices", "lady midnight", "lord of shadows",
        "queen of air and darkness", "emma carstairs", "julian blackthorn",
    ],
    "The Last Hours": [
        "the last hours", "chain of gold", "chain of iron",
        "chain of thorns", "james herondale", "cordelia carstairs",
    ],
    "The Eldest Curses": [
        "eldest curses", "red scrolls of magic", "lost book of the white",
    ],
    "Tales from the Shadowhunter Academy": [
        "tales from the shadowhunter academy", "shadowhunter academy",
        "simon lewis", "ghosts of the shadow market",
    ],
    "Shadowhunters (TV)": [
        "shadowhunters the mortal instruments", "shadowhunters tv",
        "freeform shadowhunters",
    ],
    # ----- Brandon Sanderson Cosmere additions. Mistborn + Stormlight
    # already exist in the AO3 seed; we add the rest.
    "Warbreaker": [
        "warbreaker", "vasher", "vivenna", "siri", "lightsong", "nightblood",
    ],
    "Elantris": [
        "elantris", "raoden", "sarene", "hrathen",
    ],
    "Tress of the Emerald Sea": [
        "tress of the emerald sea", "secret projects sanderson",
        "tress and the emerald sea",
    ],
    "Yumi and the Nightmare Painter": [
        "yumi and the nightmare painter", "nightmare painter",
    ],
    "The Sunlit Man": [
        "the sunlit man", "nomad sanderson",
    ],
    # ----- Sarah J. Maas — ACOTAR + Throne of Glass are already in the
    # AO3 seed; add Crescent City to round out the franchise.
    "Crescent City": [
        "crescent city", "house of earth and blood",
        "house of sky and breath", "house of flame and shadow",
        "bryce quinlan", "hunt athalar", "danika fendyr",
    ],
    # ----- Star Wars sub-fandoms. The umbrella + All Media Types +
    # Sequel Trilogy + Clone Wars (2008) already live in the AO3 seed;
    # we add the live-action TV / KOTOR / Rebels lines so each gets
    # its own shelf instead of all crashing into "Star Wars".
    "The Mandalorian (TV)": [
        "the mandalorian", "din djarin", "grogu", "baby yoda",
        "mandalorian disney+", "mando din djarin",
    ],
    "Andor (TV)": [
        "andor disney+", "cassian andor andor series", "mon mothma andor",
        "andor (tv)",
    ],
    "Star Wars: The Bad Batch (Cartoon)": [
        "the bad batch", "clone force 99", "hunter wrecker tech",
        "omega bad batch",
    ],
    "Star Wars Rebels": [
        "star wars rebels", "ezra bridger", "kanan jarrus",
        "hera syndulla", "ghost crew",
    ],
    "Star Wars: Knights of the Old Republic": [
        "knights of the old republic", "kotor", "revan",
        "darth malak", "swtor",
    ],
    "Star Wars Visions": [
        "star wars visions",
    ],
    "Rogue One: A Star Wars Story": [
        "rogue one", "jyn erso", "cassian andor rogue one",
        "k-2so", "bodhi rook",
    ],
    # ----- Other standalone large fandoms.
    "Fairy Tail": [
        "fairy tail", "fairy tail manga", "natsu dragneel",
        "lucy heartfilia", "erza scarlet", "gray fullbuster",
        "hiro mashima",
    ],
    "Dungeons & Dragons (Role-Playing Game)": [
        "dungeons & dragons", "dungeons and dragons", "d&d 5e",
        "5e d&d", "ttrpg", "tabletop roleplaying game",
    ],
    "The Legend of Vox Machina (Cartoon)": [
        "legend of vox machina", "vox machina cartoon",
        "vox machina animated",
    ],
    "Vox Machina (Critical Role)": [
        "vox machina", "vex'ahlia", "vax'ildan", "percy de rolo",
        "grog strongjaw", "scanlan shorthalt",
    ],
    "The Mighty Nein (Critical Role)": [
        "mighty nein", "caleb widogast", "fjord stone",
        "jester lavorre", "yasha", "beauregard lionett",
        "molly molymauk", "veth brenatto", "kingsley tealeaf",
    ],
    "Bell's Hells (Critical Role)": [
        "bell's hells", "campaign 3 critical role", "imogen temult",
        "laudna critical role", "fearne calloway", "fcg ashton",
    ],
    "My Hero Academia: Vigilantes": [
        "my hero academia vigilantes", "vigilantes mha",
        "knuckleduster", "koichi haimawari", "pop step",
    ],
    # ----- Spy x Family (anime/manga).
    "Spy x Family": [
        "spy x family", "spy×family", "loid forger", "yor forger",
        "anya forger", "bond forger", "tatsuya endo",
    ],
    # ----- Star Trek spin-offs not in the AO3 seed. The 5 original
    # series + AOS already exist; we add the modern Paramount+ era.
    "Star Trek: Strange New Worlds": [
        "strange new worlds", "snw star trek", "captain pike",
        "la'an noonien-singh", "una chin-riley",
    ],
    "Star Trek: Lower Decks": [
        "lower decks star trek", "uss cerritos",
        "beckett mariner", "brad boimler", "tendi", "rutherford",
    ],
    "Star Trek: Picard": [
        "star trek picard", "jean-luc picard series", "raffi musiker",
        "soji asha", "rios picard",
    ],
    "Star Trek: Discovery": [
        "star trek discovery", "uss discovery", "michael burnham",
        "sylvia tilly", "saru discovery",
    ],
    "Star Trek: Enterprise": [
        "star trek enterprise", "jonathan archer", "uss nx-01",
        "trip tucker", "t'pol",
    ],
    "Star Trek: Prodigy": [
        "star trek prodigy", "uss protostar", "dal r'el", "gwyn",
        "kid star trek",
    ],
    # ----- Pokemon spin-offs beyond the umbrella "All Media Types".
    "Pokémon Adventures / Pokémon Special (Manga)": [
        "pokemon adventures", "pokemon special manga",
        "pokespe", "pokémon adventures", "red green blue manga",
    ],
    "Detective Pikachu": [
        "detective pikachu", "pikachu detective", "tim goodman",
    ],
    "Pokémon GO": [
        "pokemon go", "pokémon go", "niantic pokemon",
    ],
    "Honkai Impact 3rd": [
        "honkai impact 3rd", "honkai impact", "houkai impact",
    ],
    # ----- Vivziepop animated shows.
    "Hazbin Hotel": [
        "hazbin hotel", "charlie morningstar", "alastor radio demon",
        "angel dust", "lucifer morningstar", "vivziepop",
    ],
    "Helluva Boss": [
        "helluva boss", "blitzo", "stolas goetia", "moxxie",
        "millie helluva", "imp city",
    ],
    # ----- The Owl House (Dana Terrace, Disney).
    "The Owl House": [
        "the owl house", "luz noceda", "amity blight", "eda clawthorne",
        "king owl house", "lumity", "boiling isles",
    ],
    # ----- The Boys (Amazon TV series + Garth Ennis comics).
    "The Boys (TV)": [
        "the boys amazon", "the boys (tv)", "homelander", "billy butcher",
        "starlight the boys", "hughie campbell", "soldier boy",
    ],
    "The Boys (Comics)": [
        "the boys comics", "garth ennis the boys", "dynamite the boys",
    ],
    "Gen V (TV)": [
        "gen v", "godolkin university", "marie moreau",
    ],
    # ----- Steven Universe: Future spinoff (the original SU is in the seed).
    "Steven Universe: Future": [
        "steven universe future", "su future",
    ],
    # ----- Studio Ghibli filmography.
    "Spirited Away": [
        "spirited away", "sen to chihiro", "no-face", "haku spirited",
    ],
    "Howl's Moving Castle": [
        "howl's moving castle", "howls moving castle", "sophie hatter",
        "calcifer", "diana wynne jones howl",
    ],
    "Princess Mononoke": [
        "princess mononoke", "mononoke hime", "san mononoke", "ashitaka",
    ],
    "My Neighbor Totoro": [
        "my neighbor totoro", "tonari no totoro", "totoro studio ghibli",
    ],
    "Castle in the Sky": [
        "castle in the sky", "laputa castle in the sky", "tenku no shiro laputa",
    ],
    "Kiki's Delivery Service": [
        "kiki's delivery service", "kikis delivery service", "majo no takkyubin",
    ],
    "Ponyo": [
        "ponyo on the cliff", "gake no ue no ponyo",
    ],
    "The Tale of the Princess Kaguya": [
        "tale of the princess kaguya", "kaguya hime no monogatari",
    ],
    "The Wind Rises": [
        "the wind rises", "kaze tachinu", "jiro horikoshi",
    ],
    "Nausicaä of the Valley of the Wind": [
        "nausicaa of the valley of the wind", "nausicaä", "kaze no tani no nausicaa",
    ],
    # ----- Bridgerton companion / Buffyverse / Angel.
    "Queen Charlotte: A Bridgerton Story": [
        "queen charlotte bridgerton", "queen charlotte (tv)",
        "young queen charlotte", "king george iii bridgerton",
    ],
    "Angel: the Series": [
        "angel the series", "angel buffy spinoff", "angel investigations",
        "wolfram and hart", "fred burkle", "wesley wyndam-pryce",
    ],
    # ----- Procedurals & long-runners not in the seed.
    "House M.D.": [
        "house md", "house m.d.", "gregory house", "james wilson",
        "hugh laurie house", "princeton-plainsboro",
    ],
    # ----- Arrowverse (CW DC shows). They share crossovers so they're
    # almost always tagged together — group them so the Help page
    # shelves them as a single franchise.
    "Arrow (TV)": [
        "arrow tv", "oliver queen arrow", "felicity smoak", "green arrow tv",
    ],
    "The Flash (TV 2014)": [
        "the flash (tv)", "barry allen tv", "cw flash", "iris west tv",
    ],
    "Supergirl (TV 2015)": [
        "supergirl cw", "kara danvers", "supergirl tv", "alex danvers",
    ],
    "Legends of Tomorrow (TV)": [
        "legends of tomorrow", "sara lance", "ava sharpe", "waverider",
    ],
    "Batwoman (TV)": [
        "batwoman cw", "kate kane tv", "ryan wilder",
    ],
    "Black Lightning (TV)": [
        "black lightning", "jefferson pierce", "thunder lightning",
    ],
    "Stargirl (TV)": [
        "stargirl cw", "courtney whitmore", "jsa stargirl",
    ],
    "Titans (TV)": [
        "titans tv", "dc titans", "raven dick grayson titans",
    ],
    # ----- League of Legends + Arcane.
    "League of Legends": [
        "league of legends", "lol video game", "summoner's rift",
        "riot games lol",
    ],
    "Arcane: League of Legends (Cartoon)": [
        "arcane league of legends", "arcane netflix", "vi arcane",
        "jinx arcane", "caitlyn arcane", "piltover", "zaun arcane",
    ],
    # ----- Castlevania (Netflix).
    "Castlevania (Cartoon)": [
        "castlevania netflix", "trevor belmont", "alucard castlevania",
        "sypha belnades", "dracula castlevania",
    ],
    "Castlevania: Nocturne (Cartoon)": [
        "castlevania nocturne", "richter belmont nocturne",
        "annette nocturne", "maria renard",
    ],
    # ----- Wheel of Time.
    "The Wheel of Time - Robert Jordan": [
        "wheel of time", "robert jordan wot", "rand al'thor",
        "mat cauthon", "perrin aybara", "egwene al'vere", "nynaeve al'meara",
    ],
    "The Wheel of Time (TV)": [
        "wheel of time amazon", "wheel of time tv", "wot prime video",
    ],
    # ----- Good Omens novel.
    "Good Omens - Pratchett & Gaiman": [
        "good omens novel", "terry pratchett neil gaiman good omens",
        "aziraphale crowley book",
    ],
    # ----- Dragon Ball spin-offs (umbrella is in the seed).
    "Dragon Ball Z": [
        "dragon ball z", "dbz", "saiyan saga", "namek saga",
        "android saga", "buu saga",
    ],
    "Dragon Ball Super": [
        "dragon ball super", "dbs", "tournament of power", "moro arc",
        "granolah arc",
    ],
    "Dragon Ball GT": [
        "dragon ball gt", "dbgt",
    ],
    # ----- Newer K-pop groups (the older ones — BTS, ENHYPEN, TWICE,
    # SEVENTEEN, Stray Kids, ATEEZ, BLACKPINK — are in the seed).
    "TXT (Band)": [
        "tomorrow x together", "txt band", "moa fandom",
        "yeonjun soobin beomgyu taehyun huening kai",
    ],
    "aespa (Band)": [
        "aespa", "karina aespa", "winter aespa", "ningning aespa",
        "giselle aespa", "my ae",
    ],
    "NewJeans (Band)": [
        "newjeans", "njz", "minji hanni danielle haerin hyein",
    ],
    "LE SSERAFIM (Band)": [
        "le sserafim", "lesserafim", "sakura chaewon yunjin kazuha eunchae",
    ],
    "ITZY (Band)": [
        "itzy", "midzy fandom", "yeji lia ryujin chaeryeong yuna",
    ],
    "IVE (Band)": [
        "ive band", "dive fandom", "yujin gaeul rei wonyoung liz leeseo",
    ],
    # ----- Sci-fi novels / horror.
    "The Locked Tomb - Tamsyn Muir": [
        "the locked tomb", "tamsyn muir", "ninth house tamsyn",
    ],
    "Gideon the Ninth": [
        "gideon the ninth", "gideon nav", "harrowhark nonagesimus first book",
    ],
    "Harrow the Ninth": [
        "harrow the ninth", "harrowhark second book",
    ],
    "Nona the Ninth": [
        "nona the ninth", "nona's life",
    ],
    "The Murderbot Diaries - Martha Wells": [
        "murderbot", "martha wells secunit",
        "all systems red", "artificial condition",
        "rogue protocol", "exit strategy", "network effect",
        "fugitive telemetry", "system collapse murderbot",
    ],
    "Wings of Fire - Tui T. Sutherland": [
        "wings of fire", "tui sutherland", "dragonet prophecy",
        "clay wof", "tsunami wof", "glory rainwing", "starflight wof",
        "sunny sandwing", "moon wof",
    ],
    "The Inheritance Cycle - Christopher Paolini": [
        "inheritance cycle", "eragon paolini", "saphira eragon",
        "alagaesia", "eldest paolini", "brisingr", "inheritance paolini",
    ],
    "Mortal Engines - Philip Reeve": [
        "mortal engines", "philip reeve", "hester shaw",
        "tom natsworthy", "predator cities quartet",
    ],
    "King of Scars Duology - Leigh Bardugo": [
        "king of scars", "rule of wolves", "nikolai lantsov king",
        "zoya nazyalensky",
    ],
    "The Lunar Chronicles - Marissa Meyer": [
        "lunar chronicles", "marissa meyer cinder", "scarlet meyer",
        "cress meyer", "winter meyer", "linh cinder",
    ],
    "Red Rising - Pierce Brown": [
        "red rising", "pierce brown", "darrow of lykos",
        "golden son brown", "morning star brown", "iron gold",
        "dark age brown", "light bringer brown",
    ],
    "The Expanse - James S. A. Corey": [
        "the expanse novels", "james s.a. corey", "leviathan wakes",
        "rocinante expanse", "naomi nagata book", "amos burton book",
    ],
    "The Expanse (TV)": [
        "the expanse tv", "syfy expanse", "amazon expanse",
        "thomas jane miller", "steven strait holden",
    ],
    "Foundation - Isaac Asimov": [
        "foundation asimov", "hari seldon", "psychohistory",
        "second foundation", "foundation and empire",
    ],
    "Foundation (TV)": [
        "foundation apple tv", "lee pace foundation", "jared harris seldon",
    ],
    # ----- Dune (novels + Villeneuve film duology).
    "Dune - Frank Herbert": [
        "dune novels", "frank herbert dune", "paul atreides book",
        "leto atreides ii", "god emperor dune", "children of dune",
        "dune messiah", "heretics of dune", "chapterhouse dune",
    ],
    "Dune (2021)": [
        "dune 2021", "denis villeneuve dune", "timothee chalamet paul",
        "rebecca ferguson jessica dune",
    ],
    "Dune: Part Two (2024)": [
        "dune part two", "dune part 2", "feyd-rautha villeneuve",
        "zendaya chani dune",
    ],
    # ----- A Song of Ice and Fire universe.
    "A Song of Ice and Fire - George R. R. Martin": [
        "a song of ice and fire", "asoiaf", "george r.r. martin",
        "westeros novels", "dance with dragons", "storm of swords",
        "feast for crows", "winds of winter",
    ],
    "House of the Dragon (TV)": [
        "house of the dragon", "hotd", "rhaenyra targaryen tv",
        "alicent hightower tv", "daemon targaryen tv", "dance of dragons tv",
    ],
    # ----- Sandman (Neil Gaiman) — comics + Netflix.
    "The Sandman - Neil Gaiman": [
        "the sandman comics", "neil gaiman sandman", "dream of the endless",
        "morpheus sandman", "death endless", "delirium endless",
    ],
    "The Sandman (TV)": [
        "the sandman netflix", "tom sturridge dream",
    ],
    # ----- Avatar (James Cameron movies/Pandora).
    "Avatar (Pandora - James Cameron)": [
        "avatar james cameron", "pandora avatar", "na'vi", "jake sully",
        "neytiri", "avatar the way of water", "avatar 2", "avatar 3",
    ],
    # ----- More anime/manga.
    "Frieren: Beyond Journey's End": [
        "frieren beyond journey's end", "sousou no frieren",
        "frieren elf", "fern frieren", "stark frieren",
    ],
    "Vinland Saga": [
        "vinland saga", "thorfinn karlsefni", "askeladd vinland",
        "canute vinland", "makoto yukimura",
    ],
    "Spy Classroom": [
        "spy classroom", "spy kyoushitsu", "lily spy classroom", "klaus spy classroom",
    ],
    # ----- Stephen King multiverse.
    "It - Stephen King": [
        "it stephen king", "pennywise", "losers' club",
        "derry maine king", "it 2017", "it chapter two",
    ],
    "The Dark Tower - Stephen King": [
        "the dark tower", "roland deschain", "ka-tet",
        "gunslinger stephen king",
    ],
    "The Shining - Stephen King": [
        "the shining", "jack torrance", "danny torrance",
        "overlook hotel", "doctor sleep",
    ],
    "Carrie - Stephen King": [
        "carrie white", "carrie stephen king",
    ],
    "Misery - Stephen King": [
        "misery stephen king", "annie wilkes",
    ],
    "Salem's Lot - Stephen King": [
        "salem's lot", "salems lot", "ben mears king",
    ],
    # ----- Horror video-game franchises.
    "Silent Hill": [
        "silent hill", "konami silent hill", "pyramid head",
        "silent hill 2", "silent hill 3", "james sunderland",
    ],
    "Resident Evil (Video Games)": [
        "resident evil games", "biohazard", "leon kennedy",
        "chris redfield", "jill valentine", "claire redfield",
        "ada wong", "umbrella corporation",
    ],
    "Resident Evil (Movies)": [
        "resident evil movies", "milla jovovich alice",
        "resident evil welcome to raccoon city",
    ],
    # ----- More YA dystopia / continuations.
    "Divergent Trilogy - Veronica Roth": [
        "divergent", "veronica roth", "tris prior", "tobias eaton",
        "insurgent", "allegiant",
    ],
    "The Maze Runner - James Dashner": [
        "the maze runner", "james dashner", "thomas maze runner",
        "newt maze runner", "minho maze runner", "wckd",
        "the scorch trials", "death cure",
    ],
    "The Ballad of Songbirds and Snakes": [
        "ballad of songbirds and snakes", "coriolanus snow prequel",
        "lucy gray baird",
    ],
    # ----- Holly Black Spiderwick + Stolen Heir.
    "The Spiderwick Chronicles": [
        "spiderwick chronicles", "tony diterlizzi holly black",
        "jared grace", "simon grace", "mallory grace",
    ],
    "The Stolen Heir Duology - Holly Black": [
        "stolen heir", "black heart holly black", "wren stolen heir",
        "oak greenbriar",
    ],
    # ----- Discworld sub-series. The umbrella is in the seed; readers
    # often want to file books by sub-thread (Watch, Witches, Death,
    # Rincewind, Tiffany Aching) so each lands on its own shelf.
    "Discworld: City Watch": [
        "city watch", "sam vimes", "carrot ironfoundersson", "ankh-morpork city watch",
        "fred colon", "nobby nobbs",
    ],
    "Discworld: Witches": [
        "discworld witches", "granny weatherwax", "nanny ogg",
        "magrat garlick", "agnes nitt",
    ],
    "Discworld: Death": [
        "discworld death", "mort discworld", "susan sto helit",
        "death and bills", "soul music pratchett",
    ],
    "Discworld: Rincewind": [
        "rincewind", "the colour of magic", "light fantastic",
        "interesting times", "the last continent",
    ],
    "Discworld: Tiffany Aching": [
        "tiffany aching", "wee free men", "nac mac feegle",
        "hat full of sky", "wintersmith pratchett", "i shall wear midnight",
    ],
    # ----- Mistborn Era 2 (Wax & Wayne).
    "Mistborn: Wax & Wayne (Era 2)": [
        "wax and wayne", "alloy of law", "shadows of self",
        "bands of mourning", "lost metal", "waxillium ladrian",
    ],
    # ----- Elden Ring.
    "Elden Ring (Video Game)": [
        "elden ring", "tarnished", "the lands between",
        "marika", "radagon", "ranni the witch", "malenia blade of miquella",
    ],
    "Doctor Who": ["doctor who", "tardis", "the doctor", "gallifrey"],
    "Supernatural": ["supernatural fic", "dean winchester", "sam winchester", "castiel"],
    "Game of Thrones": ["game of thrones", "westeros", "jon snow", "daenerys", "targaryen", "stark family"],
    "Hunger Games": ["hunger games", "katniss everdeen", "panem", "district 12"],
    "Naruto": ["naruto uzumaki", "konoha", "sasuke uchiha", "hokage", "akatsuki"],
    "My Hero Academia": ["my hero academia", "izuku midoriya", "u.a. high", "all might", "bakugou"],
    "BTS": ["bts fanfic", "jeon jungkook", "kim taehyung", "park jimin", "min yoongi"],
    "One Direction": ["one direction", "harry styles", "louis tomlinson", "larry stylinson"],
    # ── Stargate franchise ────────────────────────────────────────────
    # AO3 canonical names. SG-1 keywords are intentionally narrow (cast
    # of SG-1, Goa'uld, Cheyenne Mountain) so they don't fire on Atlantis
    # works, and vice-versa. The bare word "stargate" alone is NOT in any
    # list — it would trip every sub-fandom — so the AI classifier
    # decides ambiguous works.
    "Stargate SG-1": [
        "stargate sg-1", "stargate sg1", "sg-1 team",
        "jack o'neill", "jack oneill", "daniel jackson",
        "samantha carter", "sam carter", "teal'c", "teal c",
        "general hammond", "cheyenne mountain", "goa'uld", "goauld",
        "asgard", "tok'ra", "tokra", "stargate program", "stargate command",
        "sgc",
    ],
    "Stargate Atlantis": [
        "stargate atlantis", "sga ",  # trailing space to avoid SGU matches
        "atlantis expedition", "john sheppard", "rodney mckay",
        "mckay/sheppard", "mcshep", "teyla emmagan", "ronon dex",
        "elizabeth weir", "carson beckett", "pegasus galaxy", "wraith",
        "puddle jumper", "ancients", "lantean",
    ],
    "Stargate Universe": [
        "stargate universe", "sgu ", "stargate sgu",
        "everett young", "nicholas rush", "eli wallace", "chloe armstrong",
        "matthew scott", "ronald greer", "icarus base", "destiny ship",
        "the destiny",
    ],
    "Stargate (Movies)": [
        "stargate movie", "stargate (movies)", "stargate 1994",
        "stargate film", "ra abydos", "abydonian",
    ],
}


# Merge in the bundled AO3 top-fandoms seed (~100 popular fandoms across
# all media types) without overriding any hand-tuned entries above. The
# bundled file uses AO3-canonical names — the existing 16 short-name
# fandoms above stay because they're the canonical form for THIS user's
# library and renaming them would migrate every existing book's shelf.
try:
    from data.ao3_top_fandoms import AO3_TOP_FANDOMS  # noqa: WPS433
    for _canon, _kws in AO3_TOP_FANDOMS.items():
        FANDOM_KEYWORDS.setdefault(_canon, _kws)
    del _canon, _kws  # housekeeping
except Exception as _e:  # pragma: no cover — bundled file is always present
    logger.warning("Could not load AO3 top-fandoms seed: %s", _e)

FANFIC_SIGNALS = [
    "fanfiction", "fan fiction", "fanfic", "ao3", "archive of our own",
    "fanfiction.net", "wattpad", "x reader", "x-reader", "reader insert",
    "y/n", "self-insert", "slash fic", "shipping", "alternate universe",
    "canon divergence", "what if", "one-shot", "drabble"
]

NONFICTION_SIGNALS = [
    "memoir", "biography", "autobiography", "history of", "essay", "essays",
    "guide to", "how to", "handbook", "textbook", "self-help", "nonfiction",
    "non-fiction", "cookbook", "manual", "reference"
]

# Phase-6 split #2: EPUB metadata + relationship/fandom canonicalization
# helpers live in ``utils/epub_metadata``.  Re-exported here so existing
# call sites (tags route imports extract_epub_metadata; fandoms route
# imports _canonicalize_fandom; exports/refresh/url_lists/upload_books in
# this file all reach for these symbols via ``routes.books``) keep
# working unchanged.
from utils.epub_metadata import (  # noqa: E402, F401
    SERIES_TITLE_PATTERNS,
    _FANDOM_SPLIT_RE,
    _canonicalize_fandom,
    _canonicalize_relationship,
    _suggest_fandom_merges,
    detect_series_from_title,
    extract_epub_metadata,
    extract_urls_from_epub,
    format_links_txt,
    update_epub_metadata,
)


_CHAPTER_NORMALIZE_RE = re.compile(r'\s+')
_CHAPTER_PREFIX_RE = re.compile(r'^\s*(?:chapter|ch\.?|part|prologue|epilogue)\s*[:\-\.]?\s*\d*[:\-\.]?\s*', re.IGNORECASE)


# Phase-6 split: the chapter helpers now live in ``utils/epub_chapters``.
# They're re-exported here so existing imports keep working unchanged
# (tests + diff route + refresh helper all do ``from routes.books import
# extract_chapters`` etc).
from utils.epub_chapters import (  # noqa: E402, F401
    _normalize_chapter_title,
    extract_chapters,
    diff_chapters,
)





# ============================================================
# FANFIC REFRESH — pull latest version of a fanfic from its source URL
# ============================================================
# URL canonicalization, source detection, and the per-host regex bank
# all live in `utils/url_canonical` — this module just re-exports them
# so existing call sites (and tests) keep working unchanged.
from utils.url_canonical import (  # noqa: E402
    _AO3_HOST_RE,
    _AO3_HOST_SUBSTRINGS,
    _AO3_NON_WORK_PATTERNS,
    _AO3_WORK_CANON_RE,
    _AFF_CANON_RE,
    _FFNET_CANON_RE,
    _FP_CANON_RE,
    _PS_CANON_RE,
    _QQ_CANON_RE,
    _RR_CANON_RE,
    _SB_CANON_RE,
    _SV_CANON_RE,
    _TWILIGHTED_CANON_RE,
    FANFIC_SOURCE_PATTERNS,
    _is_ao3_host,
    classify_ao3_non_work,
    normalize_fanfic_url,
)

FANFICFARE_USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64; rv:128.0) Gecko/20100101 Firefox/128.0"
)


# Phase-5 follow-up: url_lists.py owns the dedupe / pull / export-xlsx routes
# but `upload_books` and `claim_source_url` here still need these helpers, so
# re-import them at the books.py top-level (the dedicated import block that
# used to live ~line 1426 was moved out with the routes).
from utils.url_canonical import (  # noqa: E402
    _URL_RE,
    _canonical_fanfic_url,
    _looks_like_url_list,
    _source_for,  # test_new_features imports this via ``routes.books``
)

# Phase 5 cleanup: helpers that were extracted to other modules but are still
# referenced from this file (upload_books, list_library_xlsx, etc.).

async def _write_local_and_mirror_to_r2(
    local_path: Path,
    payload: bytes,
    user_id: str,
    book_id: str,
    ext: str,
) -> None:
    """Write ``payload`` to ``local_path`` AND immediately mirror it to
    R2 if cloud storage is enabled.

    Same defensive pattern as ``upload_books``'s bulk mirror — see
    2026-06-21 incident in CHANGELOG.md.  Local disk is the cache,
    R2 is the source of truth.  Pre-2026-06-21, the entire codebase
    wrote bytes to local disk and trusted the every-10-min storage
    backfill cron to push them to R2; pod restarts inside that 10-min
    window destroyed the bytes permanently (we lost 53 of 65 books to
    this on prod).

    Mirror failures here log a warning and rely on the cron retry —
    they MUST NOT raise back to the caller, because the bytes are
    already safely on local disk and the caller has no idempotent
    retry path.
    """
    local_path.write_bytes(payload)
    try:
        from utils.storage_cloud import (
            is_enabled as _cloud_on,
            mirror_up as _r2_mirror_up,
            storage_key_for as _r2_key,
        )
        if _cloud_on():
            await asyncio.to_thread(
                _r2_mirror_up, local_path, _r2_key(user_id, book_id, ext),
            )
    except Exception as e:  # noqa: BLE001
        logger.warning(
            "_write_local_and_mirror_to_r2: R2 mirror failed for %s "
            "(book=%s ext=%s): %s — file safe on local disk, cron will retry.",
            local_path.name, book_id, ext, e,
        )


def _safe_folder(name: str) -> str:
    """Mirror of routes/exports.py::_safe_folder — sanitised dir/file name."""
    import re as _re
    out = _re.sub(r"[^\w\-. ]+", "_", (name or "").strip())
    return out[:60] or "unknown"


async def _dedupe_url_list(text: str, user_id: str):
    """Lazy-import bridge so `upload_books` keeps working after url_lists.py
    moved out of this file. Imported on demand to dodge the circular."""
    from routes.url_lists import _dedupe_url_list as _impl  # noqa: WPS433
    return await _impl(text, user_id)


# ----------------------------------------------------------------------
# EPUB TEMPLATE APPLIER
# Implementation lives in `utils/epub_template`. Re-exported here so the
# call sites in this module (and the test suite) keep working unchanged.
# ----------------------------------------------------------------------
from utils.epub_template import (  # noqa: E402
    SHELFSORT_TEMPLATE_CSS,
    SHELFSORT_TEMPLATE_MARKER,
    _html_escape,
    _build_intro_xhtml,
    apply_template_to_epub,
)

from utils.status_detector import (  # noqa: E402
    detect_status,
    effective_status,
    COMPLETE as STATUS_COMPLETE,
    ONGOING as STATUS_ONGOING,
)
from utils.constants import TRASH_SHELF, TRASH_GRACE_DAYS  # noqa: E402



# ---- Tag helpers (moved to utils.tags as part of books.py refactor Phase 2)
# We still re-export the underscore-prefixed names here so any pending
# callers in this file (the upload + bulk-edit pipelines) keep working.
from utils.tags import (  # noqa: E402
    TAG_MAX_LENGTH,  # noqa: F401
    TAG_MAX_PER_BOOK,  # noqa: F401
    _normalize_tag,  # noqa: F401
    _normalize_tags,
)


class FanficNotFoundError(Exception):
    """FanFicFare couldn't fetch this fanfic — mark the book as unavailable."""
    pass


def find_source_url(links: List[Dict[str, str]]) -> Optional[str]:
    """Return the first URL in the list that points to a supported fanfic source,
    already normalized to its canonical form."""
    for item in links:
        url = (item.get('url') or '').strip()
        canon = normalize_fanfic_url(url)
        if canon:
            return canon
    return None


def extract_fanfic_urls(links: List[Dict[str, str]]) -> List[str]:
    """Return every canonical fanfic-permalink URL found in the EPUB's link set.

    We only keep URLs that match `FANFIC_SOURCE_PATTERNS` (AO3 /works/N, FFnet
    /s/N, RoyalRoad /fiction/N, etc.) so that duplicate detection doesn't trip
    on boilerplate navigation links shared by every AO3 EPUB. URLs are
    normalized (mobile host stripped, `www.` collapsed, AO3 collection prefix
    removed, chapter id dropped, http→https, etc.) so different surface forms
    of the same work dedupe correctly.
    """
    seen: set = set()
    out: List[str] = []
    for item in links or []:
        url = (item.get('url') or '').strip()
        canon = normalize_fanfic_url(url)
        if canon and canon not in seen:
            seen.add(canon)
            out.append(canon)
    return out


def _clean_author_string(raw: Optional[str]) -> str:
    """Tidy up messy author fields before storing.

    Handles common EPUB metadata patterns that make dedup + display worse:
      - 'by John Smith' → 'John Smith'
      - 'Smith, John & Doe, Jane' → 'Smith, John & Doe, Jane' (preserved,
        but trailing/leading separators stripped)
      - 'John Smith (a.k.a. Pseudonym)' → 'John Smith' (drop parenthetical)
      - 'Pseudonym [pen name]' → 'Pseudonym' (drop bracketed annotation)
      - 'anonymous', 'unknown author', '' → 'Unknown'
      - Collapse internal whitespace.

    We deliberately do NOT lowercase or reformat the case — only the
    matching helper does that, so display stays human-friendly.
    """
    s = (raw or "").strip()
    if not s:
        return "Unknown"
    # Drop parenthetical and bracketed annotations like "(pen name)" or "[a.k.a. X]"
    s = re.sub(r"\s*[\(\[][^)\]]*[\)\]]", "", s).strip()
    # Strip leading "by " (case-insensitive)
    s = re.sub(r"^(?:by|written by|author[:\s])\s+", "", s, flags=re.IGNORECASE).strip()
    # Trim stray separators ("John Smith, " or "& Jane")
    s = s.strip(" ,&;|/")
    # Collapse whitespace
    s = re.sub(r"\s+", " ", s)
    # Canonicalize common "unknown" sentinels
    low = s.lower()
    if low in ("anonymous", "anon", "anon.", "unknown", "unknown author", "n/a", "na", "various", "various authors"):
        return {"various": "Various", "various authors": "Various"}.get(low, "Unknown")
    return s


def _normalize_title_for_match(title: Optional[str]) -> str:
    return re.sub(r"\s+", " ", (title or "").strip()).lower()


def _normalize_author_for_match(author: Optional[str]) -> str:
    """Normalize for cross-row comparison: lowercase, drop dots, collapse
    whitespace, and merge runs of single-letter "initials" so 'J. K. Rowling'
    and 'JK Rowling' compare equal. Empty stays empty so callers can detect
    missing-author and fall back to title-only matching."""
    s = re.sub(r"\.", "", (author or "")).strip()
    s = re.sub(r"\s+", " ", s).lower()
    # Concatenate runs of single-letter words: 'j k rowling' → 'jk rowling'
    s = re.sub(
        r"\b([a-z])(\s+[a-z]\b)+",
        lambda m: m.group(0).replace(" ", ""),
        s,
    )
    return s


# NOTE: URL-list endpoints (dedupe / export-xlsx / pull) and their
# helpers (_backfill_user_fanfic_urls, _dedupe_url_list) were moved to
# routes/url_lists.py in the Phase 5 refactor (2026-06-14).



async def find_duplicate_candidates(
    user_id: str,
    *,
    title: Optional[str],
    author: Optional[str] = None,
    source_url: Optional[str],
    fanfic_urls: Optional[List[str]] = None,
    exclude_book_id: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Find existing books in the user's library that look like duplicates.

    Match rules (any of):
      - normalized title + author equality (case-insensitive, whitespace-collapsed,
        dots stripped from author). When either side has no author on file we
        fall back to title-only matching so books that legitimately lack an
        author still dedupe.
      - exact source_url equality
      - any shared canonical fanfic URL (intersection on `fanfic_urls`)

    Archived versions are searched too — when a match lands on an archived
    book we walk the `replaced_by` chain to its current head and surface the
    head as the match (with `historical_version` added to match_reasons),
    so the upload can be offered as a historical version of a current copy.

    Returns a list of `{book_id, title, author, match_reasons: [...]}` dicts.
    """
    norm_title = _normalize_title_for_match(title)
    norm_author = _normalize_author_for_match(author)
    urls = [u for u in (fanfic_urls or []) if u]

    or_clauses: List[Dict[str, Any]] = []
    if norm_title:
        # Narrow the title regex pre-filter; we still verify title+author
        # equality in Python below.
        escaped = re.escape(norm_title)
        or_clauses.append({"title": {"$regex": f"^\\s*{escaped}\\s*$", "$options": "i"}})
    if source_url:
        or_clauses.append({"source_url": source_url})
    if urls:
        or_clauses.append({"fanfic_urls": {"$in": urls}})

    if not or_clauses:
        return []

    query: Dict[str, Any] = {"user_id": user_id, "$or": or_clauses}
    if exclude_book_id:
        query["book_id"] = {"$ne": exclude_book_id}

    projection = {"_id": 0, "book_id": 1, "title": 1, "author": 1, "source_url": 1, "fanfic_urls": 1, "category": 1, "replaced_by": 1}
    matches_by_head: Dict[str, Dict[str, Any]] = {}

    async def _walk_to_head(doc: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Follow `replaced_by` until we hit a current (non-archived) copy."""
        current = doc
        seen: set = set()
        while current.get("replaced_by"):
            if current["book_id"] in seen:
                return None  # cycle guard
            seen.add(current["book_id"])
            nxt = await db.books.find_one(
                {"book_id": current["replaced_by"], "user_id": user_id},
                projection,
            )
            if not nxt:
                return None
            current = nxt
        if current.get("category") == OLD_STORIES_SHELF:
            return None  # orphaned archived chain
        return current

    async for doc in db.books.find(query, projection):
        is_archived = doc.get("category") == OLD_STORIES_SHELF or bool(doc.get("replaced_by"))
        head_doc = doc if not is_archived else await _walk_to_head(doc)
        if not head_doc:
            continue

        reasons: List[str] = []
        if norm_title and _normalize_title_for_match(doc.get("title")) == norm_title:
            # Tightened rule: when both sides have an author, they must
            # match too — otherwise two different books with the same title
            # (e.g. retellings, generic titles like "Untitled") get
            # falsely paired. Fall back to title-only when either side is
            # missing an author.
            doc_norm_author = _normalize_author_for_match(doc.get("author"))
            if not norm_author or not doc_norm_author:
                reasons.append("title")
            elif doc_norm_author == norm_author:
                reasons.append("title+author")
        if source_url and doc.get("source_url") == source_url:
            reasons.append("source_url")
        if urls:
            shared = [u for u in (doc.get("fanfic_urls") or []) if u in urls]
            if shared:
                reasons.append("url")
        if not reasons:
            continue
        if is_archived:
            reasons.append("historical_version")

        head_id = head_doc["book_id"]
        if head_id == exclude_book_id:
            continue
        existing = matches_by_head.get(head_id)
        if existing:
            # Merge reasons (de-duped)
            existing["match_reasons"] = sorted(set(existing["match_reasons"]) | set(reasons))
        else:
            matches_by_head[head_id] = {
                "book_id": head_id,
                "title": head_doc.get("title") or "",
                "author": head_doc.get("author") or "",
                "match_reasons": sorted(set(reasons)),
            }

    return list(matches_by_head.values())


async def _apply_duplicate_policy(
    user_id: str,
    new_book_id: str,
    target_book_id: Optional[str],
    policy: str,
) -> Optional[Dict[str, Any]]:
    """Apply a default-policy auto-resolution to a freshly-uploaded book.

    Returns a dict describing what was done, or None if the policy couldn't
    apply (e.g., no target). The expensive chapter-diff step from the
    interactive resolve flow is skipped for batch uploads — users running on
    a stand policy chose convenience over the bell badge.

    Side effect: every change is recorded under the book's `dupe_action_meta`
    field with the previous values so the action can be undone via
    `POST /api/books/{book_id}/undo-resolve`.
    """
    now_iso = datetime.now(timezone.utc).isoformat()

    if policy == "keep_both":
        await db.books.update_one(
            {"book_id": new_book_id, "user_id": user_id},
            {
                "$unset": {"duplicate_pending": "", "duplicate_of": ""},
                "$set": {"dupe_action_meta": {"action": "keep_both", "applied_at": now_iso}},
            },
        )
        return {"action": "keep_both", "undoable": False}

    if policy == "discard":
        # Soft-delete: move to Trash shelf with a 30-day grace window so the
        # user can restore. A background sweep hard-deletes books whose
        # `trash_expires_at` is in the past.
        new_doc_before = await db.books.find_one({"book_id": new_book_id, "user_id": user_id})
        if not new_doc_before:
            return None
        expires_at = (datetime.now(timezone.utc) + timedelta(days=TRASH_GRACE_DAYS)).isoformat()
        await db.books.update_one(
            {"book_id": new_book_id, "user_id": user_id},
            {
                "$set": {
                    "category": TRASH_SHELF,
                    "trash_expires_at": expires_at,
                    "dupe_action_meta": {
                        "action": "discard",
                        "prev_category_new": new_doc_before.get("category"),
                        "applied_at": now_iso,
                    },
                },
                "$unset": {"duplicate_pending": "", "duplicate_of": ""},
            },
        )
        return {"action": "discard", "undoable": True, "trash_expires_at": expires_at}

    # The remaining two need a current head; bail if there isn't one
    if not target_book_id:
        return None
    target = await db.books.find_one({"book_id": target_book_id, "user_id": user_id})
    if not target or target.get("category") == OLD_STORIES_SHELF or target.get("replaced_by"):
        return None

    new_doc_before = await db.books.find_one({"book_id": new_book_id, "user_id": user_id})
    if not new_doc_before:
        return None

    if policy == "historical":
        await db.books.update_one(
            {"book_id": new_book_id, "user_id": user_id},
            {
                "$set": {
                    "category": OLD_STORIES_SHELF,
                    "replaced_by": target_book_id,
                    "replaced_at": now_iso,
                    "dupe_action_meta": {
                        "action": "historical",
                        "target_book_id": target_book_id,
                        "prev_category_new": new_doc_before.get("category"),
                        "applied_at": now_iso,
                    },
                },
                "$unset": {"duplicate_pending": "", "duplicate_of": ""},
            },
        )
        return {
            "action": "historical",
            "target_book_id": target_book_id,
            "undoable": True,
        }

    if policy == "new_version":
        now_dt = datetime.now(timezone.utc)
        updated_shelf = _updated_shelf_name(now_dt)
        await db.books.update_one(
            {"book_id": new_book_id, "user_id": user_id},
            {
                "$set": {
                    "category": updated_shelf,
                    "replaces": target_book_id,
                    "last_refreshed_at": now_iso,
                    "update_seen": False,
                    "dupe_action_meta": {
                        "action": "new_version",
                        "target_book_id": target_book_id,
                        "prev_category_new": new_doc_before.get("category"),
                        "prev_category_target": target.get("category"),
                        "applied_at": now_iso,
                    },
                },
                "$unset": {"duplicate_pending": "", "duplicate_of": ""},
            },
        )
        await db.categories.update_one(
            {"user_id": user_id, "name": updated_shelf},
            {"$setOnInsert": {
                "user_id": user_id,
                "name": updated_shelf,
                "created_at": now_iso,
                "auto_created": True,
            }},
            upsert=True,
        )
        await db.books.update_one(
            {"book_id": target_book_id, "user_id": user_id},
            {"$set": {
                "category": OLD_STORIES_SHELF,
                "replaced_by": new_book_id,
                "replaced_at": now_iso,
            }},
        )
        return {
            "action": "new_version",
            "target_book_id": target_book_id,
            "updated_shelf": updated_shelf,
            "undoable": True,
        }

    return None


# NOTE: `POST /books/{book_id}/undo-resolve` was moved to routes/duplicates.py
# in the Phase 4 refactor (2026-06-14).


async def fanfic_fetch_epub(source_url: str, options: Optional[Dict[str, Any]] = None) -> tuple:
    """Generate an EPUB for the given fanfic URL using FanFicFare.

    Optional `options` dict (per-user FanFicFare prefs):
      - include_author_notes: bool (default True)
      - include_images: bool (default True)
      - keep_chapter_links: bool (default False)
    """
    loop = asyncio.get_event_loop()
    options = options or {}

    # Test hook: when set, returns canned content immediately so tests don't
    # need a real internet connection.
    canned = os.environ.get("SHELFSORT_TEST_FFF_RESPONSE")
    if canned:
        try:
            obj = json.loads(canned)
        except Exception:
            obj = {}
        if obj.get("not_found"):
            raise FanficNotFoundError(obj.get("detail", "Source unavailable"))
        # `epub_b64` is base64-encoded bytes; meta is a passthrough dict
        import base64
        epub_bytes = base64.b64decode(obj.get("epub_b64", ""))
        return epub_bytes, obj.get("meta") or {}

    def _do_download():
        from fanficfare import adapters
        from fanficfare.configurable import Configuration
        from fanficfare import exceptions as fff_exc
        from urllib.parse import urlparse
        host = urlparse(source_url).hostname or ""
        try:
            config = Configuration([host], "EPUB")
            # Use a realistic browser User-Agent — AO3 / FFN / Cloudflare
            # actively block obvious scraper UAs with HTTP 403.
            try:
                config.set("defaults", "user_agent", FANFICFARE_USER_AGENT)
                config.set(host, "user_agent", FANFICFARE_USER_AGENT)
            except Exception:
                # Not all FFF builds expose the same INI sections; fall through.
                pass
            # Apply per-user FanFicFare options. FFF expects strings for ini values.
            try:
                if "include_author_notes" in options:
                    val = "true" if options["include_author_notes"] else "false"
                    config.set("epub", "include_author_notes", val)
                if "include_images" in options:
                    val = "true" if options["include_images"] else "false"
                    config.set("epub", "include_images", val)
                if "keep_chapter_links" in options:
                    val = "true" if options["keep_chapter_links"] else "false"
                    config.set("epub", "keep_summary_html", val)
            except Exception as cfg_err:
                logger.warning("Failed to apply FFF user options: %s", cfg_err)
            adapter = adapters.getAdapter(config, source_url)
        except fff_exc.UnknownSite:
            raise FanficNotFoundError(f"This site isn't supported: {host}")
        except fff_exc.InvalidStoryURL as e:
            raise FanficNotFoundError(f"Invalid story URL: {e}")
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Adapter setup failed: {e}")

        try:
            adapter.getStoryMetadataOnly()
        except fff_exc.StoryDoesNotExist as e:
            # Heuristic: FFN's Cloudflare/anti-bot pages get parsed as
            # "story doesn't exist" because the real HTML isn't there. Give
            # the user a clearer hint when the site is FFN.
            if "fanfiction.net" in (host or "").lower():
                raise FanficNotFoundError(
                    "FanFiction.net's bot protection blocked the download. The work itself is "
                    "likely still online — try the 'Upload replacement' button on the book's page "
                    "to drop in a fresh EPUB you exported from your own browser/Calibre."
                )
            raise FanficNotFoundError(f"Story not found: {e}")
        except fff_exc.HTTPErrorFFF as e:
            msg = str(e)
            if "403" in msg:
                # 403 is frequently a transient rate-limit / Cloudflare challenge.
                # Wait briefly and try once more before flagging as unavailable.
                logger.info("403 from %s — backing off 30s and retrying once", host)
                import time as _time
                _time.sleep(30)
                try:
                    adapter.getStoryMetadataOnly()
                    # Retry succeeded — fall through to writeStory below
                except fff_exc.HTTPErrorFFF as e2:
                    if "403" in str(e2):
                        raise FanficNotFoundError(
                            "Source site blocked the request (HTTP 403, retried). The site may be rate-limiting, "
                            "behind a Cloudflare challenge, or restricting this work to registered users. "
                            "Try opening the URL in a browser to check."
                        )
                    raise FanficNotFoundError(f"Couldn't reach source after retry: {e2}")
                except Exception as e2:
                    raise FanficNotFoundError(f"Couldn't reach source after retry: {e2}")
            else:
                raise FanficNotFoundError(f"Couldn't reach source: {e}")
        except fff_exc.RegularDelayException as e:
            raise HTTPException(status_code=503, detail=f"Source rate-limited: {e}")
        except Exception as e:
            raise FanficNotFoundError(f"Source error: {e}")

        # Write EPUB into a temp file
        out_fd, out_path = tempfile.mkstemp(suffix=".epub")
        os.close(out_fd)
        try:
            from fanficfare import writers
            writer = writers.getWriter("epub", config, adapter)
            writer.writeStory(outfilename=out_path, forceOverwrite=True)
            with open(out_path, "rb") as f:
                epub_bytes = f.read()
        finally:
            try:
                os.unlink(out_path)
            except Exception:
                pass

        story = adapter.story
        # Capture every field we'll need to build the template-style intro page.
        meta = {
            "chapters": int(story.getMetadata("numChapters") or 0),
            "rawExtendedMeta": {
                "dateUpdated": story.getMetadata("dateUpdated"),
                "datePublished": story.getMetadata("datePublished"),
                "words": int(story.getMetadata("numWords") or 0) if story.getMetadata("numWords") else None,
                "status": story.getMetadata("status"),
                "rating": story.getMetadata("rating"),
                "language": story.getMetadata("language"),
                "reviews": story.getMetadata("reviews"),
                "favs": story.getMetadata("favs"),
                "follows": story.getMetadata("follows"),
                "genre": story.getMetadata("genre"),
                "category": story.getMetadata("category"),
            },
            "title": story.getMetadata("title"),
            "author": _clean_author_string(story.getMetadata("author")),
            "description": story.getMetadata("description"),
            "source_url": source_url,
            "site": host,
        }
        return epub_bytes, meta

    try:
        return await loop.run_in_executor(None, _do_download)
    except FanficNotFoundError:
        raise
    except HTTPException:
        raise
    except Exception as e:
        logger.error("FanFicFare download failed: %s", e)
        raise HTTPException(status_code=502, detail=f"Download error: {e}")


def _updated_shelf_name(now: Optional[datetime] = None) -> str:
    """Return the date-stamped 'Updated stories' shelf name for refreshes today.

    Each refresh batch gets its own dated bucket, so every run of updates is
    clearly separated. Example: "Updated stories 2026-03-01"."""
    now = now or datetime.now(timezone.utc)
    return f"Updated stories {now.strftime('%Y-%m-%d')}"


OLD_STORIES_SHELF = "Old stories"
# NOTE: ``TRASH_SHELF`` and ``TRASH_GRACE_DAYS`` are imported from
# ``utils.constants`` at the top of this module — kept centralized so
# ``routes/trash.py`` and ``routes/authors.py`` share the canonical values.


async def fetch_fanfic_with_fallback(
    source_url: str,
    options: Optional[Dict[str, Any]] = None,
) -> tuple:
    """Try FanFicFare first; if it fails AND the user opted into the
    FicHub fallback, retry with FicHub. Returns the same `(epub_bytes,
    source_meta)` tuple as `fanfic_fetch_epub`.

    The fallback is serialized — even if many user requests hit this in
    parallel, they're drained through `routes.fichub_client._FETCH_LOCK`
    one at a time, with a 2s gap between consecutive FicHub fetches.
    """
    # Feature-flag kill switch — admin can pause remote fic fetching.
    from utils.feature_flags import is_enabled
    if not await is_enabled("fichub_enabled"):
        raise FanficNotFoundError("Fanfic fetching is temporarily disabled by an administrator.")
    options = options or {}
    try:
        return await fanfic_fetch_epub(source_url, options=options)
    except FanficNotFoundError as fff_err:
        if not options.get("try_fichub_fallback"):
            raise
        from routes.fichub_client import (  # local import to avoid circular
            fichub_fetch_epub,
            FichubUnsupportedURL,
            FichubError,
        )
        try:
            epub_bytes, _meta = await fichub_fetch_epub(source_url)
            logger.info("FicHub fallback succeeded for %s", source_url)
            return epub_bytes, {"source": "fichub", "url": source_url}
        except FichubUnsupportedURL:
            # Re-raise the original FFF error — that's the more informative
            # message ("Story not found", "Site not supported", etc.).
            raise fff_err
        except FichubError as e:
            logger.warning(
                "FicHub fallback also failed for %s: %s", source_url, e
            )
            raise fff_err


async def apply_refresh(book: Dict[str, Any], user_id: str, source_url: str) -> Dict[str, Any]:
    """Refresh a fanfic by generating a new EPUB via FanFicFare.

    Behavior (2026-02, updated per user request): instead of overwriting the
    existing EPUB and book record, we create a NEW book in a date-stamped
    "Updated stories YYYY-MM-DD" shelf and move the original to the single
    "Old stories" shelf. Every refresh batch gets its own dated bucket so the
    history of updates stays clearly separated.

    Cross-links:
      - new book .replaces -> old book_id
      - old book .replaced_by -> new book_id
    """
    # Honor per-user FanFicFare options (incl. opt-in FicHub fallback)
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0, "fff_options": 1})
    fff_options = (user_doc or {}).get("fff_options") or {}
    epub_bytes, source_meta = await fetch_fanfic_with_fallback(source_url, options=fff_options)

    # Apply the FicHub-style template (intro page + stylesheet) unless the
    # user has explicitly opted out. Idempotent: noop on already-templated EPUBs.
    if fff_options.get("apply_template", True):
        loop = asyncio.get_event_loop()
        epub_bytes = await loop.run_in_executor(
            None, apply_template_to_epub, epub_bytes, source_meta, source_url
        )

    user_dir = STORAGE_DIR / user_id
    user_dir.mkdir(parents=True, exist_ok=True)

    # Generate a fresh book_id + path for the new copy
    new_book_id = f"book_{uuid.uuid4().hex[:12]}"
    new_epub_path = user_dir / f"{new_book_id}.epub"
    await _write_local_and_mirror_to_r2(
        new_epub_path, epub_bytes, user_id, new_book_id, ".epub",
    )

    new_meta = extract_epub_metadata(new_epub_path)
    new_cover_path = user_dir / f"{new_book_id}.cover"
    if new_meta.get("cover_bytes"):
        await _write_local_and_mirror_to_r2(
            new_cover_path, new_meta["cover_bytes"], user_id, new_book_id, ".cover",
        )

    links = extract_urls_from_epub(new_epub_path)
    (user_dir / f"{new_book_id}.links.txt").write_text(
        format_links_txt(new_meta["title"], new_meta["author"], links),
        encoding="utf-8",
    )

    now_iso = datetime.now(timezone.utc).isoformat()
    now_dt = datetime.now(timezone.utc)
    updated_shelf = _updated_shelf_name(now_dt)
    old_book_id = book["book_id"]

    # 1) Insert the new book in the date-stamped "Updated stories" shelf
    new_doc = {
        "book_id": new_book_id,
        "user_id": user_id,
        "filename": _templated_filename(new_meta.get("title"), new_meta.get("author"), new_book_id),
        "title": new_meta["title"],
        "author": new_meta["author"],
        "description": new_meta["description"],
        "language": new_meta["language"],
        "publisher": new_meta["publisher"],
        "has_cover": bool(new_meta.get("cover_bytes")),
        # Each refresh batch lives in its own dated bucket
        "category": updated_shelf,
        "fandom": book.get("fandom"),
        "series_name": book.get("series_name"),
        "series_index": book.get("series_index"),
        "tags": book.get("tags") or [],
        "confidence": book.get("confidence", 0.0),
        "classifier": book.get("classifier", "metadata"),
        "size_bytes": len(epub_bytes),
        "links_count": len(links),
        "source_url": source_url,
        "last_refreshed_at": now_iso,
        "source_meta": source_meta,
        "replaces": old_book_id,
        "created_at": now_iso,
    }
    await db.books.insert_one(new_doc)

    # Register the dated shelf as a custom category so it surfaces in the UI
    # chip list. Idempotent — same date is reused across a day's refreshes.
    await db.categories.update_one(
        {"user_id": user_id, "name": updated_shelf},
        {"$setOnInsert": {
            "user_id": user_id,
            "name": updated_shelf,
            "created_at": now_iso,
            "auto_created": True,
        }},
        upsert=True,
    )

    # 2) Move the old book to the "Old stories" shelf with a back-pointer
    await db.books.update_one(
        {"book_id": old_book_id, "user_id": user_id},
        {"$set": {
            "category": OLD_STORIES_SHELF,
            "replaced_by": new_book_id,
            "replaced_at": now_iso,
        }},
    )

    # 3) Compute a quick diff summary and stash it on the new book so the
    # "fics updated" navbar badge can query it cheaply (no per-poll EPUB
    # parsing). Failures here are non-fatal — the badge will just skip this
    # book. Always sets `update_seen=False` so the badge picks it up.
    refresh_summary: Optional[Dict[str, Any]] = None
    try:
        old_epub_path = user_dir / f"{old_book_id}.epub"
        if old_epub_path.exists():
            loop = asyncio.get_event_loop()
            old_chapters = await loop.run_in_executor(None, extract_chapters, old_epub_path)
            new_chapters = await loop.run_in_executor(None, extract_chapters, new_epub_path)
            d = diff_chapters(old_chapters, new_chapters)
            refresh_summary = {
                "chapters_added": d["summary"]["chapters_added"],
                "chapters_changed": d["summary"]["chapters_changed"],
                "chapters_removed": d["summary"]["chapters_removed"],
                "words_delta": d["summary"]["words_delta"],
                "first_changed_href": (d.get("first_changed_chapter") or {}).get("new_href", ""),
                "first_changed_title": (d.get("first_changed_chapter") or {}).get("title", ""),
                "first_changed_kind": (d.get("first_changed_chapter") or {}).get("kind", ""),
            }
    except Exception as e:
        logger.warning("refresh_summary diff failed for %s -> %s: %s", old_book_id, new_book_id, e)

    await db.books.update_one(
        {"book_id": new_book_id, "user_id": user_id},
        {"$set": {
            "refresh_summary": refresh_summary,
            "update_seen": False,
        }},
    )

    return {
        "new_book_id": new_book_id,
        "old_book_id": old_book_id,
        "title": new_meta["title"],
        "author": new_meta["author"],
        "last_refreshed_at": now_iso,
        "updated_shelf": updated_shelf,
    }


# Phase-6 split #3: classifier helpers (heuristic + Claude) now live in
# ``utils/classifier``.  Re-exported so existing imports (admin route
# imports classify_by_metadata; tags route imports classify_with_ai;
# tests reach for both via routes.books) keep working unchanged.
from utils.classifier import (  # noqa: E402, F401
    classify_book,
    classify_by_metadata,
    classify_with_ai,
)


# ============================================================
# BOOK ROUTES

NEEDS_CONVERSION_EXTS = {
    ".pdf", ".mobi", ".azw", ".azw3", ".kf8", ".kfx",
    ".docx", ".doc", ".rtf", ".fb2", ".lit", ".lrf", ".pdb", ".txt", ".html", ".htm",
}
NEEDS_CONVERSION_SHELF = "Needs conversion"


def _convert_to_epub_sync(src_path: Path, dest_path: Path) -> Optional[str]:
    """Run `ebook-convert <src> <dest>` synchronously. Returns None on success,
    or an error message on failure. Called from an executor so the FastAPI
    event loop stays responsive."""
    import subprocess
    try:
        proc = subprocess.run(
            ["ebook-convert", str(src_path), str(dest_path)],
            capture_output=True,
            text=True,
            timeout=180,  # 3 min cap per book — heavy PDFs can be slow
        )
        if proc.returncode != 0:
            tail = (proc.stderr or proc.stdout or "")[-400:]
            return f"ebook-convert failed (rc={proc.returncode}): {tail.strip()}"
        if not dest_path.exists() or dest_path.stat().st_size < 256:
            return "ebook-convert produced no usable output"
        return None
    except FileNotFoundError:
        return "ebook-convert is not installed on the server"
    except subprocess.TimeoutExpired:
        return "ebook-convert timed out (>3 min)"
    except Exception as e:
        return f"ebook-convert crashed: {e}"


async def convert_to_epub(src_path: Path, dest_path: Path) -> Optional[str]:
    # Feature-flag kill switch — admin can pause Calibre conversions.
    from utils.feature_flags import is_enabled
    if not await is_enabled("calibre_convert_enabled"):
        return "Calibre conversion is temporarily disabled by an administrator."
    # 2026-06-21 — Cap concurrent Calibre invocations to 2 to avoid OOM
    # on the production Launch tier (2 GiB pod limit).  ebook-convert
    # holds ~200-400 MB transiently per conversion, and clamd already
    # claims ~960 MB persistently — Emergent Support flagged in the
    # ClamAV bake-in thread that "concurrent ebook conversions could
    # push the pod past the 2 GiB limit. Cap concurrent conversions to
    # 1 or 2 in your app code."  Chose 2 (not 1) so a single user with
    # a queue doesn't fully block a second user from converting at the
    # same time, while still staying inside the headroom budget.
    sem = _get_calibre_semaphore()
    async with sem:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, _convert_to_epub_sync, src_path, dest_path)


# Lazy semaphore init — asyncio.Semaphore must be created inside a
# running event loop, otherwise it captures the wrong loop and raises
# "Future attached to a different loop" under heavy traffic.
_calibre_sem: Optional[asyncio.Semaphore] = None


def _get_calibre_semaphore() -> asyncio.Semaphore:
    global _calibre_sem
    if _calibre_sem is None:
        _calibre_sem = asyncio.Semaphore(2)
    return _calibre_sem


# Persistent conversion-job tracking — backed by MongoDB so jobs survive
# backend restarts, tab closes, and cross-device sessions. A TTL index on
# `expires_at` cleans up finished jobs after the 4-hour visibility window.
CONVERSION_VISIBILITY_HOURS = 4
_conversion_index_ensured = False


async def _ensure_conversion_index() -> None:
    """Lazily create a TTL index on conversion_jobs.expires_at."""
    global _conversion_index_ensured
    if _conversion_index_ensured:
        return
    try:
        await db.conversion_jobs.create_index("expires_at", expireAfterSeconds=0)
        await db.conversion_jobs.create_index([("user_id", 1), ("started_at", -1)])
        _conversion_index_ensured = True
    except Exception as e:
        logger.warning("Failed to create conversion_jobs indexes: %s", e)


async def _conversion_start(user_id: str, job: Dict[str, Any]) -> None:
    await _ensure_conversion_index()
    doc = {
        **job,
        "user_id": user_id,
        "status": "processing",
        # expires_at intentionally omitted so the TTL doesn't apply while
        # the job is still running.
    }
    await db.conversion_jobs.insert_one(doc)


async def _conversion_end(user_id: str, job_id: str, *, error: Optional[str] = None) -> None:
    now = datetime.now(timezone.utc)
    expires = now + timedelta(hours=CONVERSION_VISIBILITY_HOURS)
    await db.conversion_jobs.update_one(
        {"id": job_id, "user_id": user_id},
        {
            "$set": {
                "status": "failed" if error else "done",
                "error": error,
                "finished_at": now.isoformat(),
                "expires_at": expires,
            }
        },
    )


# /conversions/* and /library/originals/* endpoints live in
# ``routes/conversions.py`` (extracted 2026-06-13). The helpers above
# (``convert_to_epub``, ``_conversion_start``, ``_conversion_end``,
# ``_ensure_conversion_index``, and the visibility-window constant) stay
# here because ``upload_books`` below also uses them.


@api_router.post("/books/upload")
async def upload_books(
    request: Request,
    files: List[UploadFile] = File(...),
    keep_originals: List[str] = Form([]),
    user: User = Depends(get_current_user),
):
    # Feature-flag kill switch — admin can pause uploads in maintenance.
    from utils.feature_flags import is_enabled
    if not await is_enabled("uploads_enabled"):
        raise HTTPException(status_code=503, detail="Uploads are temporarily disabled by an administrator.")
    user_dir = STORAGE_DIR / user.user_id
    user_dir.mkdir(parents=True, exist_ok=True)
    # Load fandom aliases once for the whole batch so per-book canonicalization
    # picks up user-defined merges (e.g. "HP" -> "Harry Potter"). Global
    # admin-managed aliases are merged in; per-user overrides on conflict.
    _udoc = await db.users.find_one(
        {"user_id": user.user_id}, {"_id": 0, "fandom_aliases": 1}
    ) or {}
    user_aliases = _udoc.get("fandom_aliases") or {}
    from routes.admin import get_global_fandom_aliases_dict
    global_aliases = await get_global_fandom_aliases_dict()
    fandom_aliases = {**global_aliases, **user_aliases}
    # Filenames the user explicitly asked to keep as the original format
    # (no Calibre conversion). They land on /library/originals separately
    # from the main EPUB library.
    keep_original_set = {n for n in keep_originals if n}
    results = []
    url_list_reports: List[Dict[str, Any]] = []
    upload_suggestions: List[Dict[str, Any]] = []
    cross_format_dupes: List[Dict[str, Any]] = []
    # Story-shaped URLs we found inside uploaded EPUBs whose host isn't on
    # the accepted-sources list. Collected across every file in the batch
    # and flushed to the `unknown_sources` collection just before the
    # response so the toast can echo back the new hosts.
    upload_unknown_urls: List[Dict[str, Any]] = []  # {url, book_id, title, author}

    for f in files:
        lower = (f.filename or "").lower()
        ext = "." + lower.rsplit(".", 1)[-1] if "." in lower else ""

        # ---- Antivirus pre-scan (2026-06-18) --------------------------
        # Synchronous ClamAV scan before ANY other processing — anything
        # the scanner flags is rejected with a hard 400 and recorded in
        # the ``av_quarantine`` collection for admin review.  Read once,
        # rewind, then proceed; downstream branches all re-read ``f`` or
        # use the buffered bytes.  Failing to run the scanner (daemon
        # not up, sig DB missing) returns ok=False and we fail-open so
        # uploads keep working — the missing-AV state surfaces in the
        # admin Health card instead of breaking user uploads.
        from utils.antivirus import scan_bytes, record_quarantine
        import asyncio as _asyncio
        _av_bytes = await f.read()
        await f.seek(0)
        _av_result = await _asyncio.to_thread(scan_bytes, _av_bytes, hint_name=(f.filename or "upload.bin"))
        if _av_result.get("infected"):
            await record_quarantine(
                user_id=user.user_id,
                filename=f.filename or "",
                scan=_av_result,
                source="upload",
                extra={"size_bytes": len(_av_bytes)},
            )
            raise HTTPException(
                status_code=400,
                detail=f"\"{f.filename or 'this file'}\" appears unsafe ({_av_result.get('signature') or 'flagged by antivirus'}). Upload blocked.",
            )
        # -------------------------------------------------------------

        # `.txt` is a special case — it could be a plain-text manuscript
        # (Calibre-convertible) OR a wishlist of fanfic URLs. If it's
        # dominantly URLs we route it through the dedupe pipeline instead of
        # converting it as a book.
        if ext == ".txt":
            try:
                raw_bytes = await f.read()
                text = raw_bytes.decode("utf-8", errors="ignore")
            except Exception:
                text, raw_bytes = "", b""
            looks_like_url_list = _looks_like_url_list(text)
            if looks_like_url_list:
                report = await _dedupe_url_list(text, user.user_id)
                report["filename"] = f.filename
                url_list_reports.append(report)
                continue
            # Not a URL list — restore the read pointer so the standard
            # Calibre-convert branch below picks it up. We re-write the file
            # to disk and skip ahead.
            await f.seek(0)

        # Non-EPUB but a known ebook format → auto-convert to EPUB via
        # Calibre's `ebook-convert`, then fall through to the normal EPUB
        # pipeline below (metadata / classification / fanfic / template).
        # On conversion failure we keep the original file under the
        # "Needs conversion" shelf with a friendly error message.
        original_format: Optional[str] = None
        if ext != ".epub" and ext in NEEDS_CONVERSION_EXTS:
            book_id = f"book_{uuid.uuid4().hex[:12]}"
            src_target = user_dir / f"{book_id}{ext}"
            content = await f.read()
            src_target.write_bytes(content)

            # Path 1 — "Keep original": user wants this file on the Originals
            # shelf without Calibre conversion. We do a quick title/author
            # guess from the filename (and cross-format dup check against
            # existing EPUBs) and store an original-only doc.
            if (f.filename or "") in keep_original_set:
                base_name = (f.filename or "Untitled").rsplit(".", 1)[0]
                # Title - Author pattern, common from manual exports
                guess_title = base_name
                guess_author = "Unknown"
                if " - " in base_name:
                    left, right = base_name.rsplit(" - ", 1)
                    if len(left) > 1 and len(right) > 1:
                        guess_title, guess_author = left.strip(), right.strip()
                # Cross-format duplicate detection — match title+author
                # case-insensitively against existing EPUB books.
                dup_match = await db.books.find_one(
                    {
                        "user_id": user.user_id,
                        "original_only": {"$ne": True},
                        "title": {"$regex": f"^{re.escape(guess_title)}$", "$options": "i"},
                        "author": {"$regex": f"^{re.escape(guess_author)}$", "$options": "i"},
                    },
                    {"_id": 0, "book_id": 1, "title": 1, "author": 1},
                )
                dup_ids = [dup_match["book_id"]] if dup_match else []
                if dup_match:
                    cross_format_dupes.append({
                        "new_filename": f.filename,
                        "new_book_id": book_id,
                        "matched_book_id": dup_match["book_id"],
                        "matched_title": dup_match.get("title"),
                        "matched_author": dup_match.get("author"),
                    })
                now_iso = datetime.now(timezone.utc).isoformat()
                doc = {
                    "book_id": book_id,
                    "user_id": user.user_id,
                    "filename": f.filename,
                    "title": guess_title,
                    "author": guess_author,
                    "description": f"Original {ext.lstrip('.').upper()} kept as-is (no Calibre conversion).",
                    "language": "",
                    "publisher": "",
                    "has_cover": False,
                    # Use a distinct shelf so these don't pollute the main library.
                    "category": "Originals",
                    "fandom": None,
                    "confidence": 1.0,
                    "classifier": "kept-original",
                    "tags": [],
                    "size_bytes": len(content),
                    "links_count": 0,
                    "source_url": None,
                    "fanfic_urls": [],
                    "last_refreshed_at": None,
                    "series_name": None,
                    "series_index": None,
                    "original_only": True,
                    "original_format": ext.lstrip("."),
                    "cross_format_duplicate_of": dup_ids,
                    "created_at": now_iso,
                }
                await db.books.insert_one(doc)
                results.append({k: v for k, v in doc.items() if k != "_id"})
                continue

            # Path 2 — normal "Convert" flow (existing behavior).
            epub_target = user_dir / f"{book_id}.epub"
            job_id = uuid.uuid4().hex
            await _conversion_start(user.user_id, {
                "id": job_id,
                "book_id": book_id,
                "title": (f.filename or "Untitled").rsplit(".", 1)[0],
                "original_format": ext.lstrip("."),
                "started_at": datetime.now(timezone.utc).isoformat(),
            })
            err = None
            try:
                err = await convert_to_epub(src_target, epub_target)
            finally:
                await _conversion_end(user.user_id, job_id, error=err)
            if err:
                base_name = (f.filename or "Untitled").rsplit(".", 1)[0]
                now_iso = datetime.now(timezone.utc).isoformat()
                doc = {
                    "book_id": book_id,
                    "user_id": user.user_id,
                    "filename": f.filename,
                    "title": base_name,
                    "author": "Unknown",
                    "description": (
                        f"Uploaded as .{ext.lstrip('.')} but auto-conversion failed: {err}. "
                        f"Convert it manually with Calibre's 'Convert books' tool and re-upload."
                    ),
                    "language": "",
                    "publisher": "",
                    "has_cover": False,
                    "category": NEEDS_CONVERSION_SHELF,
                    "fandom": None,
                    "confidence": 1.0,
                    "classifier": "needs-conversion",
                    "size_bytes": len(content),
                    "links_count": 0,
                    "source_url": None,
                    "last_refreshed_at": None,
                    "series_name": None,
                    "series_index": None,
                    "needs_conversion": True,
                    "original_format": ext.lstrip("."),
                    "conversion_error": err,
                    "created_at": now_iso,
                }
                await db.books.insert_one(doc)
                results.append({k: v for k, v in doc.items() if k != "_id"})
                continue
            # Conversion succeeded — keep the original file too (so the user
            # has the source) but route the rest of the pipeline at the EPUB.
            original_format = ext.lstrip(".")
            content = epub_target.read_bytes()
            target = epub_target
            # Fall through to the standard EPUB processing below using the
            # already-written EPUB. We jump straight to metadata extraction by
            # reusing the local `book_id` we generated above.
        elif ext != ".epub":
            results.append({"filename": f.filename, "error": "Not an EPUB"})
            continue
        else:
            book_id = f"book_{uuid.uuid4().hex[:12]}"
            target = user_dir / f"{book_id}.epub"
            content = await f.read()
            target.write_bytes(content)

        meta = extract_epub_metadata(target)

        # Short-circuit: if the EPUB can't be opened at all, file it under
        # "Can't Open" and skip classification / AI / links / series detection.
        if meta.get("parse_failed"):
            doc = {
                "book_id": book_id,
                "user_id": user.user_id,
                "filename": f.filename,
                "title": meta.get("title") or f.filename,
                "author": "Unknown",
                "description": "",
                "language": "",
                "publisher": "",
                "has_cover": False,
                "category": "Can't Open",
                "fandom": None,
                "confidence": 1.0,
                "classifier": "broken-epub",
                "size_bytes": len(content),
                "links_count": 0,
                "source_url": None,
                "last_refreshed_at": None,
                "series_name": None,
                "series_index": None,
                "epub_unreadable": True,
                "epub_parse_error": meta.get("parse_error"),
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.books.insert_one(doc)
            results.append({k: v for k, v in doc.items() if k != "_id"})
            continue

        classification = await classify_book(meta)

        # Save cover separately if exists
        cover_path = user_dir / f"{book_id}.cover"
        if meta.get('cover_bytes'):
            cover_path.write_bytes(meta['cover_bytes'])

        # Extract URLs and save to a notepad-friendly .txt file
        links = extract_urls_from_epub(target)
        links_path = user_dir / f"{book_id}.links.txt"
        links_path.write_text(
            format_links_txt(meta['title'], meta['author'], links),
            encoding='utf-8',
        )
        source_url = find_source_url(links)
        fanfic_urls = extract_fanfic_urls(links)

        # Stash URLs that look story-shaped but didn't canonicalize so we
        # can record their hosts as "potential new sources" after the
        # batch finishes (one Mongo write per host, not per URL).
        for _link in links or []:
            _u = (_link.get("url") or "").strip()
            if _u and not normalize_fanfic_url(_u):
                upload_unknown_urls.append({
                    "url": _u, "book_id": book_id,
                    "title": meta.get("title"), "author": meta.get("author"),
                })

        # Series detection: prefer EPUB Calibre meta, fall back to title regex
        series_name = meta.get('series_name')
        series_index = meta.get('series_index')
        if not series_name:
            sn, si = detect_series_from_title(meta['title'])
            if sn:
                series_name = sn
                series_index = si if si is not None else series_index

        doc = {
            "book_id": book_id,
            "user_id": user.user_id,
            "filename": f.filename,
            "title": meta['title'],
            "author": meta['author'],
            "description": meta['description'],
            "language": meta['language'],
            "publisher": meta['publisher'],
            "has_cover": bool(meta.get('cover_bytes')),
            "category": classification['category'],
            "fandom": _canonicalize_fandom(classification.get('fandom'), fandom_aliases),
            "confidence": classification.get('confidence'),
            "classifier": classification.get('classifier'),
            "size_bytes": len(content),
            "links_count": len(links),
            "source_url": source_url,
            "fanfic_urls": fanfic_urls,
            "last_refreshed_at": None,
            "series_name": series_name,
            "series_index": series_index,
            "relationships": meta.get("relationships") or [],
            "rating": meta.get("rating"),
            "warnings": meta.get("warnings") or [],
            "categories": meta.get("categories") or [],
            "ao3_freeform_tags": meta.get("ao3_freeform_tags") or [],
            # Auto-detected completion status (complete | ongoing). User
            # override lives at `manual_status`; effective_status() picks
            # the override when set. Detection runs only at upload time —
            # users said they don't want re-detection on refresh (5a).
            "status": detect_status(
                title=meta.get("title"),
                description=meta.get("description"),
                raw_meta_text=meta.get("rawExtendedMeta_text"),
                tags=meta.get("tags") or [],
            ),
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        if original_format:
            # Surface the source format so the UI can show e.g. "Converted from PDF"
            doc["original_format"] = original_format
            doc["converted_from"] = original_format

        # Duplicate detection — flag, don't block. The UI pops a modal letting
        # the user choose: keep both / discard this upload / promote as new
        # version of the existing book.
        dupes = await find_duplicate_candidates(
            user.user_id,
            title=meta['title'],
            author=meta.get('author'),
            source_url=source_url,
            fanfic_urls=fanfic_urls,
        )
        if dupes:
            doc["duplicate_pending"] = True
            doc["duplicate_of"] = dupes

        await db.books.insert_one(doc)
        # Hook in full-text index — extract the EPUB body so the new book
        # is searchable from `/library/search/fulltext` immediately. Any
        # failure here is logged inside the helper; we never want a
        # fulltext glitch to break the upload itself, so we swallow.
        try:
            from utils.epub_fulltext import extract_epub_text, upsert_fulltext, count_words  # noqa: WPS433
            from pathlib import Path as _P  # noqa: WPS433
            # Reconstruct the on-disk path from STORAGE_DIR + user_id + book_id.
            _epub_path = STORAGE_DIR / user.user_id / f"{doc['book_id']}.epub"
            _ft_text = extract_epub_text(_epub_path)
            await upsert_fulltext(db, doc["book_id"], user.user_id, _ft_text)
            _wc = count_words(_ft_text)
            if _wc > 0:
                await db.books.update_one(
                    {"book_id": doc["book_id"]},
                    {"$set": {"word_count": _wc}},
                )
                doc["word_count"] = _wc
        except Exception as _ft_exc:
            logger.warning("fulltext index on upload failed for %s: %s", doc.get("book_id"), _ft_exc)
        results.append({k: v for k, v in doc.items() if k != '_id'})

    # Auto-resolve based on the user's default duplicate policy. When the
    # policy is "ask" we leave duplicate_pending on every flagged book so the
    # UI pops the modal. For other policies we apply the action immediately.
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "duplicate_policy": 1})
    from routes.user_prefs import DUPE_POLICY_DEFAULT  # extracted module
    policy = (user_doc or {}).get("duplicate_policy") or DUPE_POLICY_DEFAULT
    auto_resolved = 0
    actions: List[Dict[str, Any]] = []
    if policy != "ask":
        for i, doc in enumerate(results):
            if not doc.get("duplicate_pending"):
                continue
            target_id = (doc.get("duplicate_of") or [{}])[0].get("book_id")
            applied = await _apply_duplicate_policy(
                user.user_id, doc["book_id"], target_id, policy,
            )
            if applied:
                auto_resolved += 1
                actions.append({
                    "book_id": doc["book_id"],
                    "title": doc.get("title") or "",
                    "action": applied.get("action"),
                    "target_book_id": applied.get("target_book_id"),
                    "undoable": applied.get("undoable", False),
                })
                # Reflect the auto-resolve in the response so the UI knows
                if applied.get("deleted"):
                    results[i] = {**doc, "duplicate_pending": False, "duplicate_resolved": "discard", "removed": True}
                else:
                    fresh = await db.books.find_one({"book_id": doc["book_id"], "user_id": user.user_id})
                    if fresh:
                        fresh.pop("_id", None)
                        fresh["duplicate_resolved"] = applied.get("action")
                        results[i] = fresh

    # Fuzzy match suggestions — look at every fandom that landed in this
    # batch; if it's a brand-new fandom and close (≤2 edits) to an existing
    # one, surface a suggestion the UI can pop as a toast.
    batch_fandoms = {b.get("fandom") for b in results if isinstance(b, dict) and b.get("fandom")}
    if batch_fandoms:
        existing_rows = await db.books.aggregate([
            {"$match": {"user_id": user.user_id, "fandom": {"$ne": None, "$exists": True}}},
            {"$group": {"_id": "$fandom"}},
        ]).to_list(5000)
        existing_fandoms = [r["_id"] for r in existing_rows if r.get("_id")]
        # Only suggest when the just-uploaded fandom is rare in the library
        # (otherwise it's clearly already an "established" shelf).
        counts: Dict[str, int] = {}
        for r in existing_rows:
            counts[r["_id"]] = counts.get(r["_id"], 0) + 1
        for nf in batch_fandoms:
            sug = _suggest_fandom_merges(nf, [e for e in existing_fandoms if e != nf])
            if sug:
                upload_suggestions.append({"new_fandom": nf, "suggestions": sug})

    # Unknown-source detector: flush all story-shaped URLs that didn't
    # canonicalize as a single Mongo upsert per distinct host. We record
    # the most recently-seen sample per host along with the book title/
    # author/id so the admin endpoint can show context.
    from utils.unknown_sources import record_unknown_sources
    unknown_hosts_recorded: List[str] = []
    if upload_unknown_urls:
        # Group by host so we attach the latest book context to each host.
        from utils.unknown_sources import _host_of, looks_like_fanfic_url
        seen_hosts: set = set()
        for item in upload_unknown_urls:
            u = item["url"]
            if not looks_like_fanfic_url(u):
                continue
            h = _host_of(u)
            if not h or h in seen_hosts:
                continue
            seen_hosts.add(h)
            rec = await record_unknown_sources(
                db, [u], context="upload",
                user_id=user.user_id,
                book_id=item.get("book_id"),
                book_title=item.get("title"),
                book_author=item.get("author"),
            )
            unknown_hosts_recorded.extend(rec)

    # Best-effort: notify friends who already collect any of the same
    # fandoms in this batch. Never raises — see helper for rules.
    await _notify_friends_of_shared_fandom_uploads(
        user.user_id,
        (user.name or user.email or "A friend"),
        results,
    )

    # 2026-06-21 — Synchronously mirror every freshly-created file to
    # R2 before returning success.  Pre-fix, ``upload_books`` wrote to
    # local disk only and relied on the every-10-min storage backfill
    # cron to push to R2.  If the pod restarted in that window (idle
    # scale-down, deploy, OOM kill), the bytes were lost FOREVER —
    # which is exactly the regression the user hit on 2026-06-21
    # when their /admin/storage-migration-progress showed only 18% of
    # books actually in R2.  This loop closes the gap by mirroring
    # every book + cover synchronously before we tell the user "upload
    # succeeded".  Each mirror is best-effort — if R2 is briefly
    # unreachable we log a warning and continue (the cron will retry
    # within 10 min), but we DON'T silently swallow it.
    from utils.storage_cloud import (
        is_enabled as _cloud_on,
        mirror_up as _r2_mirror_up,
        storage_key_for as _r2_key,
    )
    if _cloud_on() and results:
        import asyncio as _asyncio
        for r in results:
            bid = r.get("book_id")
            if not bid:
                continue
            # Mirror the EPUB (always), the original-format source if
            # this was a converted upload, and the cover (if any).
            mirror_targets: List[tuple[Path, str]] = []
            epub_local = user_dir / f"{bid}.epub"
            if epub_local.exists():
                mirror_targets.append((epub_local, _r2_key(user.user_id, bid, ".epub")))
            orig_fmt = (r.get("original_format") or "").strip()
            if orig_fmt:
                src_ext = f".{orig_fmt}"
                src_local = user_dir / f"{bid}{src_ext}"
                if src_local.exists():
                    mirror_targets.append((src_local, _r2_key(user.user_id, bid, src_ext)))
            cover_local = user_dir / f"{bid}.cover"
            if cover_local.exists():
                mirror_targets.append((cover_local, _r2_key(user.user_id, bid, ".cover")))
            for local_path, key in mirror_targets:
                try:
                    ok = await _asyncio.to_thread(_r2_mirror_up, local_path, key)
                    if not ok:
                        logger.warning(
                            "upload_books: R2 mirror returned False for %s (key=%s) — "
                            "file safe on local disk, cron will retry.",
                            local_path.name, key,
                        )
                except Exception as e:  # noqa: BLE001
                    logger.warning(
                        "upload_books: R2 mirror raised for %s (key=%s): %s — "
                        "file safe on local disk, cron will retry.",
                        local_path.name, key, e,
                    )

    return {
        "uploaded": len(results),
        "books": results,
        "auto_resolved": auto_resolved,
        "policy": policy,
        "actions": actions,
        "url_lists": url_list_reports,
        "fandom_suggestions": upload_suggestions,
        "cross_format_duplicates": cross_format_dupes,
        "unknown_sources_found": unknown_hosts_recorded,
    }


async def _notify_friends_of_shared_fandom_uploads(
    uploader_id: str,
    uploader_display: str,
    uploaded_results: List[Dict[str, Any]],
) -> None:
    """When a user uploads fanfic in fandoms their friends also collect,
    drop one in-app notification per (friend, fandom) so the friend can
    peek at the new arrival. Best-effort only — failures are logged and
    swallowed so an upload never 500s on a notification hiccup.

    Rules:
      • Only books with a `fandom` value count (skips non-fic / original fic).
      • Books that were removed by an auto-resolve "discard" policy are
        skipped (`removed: True`).
      • One notification per (friend, fandom) per batch — not per book.
      • Hard cap of 50 notifications per upload to prevent runaway spam.
    """
    from routes.notifications import create_notification
    try:
        # 1) Distinct fandoms in this batch that we'd want to ping about.
        batch_fandoms: set = set()
        for b in uploaded_results or []:
            if not isinstance(b, dict):
                continue
            if b.get("removed"):
                continue
            fd = b.get("fandom")
            if fd and isinstance(fd, str) and fd.strip():
                batch_fandoms.add(fd.strip())
        if not batch_fandoms:
            return

        # 2) Accepted friends only.
        friend_rows = await db.friendships.find(
            {
                "status": "accepted",
                "$or": [{"user_a": uploader_id}, {"user_b": uploader_id}],
            },
            {"_id": 0, "user_a": 1, "user_b": 1},
        ).to_list(length=2000)
        friend_ids = [
            (r["user_b"] if r["user_a"] == uploader_id else r["user_a"])
            for r in friend_rows
        ]
        if not friend_ids:
            return

        # 3) For each friend, find which of the batch fandoms they also have.
        emitted = 0
        cap = 50
        for fid in friend_ids:
            if emitted >= cap:
                break
            rows = await db.books.find(
                {"user_id": fid, "fandom": {"$in": list(batch_fandoms)}},
                {"_id": 0, "fandom": 1},
            ).to_list(length=500)
            shared = sorted({r["fandom"] for r in rows if r.get("fandom")})
            for fandom in shared:
                if emitted >= cap:
                    break
                await create_notification(
                    fid,
                    kind="friend_new_book",
                    title=f"{uploader_display} just added a new {fandom} fic",
                    body="Peek their shelf to see what's new.",
                    link="/friends",
                )
                emitted += 1
    except Exception as e:  # pragma: no cover — defensive
        logger.warning(f"friend-fandom notifications skipped: {e}")


# NOTE: `GET /library/trends` was moved to routes/library_views.py
# in the Phase 5E refactor (2026-06-14).



# /library/originals* endpoints and the convert_original_to_epub helper
# live in ``routes/conversions.py`` (extracted 2026-06-13).



@api_router.get("/books")
async def list_books(
    request: Request,
    category: Optional[str] = None,
    fandom: Optional[str] = None,
    relationship: Optional[str] = None,
    q: Optional[str] = None,
    smart: Optional[str] = None,
    include_originals: bool = False,
    rating: Optional[str] = None,
    ao3_category: Optional[str] = None,
    warning: Optional[str] = None,
    exclude_warning: Optional[str] = None,
    user: User = Depends(get_current_user),
):
    query: Dict[str, Any] = {"user_id": user.user_id}
    if category:
        query['category'] = category
    else:
        # Trash is opt-in — only show when the user explicitly asks for it
        query['category'] = {"$ne": TRASH_SHELF}
    if fandom:
        query['fandom'] = fandom
    if relationship:
        query['relationships'] = relationship
    # AO3 metadata filters (added 2026-06-13). Each is exact-match on a
    # canonical value (e.g. "Mature", "M/M", "Graphic Depictions Of Violence").
    if rating:
        query['rating'] = rating
    if ao3_category:
        query['categories'] = ao3_category
    if warning:
        query['warnings'] = warning
    if exclude_warning:
        # "Hide books warned for X" — content-safety filter. Returns books
        # whose ``warnings`` array does NOT contain the given value.
        query.setdefault('warnings', {})
        if isinstance(query['warnings'], dict):
            query['warnings']['$ne'] = exclude_warning
        else:
            # warning was also set — combine into $and so both apply.
            query['$and'] = query.get('$and', []) + [
                {'warnings': query['warnings']},
                {'warnings': {'$ne': exclude_warning}},
            ]
            del query['warnings']
    # Originals (kept-as-is non-EPUBs) live on /library/originals — exclude
    # them from the main library unless explicitly asked.
    if not include_originals and not (category == "Originals"):
        query['original_only'] = {"$ne": True}

    or_clauses: List[List[Dict[str, Any]]] = []
    if q:
        or_clauses.append([
            {"title": {"$regex": q, "$options": "i"}},
            {"author": {"$regex": q, "$options": "i"}},
        ])

    if smart == "reading":
        query['progress_fraction'] = {"$gte": 0.05, "$lt": 0.95}
    elif smart == "finished":
        query['progress_fraction'] = {"$gte": 0.99}
    elif smart == "unavailable":
        query['unavailable'] = True
    elif smart == "unread":
        or_clauses.append([
            {"progress_fraction": {"$exists": False}},
            {"progress_fraction": None},
            {"progress_fraction": {"$lt": 0.05}},
        ])

    if len(or_clauses) == 1:
        query["$or"] = or_clauses[0]
    elif len(or_clauses) > 1:
        query["$and"] = [{"$or": clauses} for clauses in or_clauses]

    books = await db.books.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return {"books": books}


@api_router.get("/books/stats")
async def book_stats(user: User = Depends(get_current_user)):
    pipeline_cat = [
        {"$match": {"user_id": user.user_id}},
        {"$group": {"_id": "$category", "count": {"$sum": 1}}},
    ]
    pipeline_fandom = [
        {"$match": {"user_id": user.user_id, "fandom": {"$ne": None}}},
        {"$group": {"_id": "$fandom", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    pipeline_rel = [
        {"$match": {"user_id": user.user_id, "relationships": {"$exists": True, "$ne": []}}},
        {"$unwind": "$relationships"},
        {"$group": {"_id": "$relationships", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    cats = await db.books.aggregate(pipeline_cat).to_list(100)
    fandoms = await db.books.aggregate(pipeline_fandom).to_list(100)
    relationships = await db.books.aggregate(pipeline_rel).to_list(200)
    total = await db.books.count_documents({"user_id": user.user_id})
    reading = await db.books.count_documents({
        "user_id": user.user_id,
        "progress_fraction": {"$gte": 0.05, "$lt": 0.95},
    })
    finished = await db.books.count_documents({
        "user_id": user.user_id,
        "progress_fraction": {"$gte": 0.99},
    })
    unreadable = await db.books.count_documents({
        "user_id": user.user_id,
        "epub_unreadable": True,
    })
    return {
        "total": total,
        "reading": reading,
        "finished": finished,
        "unreadable": unreadable,
        "categories": [{"name": c['_id'], "count": c['count']} for c in cats],
        "fandoms": [{"name": f['_id'], "count": f['count']} for f in fandoms],
        "relationships": [{"name": r['_id'], "count": r['count']} for r in relationships],
        "crossover_count": sum(
            1 for f in fandoms
            if f.get('_id') and len([p for p in str(f['_id']).split(' / ') if p.strip()]) >= 2
        ),
    }


def _suggest_search_url(source_url: Optional[str], title: str, author: str) -> Optional[str]:
    """Build a 'find it again' search URL on the same site as the dead source."""
    from urllib.parse import quote_plus
    q = quote_plus(f"{title or ''} {author or ''}".strip())
    if not q:
        return None
    host = (source_url or "").lower()
    if "archiveofourown.org" in host:
        return f"https://archiveofourown.org/works/search?work_search%5Bquery%5D={q}"
    if "fanfiction.net" in host:
        return f"https://www.fanfiction.net/search/?keywords={q}&type=story"
    if "fictionpress.com" in host:
        return f"https://www.fictionpress.com/search/?keywords={q}&type=story"
    if "royalroad.com" in host:
        return f"https://www.royalroad.com/fictions/search?title={q}"
    if "spacebattles.com" in host or "sufficientvelocity.com" in host or "questionablequesting.com" in host:
        base = host.split("/")[2] if "://" in host else host
        return f"https://www.google.com/search?q=site%3A{base}+{q}"
    # Generic fallback: Google
    return f"https://www.google.com/search?q={q}"


@api_router.get("/books/export/unavailable")
async def export_unavailable_list(user: User = Depends(get_current_user)):
    """A plain .txt list of every book FanFicFare couldn't find — for manual lookup."""
    books = await db.books.find(
        {"user_id": user.user_id, "unavailable": True},
        {"_id": 0},
    ).sort("title", 1).to_list(5000)

    lines: List[str] = []
    lines.append("Shelfsort — books we couldn't fetch online")
    lines.append(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    lines.append(f"Count: {len(books)}")
    lines.append("=" * 70)
    lines.append("")
    if not books:
        lines.append("(none — every refreshable book updated successfully)")
    else:
        for i, b in enumerate(books, 1):
            shelf = b.get("category") or "Uncategorized"
            if shelf == "Fanfiction" and b.get("fandom"):
                shelf = f"Fanfiction / {b['fandom']}"
            lines.append(f"{i}. {b.get('title') or '(untitled)'}")
            lines.append(f"   Author:      {b.get('author') or 'Unknown'}")
            lines.append(f"   Shelf:       {shelf}")
            if b.get("source_url"):
                lines.append(f"   Source URL:  {b['source_url']}")
            if b.get("last_fetch_error"):
                lines.append(f"   Source said: {b['last_fetch_error']}")
            if b.get("last_fetch_attempt_at"):
                lines.append(f"   Last tried:  {b['last_fetch_attempt_at']}")
            search = _suggest_search_url(
                b.get("source_url"), b.get("title", ""), b.get("author", "")
            )
            if search:
                lines.append(f"   How to fix:  {search}")
            lines.append("")
    body = "\n".join(lines) + "\n"
    headers = {"Content-Disposition": "attachment; filename=shelfsort_cant_find_online.txt"}
    return Response(content=body, media_type="text/plain; charset=utf-8", headers=headers)


# NOTE: `GET /books/refresh-status` was moved to routes/refresh.py
# in the Phase 4 refactor (2026-06-14).


@api_router.get("/books/recent")
async def list_recent(limit: int = 8, user: User = Depends(get_current_user)):
    """Recently-opened books for the dashboard's Continue Reading rail.

    For each book we also attach the latest ``reading_cursors`` entry
    (`last_device_id`, `last_device_label`, `last_cursor_updated_at`)
    so the rail can render a "📱 Last on iPhone · 2h ago" caption when
    the most recent reading happened on a different device than the
    one currently viewing the dashboard.  This turns the cross-device
    sync from invisible plumbing into a visible delight.
    """
    cursor = db.books.find(
        {"user_id": user.user_id, "last_opened_at": {"$ne": None, "$exists": True}},
        {"_id": 0},
    ).sort("last_opened_at", -1).limit(max(1, min(int(limit), 24)))
    books = await cursor.to_list(24)

    if not books:
        return {"books": books}

    # Side-fetch the latest cursor per book in a single Mongo round trip.
    book_ids = [b["book_id"] for b in books]
    cursor_rows = await db.reading_cursors.find(
        {"user_id": user.user_id, "book_id": {"$in": book_ids}},
        {"_id": 0, "book_id": 1, "device_id": 1, "device_label": 1, "updated_at": 1},
    ).to_list(length=len(book_ids))
    cursor_by_book = {c["book_id"]: c for c in cursor_rows}
    for b in books:
        c = cursor_by_book.get(b["book_id"])
        if c:
            b["last_device_id"]         = c.get("device_id")
            b["last_device_label"]      = c.get("device_label") or ""
            ts = c.get("updated_at")
            if isinstance(ts, datetime):
                ts = ts.isoformat()
            b["last_cursor_updated_at"] = ts
    return {"books": books}


@api_router.get("/books/recent-updates")
async def recent_updates(limit: int = 8, user: User = Depends(get_current_user)):
    """Fanfics that have been refreshed and haven't been marked as seen.
    Powers the "fics updated" navbar bell badge."""
    limit = max(1, min(int(limit), 24))
    cursor = db.books.find(
        {
            "user_id": user.user_id,
            "replaces": {"$ne": None, "$exists": True},
            "update_seen": {"$ne": True},
        },
        {
            "_id": 0,
            "book_id": 1,
            "title": 1,
            "author": 1,
            "fandom": 1,
            "category": 1,
            "last_refreshed_at": 1,
            "replaces": 1,
            "refresh_summary": 1,
            "has_cover": 1,
        },
    ).sort("last_refreshed_at", -1).limit(limit)
    items = await cursor.to_list(limit)
    # Total unseen (so the badge can say "8+" if there are more)
    total_unseen = await db.books.count_documents({
        "user_id": user.user_id,
        "replaces": {"$ne": None, "$exists": True},
        "update_seen": {"$ne": True},
    })
    return {"updates": items, "total_unseen": total_unseen}


@api_router.post("/books/{book_id}/mark-update-seen")
async def mark_update_seen(book_id: str, user: User = Depends(get_current_user)):
    """Mark a single refreshed book as seen — removes it from the bell badge."""
    result = await db.books.update_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"$set": {"update_seen": True, "update_seen_at": datetime.now(timezone.utc).isoformat()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


@api_router.post("/books/mark-updates-seen")
async def mark_all_updates_seen(user: User = Depends(get_current_user)):
    """Mark every pending refreshed book as seen — clears the bell badge."""
    now_iso = datetime.now(timezone.utc).isoformat()
    result = await db.books.update_many(
        {
            "user_id": user.user_id,
            "replaces": {"$ne": None, "$exists": True},
            "update_seen": {"$ne": True},
        },
        {"$set": {"update_seen": True, "update_seen_at": now_iso}},
    )
    return {"ok": True, "marked": result.modified_count}



# FFF options, dashboard layout, format prefs, duplicate policy,
# onboarding sweeps, and fandom aliases live in routes/user_prefs.py
# (extracted 2026-06-13). The helpers ``apply_template_to_epub`` and
# ``_templated_filename`` stay in this file and are imported lazily.











@api_router.get("/books/{book_id}")
async def get_book(book_id: str, user: User = Depends(get_current_user)):
    book = await db.books.find_one({"book_id": book_id, "user_id": user.user_id}, {"_id": 0})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")
    # Cross-device awareness: attach the latest reading_cursor's device
    # info so the BookDetail page can render a "Last read on iPhone ·
    # 42% · 2h ago" hint next to the Read button.  Same payload shape
    # as /books/recent, kept consistent so the same FE helpers work.
    c = await db.reading_cursors.find_one(
        {"user_id": user.user_id, "book_id": book_id},
        {"_id": 0, "device_id": 1, "device_label": 1, "updated_at": 1, "percent": 1},
    )
    if c:
        book["last_device_id"]    = c.get("device_id")
        book["last_device_label"] = c.get("device_label") or ""
        ts = c.get("updated_at")
        if isinstance(ts, datetime):
            ts = ts.isoformat()
        book["last_cursor_updated_at"] = ts
        # If the cursor has a percent and the book doc doesn't track
        # one yet, fall back to cursor's percent for the UI.
        if c.get("percent") is not None and book.get("progress_fraction") is None:
            book["last_cursor_percent"] = c.get("percent")
    return book


@api_router.get("/books/{book_id}/reading-stats")
async def book_reading_stats(book_id: str, user: User = Depends(get_current_user)):
    """Per-book reading stats for the book-detail page.

    Returns:
      - reading_minutes: total time spent in this book (from heartbeats)
      - session_count: distinct days this book was opened/read
      - first_opened_at: ISO date of the first reading_activity row with this book
      - last_opened_at: from book document
      - sparkline: last 30 days, binary { date, active } per day
    """
    from datetime import date as _date, timedelta as _td

    book = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "book_id": 1, "reading_minutes": 1, "last_opened_at": 1, "created_at": 1, "progress_fraction": 1},
    )
    if not book:
        raise HTTPException(status_code=404, detail="Not found")

    activity = await db.reading_activity.find(
        {"user_id": user.user_id, "book_ids": book_id},
        {"_id": 0, "date": 1, "book_minutes": 1},
    ).sort("date", 1).to_list(2000)
    dates: List[str] = [a["date"] for a in activity if a.get("date")]
    # Map date -> minutes spent on THIS book that day. Older activity rows
    # (before per-book tracking landed) lack `book_minutes`; treat as 0.
    minutes_by_date: Dict[str, float] = {}
    for a in activity:
        bm = a.get("book_minutes") or {}
        minutes_by_date[a["date"]] = float(bm.get(book_id, 0))

    today = datetime.now(timezone.utc).date()
    cutoff = today - _td(days=29)
    date_set = set(dates)
    sparkline: List[Dict[str, Any]] = []
    # Find the day's max minutes (within the window) so the UI can normalize
    # bar heights without a second pass.
    window_minutes: List[float] = [
        minutes_by_date.get((cutoff + _td(days=i)).isoformat(), 0) for i in range(30)
    ]
    max_minutes = max(window_minutes) if window_minutes else 0
    for i in range(30):
        d = cutoff + _td(days=i)
        key = d.isoformat()
        mins = minutes_by_date.get(key, 0)
        sparkline.append({
            "date": key,
            "active": key in date_set,
            "minutes": int(mins),
        })

    # Reading-pace estimate: time-to-finish based on minutes-per-progress so far.
    # Only show when there's enough signal to avoid wild extrapolations:
    #   * at least 5 minutes of tracked reading (otherwise per-progress is noisy)
    #   * progress between 5% and 99% (else division explodes or book is done)
    reading_minutes = int(book.get("reading_minutes") or 0)
    progress = float(book.get("progress_fraction") or 0)
    estimated_minutes_left: Optional[int] = None
    if reading_minutes >= 5 and 0.05 <= progress < 0.99:
        try:
            estimated_minutes_left = max(0, int(round(
                (reading_minutes / progress) * (1 - progress)
            )))
            # Sanity cap at 1 week of reading (10080 min) — clamps wild outliers
            estimated_minutes_left = min(estimated_minutes_left, 10080)
        except (ZeroDivisionError, ValueError):
            estimated_minutes_left = None

    return {
        "book_id": book_id,
        "reading_minutes": reading_minutes,
        "session_count": len(dates),
        "first_opened_at": dates[0] if dates else None,
        "last_opened_at": book.get("last_opened_at"),
        "sparkline": sparkline,
        "sparkline_max_minutes": int(max_minutes),
        "progress_fraction": progress,
        "estimated_minutes_left": estimated_minutes_left,
    }





@api_router.get("/books/{book_id}/cover")
async def get_cover(book_id: str, request: Request):
    # Allow token in query for img src
    token = request.query_params.get('t')
    user_id = None
    if token:
        sess = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
        if sess:
            user_id = sess['user_id']
    if not user_id:
        try:
            user = await get_current_user(request)
            user_id = user.user_id
        except HTTPException:
            raise HTTPException(status_code=401, detail="Not authenticated")
    book = await db.books.find_one({"book_id": book_id, "user_id": user_id}, {"_id": 0})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")
    cover = STORAGE_DIR / user_id / f"{book_id}.cover"
    # Object-storage fallback: after a redeploy the local cache may be
    # empty even though the bytes are safely mirrored in the cloud.
    if not cover.exists():
        from utils.storage_cloud import ensure_local_cached
        if not await asyncio.to_thread(ensure_local_cached, cover, user_id, book_id, ".cover"):
            raise HTTPException(status_code=404, detail="No cover")
    return FileResponse(str(cover), media_type="image/jpeg")


@api_router.get("/books/{book_id}/download")
async def download_book(book_id: str, user: User = Depends(get_current_user)):
    book = await db.books.find_one({"book_id": book_id, "user_id": user.user_id}, {"_id": 0})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")
    fp = STORAGE_DIR / user.user_id / f"{book_id}.epub"
    if not fp.exists():
        from utils.storage_cloud import ensure_local_cached
        if not await asyncio.to_thread(ensure_local_cached, fp, user.user_id, book_id, ".epub"):
            raise HTTPException(status_code=404, detail="File missing")
    # AV pre-flight on download — catches the "old file, new signature"
    # scenario where a book that was clean at upload time gets flagged
    # by an updated ClamAV signature DB.  Cheap when clamd is running
    # (~10-50 ms).  Cache the result on the book doc so we don't rescan
    # the same file on every subsequent download.
    if book.get("av_status") != "clean":
        from utils.antivirus import scan_path, record_quarantine
        _av = await asyncio.to_thread(scan_path, fp)
        if _av.get("infected"):
            await record_quarantine(
                user_id=user.user_id,
                filename=f"{book_id}.epub",
                scan=_av,
                source="restore",
                extra={"book_id": book_id},
            )
            await db.books.update_one(
                {"book_id": book_id, "user_id": user.user_id},
                {"$set": {
                    "av_status": "infected",
                    "av_signature": _av.get("signature", ""),
                    "av_scanned_at": datetime.now(timezone.utc).isoformat(),
                }},
            )
            raise HTTPException(
                status_code=403,
                detail=f"This file was flagged by antivirus ({_av.get('signature') or 'unknown signature'}) and can no longer be downloaded.",
            )
        if _av.get("ok"):
            await db.books.update_one(
                {"book_id": book_id, "user_id": user.user_id},
                {"$set": {
                    "av_status": "clean",
                    "av_scanned_at": datetime.now(timezone.utc).isoformat(),
                }},
            )
    download_name = _templated_filename(book.get('title'), book.get('author'), book_id)
    return FileResponse(str(fp), media_type="application/epub+zip", filename=download_name)


# ---------------------------------------------------------------------
# Send to Kindle (2026-06-22) — emails the book's EPUB to the user's
# Amazon Kindle inbox.  Heavy lifting (file read, AV check, rate
# limit, Resend attachment send, email_logs write) lives in
# ``utils/send_to_kindle.py``; this route is a thin wrapper that
# enforces the auth context.
# ---------------------------------------------------------------------
@api_router.post("/books/{book_id}/send-to-kindle")
async def send_book_to_kindle_route(
    book_id: str,
    user: User = Depends(get_current_user),
):
    """Email the book's EPUB to the user's configured Kindle address.

    Gated on the ``send_to_kindle_enabled`` feature flag (default OFF,
    2026-06-22 — Resend quota brake).  When the flag is off, the
    endpoint returns 503 so the frontend can render a clear toast
    instead of users blaming the email service.

    Returns ``{"ok": true, "resend_id": "...", "to": "...", "size_bytes": int}``
    on success.  Common failure responses (caller surfaces toast text
    straight from the ``detail`` field):

    * 400 — no Kindle email configured / invalid format
    * 403 — book quarantined by antivirus
    * 404 — book or file missing
    * 413 — file > Kindle 25 MB gateway cap
    * 429 — same book already sent within the past 30 min
    * 502 — Resend rejected the send (quota, recipient bounce, etc.)
    * 503 — feature flag off or no Resend key configured
    """
    from utils.feature_flags import is_enabled  # noqa: WPS433
    if not await is_enabled("send_to_kindle_enabled"):
        from fastapi import HTTPException as _HE
        raise _HE(
            status_code=503,
            detail="Send-to-Kindle is currently disabled by the operator.",
        )
    from utils.send_to_kindle import send_book_to_kindle  # noqa: WPS433
    return await send_book_to_kindle(user_id=user.user_id, book_id=book_id)


@api_router.delete("/books/{book_id}")
async def delete_book(book_id: str, user: User = Depends(get_current_user)):
    book = await db.books.find_one({"book_id": book_id, "user_id": user.user_id}, {"_id": 0})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")
    await db.books.delete_one({"book_id": book_id, "user_id": user.user_id})
    for ext in ['.epub', '.cover', '.links.txt']:
        p = STORAGE_DIR / user.user_id / f"{book_id}{ext}"
        if p.exists():
            p.unlink()
    return {"ok": True}


def _safe_filename(name: str, ext: str) -> str:
    # Strip path separators / control chars
    base = re.sub(r'[\\/:*?"<>|\x00-\x1f]', '_', name or 'book').strip().rstrip('.')
    base = base[:120] or 'book'
    return f"{base}{ext}"


def _templated_filename(title: Optional[str], author: Optional[str], book_id: str, ext: str = ".epub") -> str:
    """Build a filename matching the attachment template: 'Title_by_Author-id.epub'.
    Underscores replace spaces, control + filesystem-unsafe chars are stripped,
    and a short 8-char book_id suffix disambiguates same-name fics."""
    def _clean(s: str) -> str:
        s = re.sub(r'[\\/:*?"<>|\x00-\x1f]', '', s or '')
        s = re.sub(r'\s+', '_', s.strip())
        return s.strip('._') or ''
    title_part = _clean(title or 'Untitled')[:80]
    author_part = _clean(author or 'Unknown')[:50]
    # Take the trailing 8 chars of the book_id for a stable, short, unique suffix
    short_id = (book_id or '').split('_')[-1][:8] or 'x'
    return f"{title_part}_by_{author_part}-{short_id}{ext}"


@api_router.get("/books/export/links")
async def export_all_links(
    category: Optional[List[str]] = Query(None),
    fandom: Optional[List[str]] = Query(None),
    relationship: Optional[List[str]] = Query(None),
    author: Optional[List[str]] = Query(None),
    format: str = "txt",
    user: User = Depends(get_current_user),
):
    """Download every URL across the user's library (or a filter).

    `format=txt` (default): one combined .txt file.
    `format=zip`: a .zip with one .txt per fandom (grouped by like fanfiction).
    `format=xlsx`: a single .xlsx workbook with one sheet per fandom, each
        row containing the book's full metadata + extracted URL count.
    """
    query: Dict[str, Any] = {"user_id": user.user_id}
    if category:
        query["category"] = {"$in": category} if len(category) > 1 else category[0]
    if fandom:
        query["fandom"] = {"$in": fandom} if len(fandom) > 1 else fandom[0]
    if relationship:
        query["relationships"] = {"$in": relationship} if len(relationship) > 1 else relationship[0]
    if author:
        query["author"] = {"$in": author} if len(author) > 1 else author[0]
    books = await db.books.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)
    if not books:
        raise HTTPException(status_code=404, detail="No books")

    user_dir = STORAGE_DIR / user.user_id

    # XLSX format — single workbook, one sheet per fandom, full metadata per row
    if format == "xlsx":
        import io as _io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        # Group books by fandom (or category for non-fanfic)
        buckets: Dict[str, List[Dict[str, Any]]] = {}
        for b in books:
            cat = b.get('category') or 'Uncategorized'
            bucket = b.get('fandom') if cat == 'Fanfiction' and b.get('fandom') else cat
            buckets.setdefault(bucket, []).append(b)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="6B46C1")
        header_align = Alignment(horizontal="left", vertical="center")
        columns = [
            ("Filename", "filename", 32),
            ("Title", "title", 36),
            ("Author", "author", 22),
            ("Fandom", "fandom", 22),
            ("Rating", "rating", 14),
            ("Categories", "categories", 16),
            ("Archive Warnings", "warnings", 26),
            ("Relationships", "relationships", 30),
            ("AO3 Tags", "ao3_freeform_tags", 28),
            ("User Tags", "tags", 22),
            ("Source URL", "source_url", 60),
        ]

        # Summary sheet first
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = "Shelfsort library export"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        ws_summary["A3"] = f"Books total: {len(books)}"
        ws_summary["A4"] = f"Fandoms / categories: {len(buckets)}"
        ws_summary["A6"] = "Fandom / Category"
        ws_summary["B6"] = "Books"
        ws_summary["A6"].font = header_font
        ws_summary["B6"].font = header_font
        ws_summary["A6"].fill = header_fill
        ws_summary["B6"].fill = header_fill
        for i, (bk, lst) in enumerate(sorted(buckets.items()), start=7):
            ws_summary[f"A{i}"] = bk
            ws_summary[f"B{i}"] = len(lst)
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 10

        def _sheet_name(name: str) -> str:
            # Excel limits: ≤31 chars, no : \ / ? * [ ]
            cleaned = re.sub(r'[:\\/?*\[\]]', '-', name)[:31] or "Sheet"
            return cleaned

        used_names: set = {"Summary"}
        for bucket_name, bucket_books in sorted(buckets.items()):
            base = _sheet_name(bucket_name)
            name = base
            suffix = 2
            while name in used_names:
                name = (base[:28] + f"_{suffix}")[:31]
                suffix += 1
            used_names.add(name)
            ws = wb.create_sheet(title=name)
            # Header row
            for col_idx, (label, _key, width) in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=label)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                ws.column_dimensions[cell.column_letter].width = width
            ws.freeze_panes = "A2"
            # Data rows
            for r_idx, b in enumerate(bucket_books, start=2):
                for c_idx, (label, key, _w) in enumerate(columns, start=1):
                    raw = b.get(key)
                    if isinstance(raw, list):
                        # List-valued columns (warnings, categories, etc.)
                        # render as a comma-joined string so Excel filters
                        # work on a single column.
                        value = ", ".join(str(x) for x in raw if x)
                    else:
                        value = raw or ""
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws.auto_filter.ref = ws.dimensions

        buf = _io.BytesIO()
        wb.save(buf)
        payload = buf.getvalue()
        xlsx_name = "shelfsort_library.xlsx"
        if fandom and len(fandom) == 1:
            xlsx_name = f"shelfsort_{_safe_folder(fandom[0])}.xlsx"
        elif category and len(category) == 1:
            xlsx_name = f"shelfsort_{_safe_folder(category[0])}.xlsx"
        elif any([fandom, category, relationship, author]):
            xlsx_name = "shelfsort_filtered.xlsx"
        return Response(
            content=payload,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{xlsx_name}"',
                "Content-Length": str(len(payload)),
            },
        )

    # ZIP format — one .txt per fandom (or category for non-fanfiction)
    if format == "zip":
        import io as _io
        # Group books by their bucket — fanfiction sub-groups by fandom,
        # everything else groups by category.
        buckets: Dict[str, List[Dict[str, Any]]] = {}
        for b in books:
            category_val = b.get('category') or 'Uncategorized'
            if category_val == 'Fanfiction':
                bucket = b.get('fandom') or 'Unsorted Fanfiction'
            else:
                bucket = category_val
            buckets.setdefault(bucket, []).append(b)

        buf = _io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            now_str = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')
            summary_lines = [
                "Shelfsort — links grouped by fandom",
                f"Generated: {now_str}",
                f"Books: {len(books)}",
                f"Fandoms / shelves: {len(buckets)}",
                "",
                "Each .txt file groups every fanfic from one fandom (or",
                "category, for non-fanfiction books). Stories are separated",
                "by blank lines so you can scan a whole fandom at a glance.",
                "",
            ]
            zf.writestr("README.txt", "\n".join(summary_lines))

            for bucket_name, bucket_books in sorted(buckets.items()):
                bucket_lines: List[str] = []
                bucket_lines.append(f"=== {bucket_name} ===")
                bucket_lines.append(
                    f"{len(bucket_books)} book{'s' if len(bucket_books) != 1 else ''} · generated {now_str}"
                )
                bucket_lines.append("")
                bucket_total = 0
                for b in bucket_books:
                    epub_path = user_dir / f"{b['book_id']}.epub"
                    bucket_lines.append(
                        f"{b.get('title','Untitled')} — {b.get('author','Unknown')}"
                    )
                    if not epub_path.exists():
                        bucket_lines.append("  (EPUB missing on disk)")
                        bucket_lines.append("")
                        continue
                    links = extract_urls_from_epub(epub_path)
                    bucket_total += len(links)
                    if not links:
                        bucket_lines.append("  (no URLs)")
                    else:
                        for item in links:
                            anchor = item.get('anchor')
                            if anchor:
                                bucket_lines.append(f"  {item['url']}  —  {anchor}")
                            else:
                                bucket_lines.append(f"  {item['url']}")
                    bucket_lines.append("")
                bucket_lines.insert(2, f"Total URLs: {bucket_total}")
                arcname = f"{_safe_folder(bucket_name)}.txt"
                zf.writestr(arcname, "\n".join(bucket_lines) + "\n")

        payload = buf.getvalue()
        zip_name = "shelfsort_links_by_fandom.zip"
        if fandom and len(fandom) == 1:
            zip_name = f"shelfsort_{_safe_folder(fandom[0])}_links.zip"
        elif category and len(category) == 1:
            zip_name = f"shelfsort_{_safe_folder(category[0])}_links.zip"
        elif any([fandom, category, relationship, author]):
            zip_name = "shelfsort_filtered_links.zip"
        return Response(
            content=payload,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{zip_name}"',
                "Content-Length": str(len(payload)),
            },
        )

    # TXT format — combined single file (default, backward-compatible)
    scope = "your library"
    if fandom and len(fandom) == 1:
        scope = f"the {fandom[0]} shelf"
    elif category and len(category) == 1:
        scope = f"the {category[0]} shelf"
    elif any([fandom, category, relationship, author]):
        scope = "the filtered selection"

    lines: List[str] = []
    lines.append(f"Shelfsort — links extracted from {scope}")
    lines.append(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    lines.append(f"Books scanned: {len(books)}")
    lines.append("=" * 70)
    lines.append("")

    total_links = 0
    for b in books:
        epub_path = user_dir / f"{b['book_id']}.epub"
        if not epub_path.exists():
            continue
        links = extract_urls_from_epub(epub_path)
        total_links += len(links)

        shelf = b.get('category') or 'Uncategorized'
        if shelf == 'Fanfiction' and b.get('fandom'):
            shelf = f"Fanfiction / {b['fandom']}"

        lines.append(f"[{shelf}] {b.get('title','')} — {b.get('author','')}")
        if not links:
            lines.append("  (no URLs)")
        else:
            for item in links:
                if item.get('anchor'):
                    lines.append(f"  {item['url']}  —  {item['anchor']}")
                else:
                    lines.append(f"  {item['url']}")
        lines.append("")

    lines.insert(3, f"Total URLs:    {total_links}")
    body = "\n".join(lines) + "\n"

    fname = "shelfsort_all_links.txt"
    if fandom and len(fandom) == 1:
        fname = f"shelfsort_{_safe_folder(fandom[0])}_links.txt"
    elif category and len(category) == 1:
        fname = f"shelfsort_{_safe_folder(category[0])}_links.txt"
    elif any([fandom, category, relationship, author]):
        fname = "shelfsort_filtered_links.txt"
    headers = {"Content-Disposition": f"attachment; filename={fname}"}
    return Response(content=body, media_type="text/plain; charset=utf-8", headers=headers)


@api_router.get("/books/quick-search")
async def quick_search_books(q: str, limit: int = 8, user: User = Depends(get_current_user)):
    """Lightweight title/author typeahead — feeds the navbar quick-search dropdown.

    Case-insensitive *substring* match against title + author (NOT full-body
    text — see `/library/search/fulltext` for the heavier search). Excludes
    trashed / replaced books.  Returns minimal fields so the dropdown stays
    snappy.
    """
    needle = (q or "").strip()
    if len(needle) < 2:
        return {"books": []}
    limit = max(1, min(limit, 20))
    safe = re.escape(needle)
    cursor = db.books.find(
        {
            "user_id": user.user_id,
            "category": {"$ne": "Trash"},
            "replaced_by": {"$exists": False},
            "$or": [
                {"title": {"$regex": safe, "$options": "i"}},
                {"author": {"$regex": safe, "$options": "i"}},
            ],
        },
        {"_id": 0, "book_id": 1, "title": 1, "author": 1, "category": 1, "fandom": 1},
    ).sort([("last_opened_at", -1), ("title", 1)]).limit(limit)
    out = []
    async for b in cursor:
        out.append({
            "book_id": b["book_id"],
            "title": b.get("title", ""),
            "author": b.get("author", ""),
            "category": b.get("category", ""),
            "fandom": b.get("fandom", []),
        })
    return {"books": out}


@api_router.get("/books/{book_id}/links")
async def get_book_links(book_id: str, user: User = Depends(get_current_user)):
    """Download the extracted URLs for a single book as a .txt file."""
    book = await db.books.find_one({"book_id": book_id, "user_id": user.user_id}, {"_id": 0})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")

    user_dir = STORAGE_DIR / user.user_id
    links_path = user_dir / f"{book_id}.links.txt"

    # Regenerate if missing (e.g., older book uploaded before this feature)
    if not links_path.exists():
        epub_path = user_dir / f"{book_id}.epub"
        if not epub_path.exists():
            raise HTTPException(status_code=404, detail="File missing")
        links = extract_urls_from_epub(epub_path)
        links_path.write_text(
            format_links_txt(book['title'], book['author'], links),
            encoding='utf-8',
        )
        await db.books.update_one(
            {"book_id": book_id, "user_id": user.user_id},
            {"$set": {"links_count": len(links)}},
        )

    filename = _safe_filename(book.get('title') or book_id, '.links.txt')
    return FileResponse(str(links_path), media_type="text/plain; charset=utf-8", filename=filename)


class ReclassifyBody(BaseModel):
    use_ai: bool = True


class ReclassifyAllBody(BaseModel):
    only_unclassified: bool = True
    category: Optional[str] = None
    fandom: Optional[str] = None


@api_router.post("/books/reclassify-all")
async def reclassify_all(body: ReclassifyAllBody, user: User = Depends(get_current_user)):
    """Run the AI classifier on every matching book, then persist the new labels."""
    query: Dict[str, Any] = {"user_id": user.user_id}
    if body.only_unclassified:
        query["category"] = "Unclassified"
    else:
        if body.category:
            query["category"] = body.category
        if body.fandom:
            query["fandom"] = body.fandom

    books = await db.books.find(query, {"_id": 0}).to_list(5000)
    if not books:
        return {"processed": 0, "changed": 0}

    user_dir = STORAGE_DIR / user.user_id
    sem = asyncio.Semaphore(3)  # cap concurrent AI calls

    async def process(b):
        async with sem:
            fp = user_dir / f"{b['book_id']}.epub"
            if not fp.exists():
                return None
            try:
                meta = extract_epub_metadata(fp)
                cls = await classify_with_ai(meta)
            except Exception as e:
                logger.error(f"AI reclass error for {b['book_id']}: {e}")
                return None
            if cls['confidence'] <= 0:
                return None
            return (b['book_id'], cls)

    results = await asyncio.gather(*[process(b) for b in books])
    changed = 0
    for r in results:
        if not r:
            continue
        bid, cls = r
        await db.books.update_one(
            {"book_id": bid, "user_id": user.user_id},
            {"$set": {
                "category": cls['category'],
                "fandom": cls.get('fandom'),
                "confidence": cls['confidence'],
                "classifier": cls['classifier'],
            }},
        )
        changed += 1
    return {"processed": len(books), "changed": changed}


@api_router.post("/books/{book_id}/reclassify")
async def reclassify_book(book_id: str, body: ReclassifyBody, user: User = Depends(get_current_user)):
    book = await db.books.find_one({"book_id": book_id, "user_id": user.user_id}, {"_id": 0})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")
    fp = STORAGE_DIR / user.user_id / f"{book_id}.epub"
    if not fp.exists():
        raise HTTPException(status_code=404, detail="File missing")
    meta = extract_epub_metadata(fp)
    classification = await classify_book(meta, force_ai=body.use_ai)
    await db.books.update_one(
        {"book_id": book_id},
        {"$set": {
            "category": classification['category'],
            "fandom": _canonicalize_fandom(classification.get('fandom')),
            "confidence": classification['confidence'],
            "classifier": classification['classifier'],
        }},
    )
    return classification


class UpdateBookBody(BaseModel):
    category: Optional[str] = None
    fandom: Optional[str] = None
    title: Optional[str] = None
    author: Optional[str] = None
    description: Optional[str] = None


# NOTE: The block of fanfic-refresh endpoints + the health-probe / sweep
# helpers (`refresh-all`, `fanfic/status`, `retry-unavailable`, plus
# `_probe_fanfic_now`, `_sweep_user_unavailable`, `_fanfic_status_cache`)
# was moved to routes/refresh.py in the Phase 4 refactor (2026-06-14).


# NOTE: `POST /books/{book_id}/mark`, `/heartbeat`, `/progress`, `/touch`
# and the shared `_log_activity` helper moved to routes/reading_activity.py
# in the Phase 5F refactor (2026-06-14).


@api_router.get("/books/{book_id}/diff")
async def book_diff(
    book_id: str,
    vs: Optional[str] = None,
    user: User = Depends(get_current_user),
):
    """Per-chapter diff between two versions of a book.

    If `vs` is omitted, auto-resolves the counterpart via the book's
    `replaces` (current is "Updated stories") or `replaced_by` (current is
    "Old stories") link.

    Returns: old + new metadata, chapter lists, and a structured diff payload.
    """
    book = await db.books.find_one({"book_id": book_id, "user_id": user.user_id}, {"_id": 0})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")

    counterpart_id = vs
    if not counterpart_id:
        counterpart_id = book.get("replaces") or book.get("replaced_by")
    if not counterpart_id:
        raise HTTPException(
            status_code=400,
            detail="No counterpart version found. Refresh this book first to create a version history, or pass ?vs={other_book_id}.",
        )

    other = await db.books.find_one(
        {"book_id": counterpart_id, "user_id": user.user_id}, {"_id": 0}
    )
    if not other:
        raise HTTPException(status_code=404, detail="Counterpart book not found")

    # Order them: old version first, new version second
    if book.get("replaced_by") == counterpart_id:
        old_doc, new_doc = book, other
    elif book.get("replaces") == counterpart_id:
        old_doc, new_doc = other, book
    else:
        # Explicit vs= without a link — use timestamps to order
        old_doc, new_doc = book, other
        if (other.get("created_at") or "") < (book.get("created_at") or ""):
            old_doc, new_doc = other, book

    user_dir = STORAGE_DIR / user.user_id
    old_path = user_dir / f"{old_doc['book_id']}.epub"
    new_path = user_dir / f"{new_doc['book_id']}.epub"
    if not old_path.exists() or not new_path.exists():
        raise HTTPException(status_code=404, detail="One or both EPUB files are missing on disk")

    loop = asyncio.get_event_loop()
    old_chapters = await loop.run_in_executor(None, extract_chapters, old_path)
    new_chapters = await loop.run_in_executor(None, extract_chapters, new_path)
    diff = diff_chapters(old_chapters, new_chapters)

    def _doc_summary(d: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "book_id": d["book_id"],
            "title": d.get("title", ""),
            "author": d.get("author", ""),
            "category": d.get("category", ""),
            "created_at": d.get("created_at"),
            "last_refreshed_at": d.get("last_refreshed_at"),
            "replaced_at": d.get("replaced_at"),
        }

    return {
        "old": {**_doc_summary(old_doc), "chapters": old_chapters},
        "new": {**_doc_summary(new_doc), "chapters": new_chapters},
        "diff": diff,
    }




class BulkIdsBody(BaseModel):
    book_ids: List[str]


class BulkMoveBody(BaseModel):
    book_ids: List[str]
    category: Optional[str] = None
    fandom: Optional[str] = None


@api_router.post("/books/bulk/delete")
async def bulk_delete(body: BulkIdsBody, user: User = Depends(get_current_user)):
    """Bulk soft-delete — books move to Trash with a 30-day grace window so
    accidental "select all → delete" mishaps are reversible. Use
    `/api/trash/empty` if you want immediate hard deletion afterwards.
    """
    if not body.book_ids:
        return {"deleted": 0}
    now_dt = datetime.now(timezone.utc)
    now_iso = now_dt.isoformat()
    expires_at = (now_dt + timedelta(days=TRASH_GRACE_DAYS)).isoformat()
    # Record each book's prior category so it can be restored from Trash
    cursor = db.books.find(
        {"book_id": {"$in": body.book_ids}, "user_id": user.user_id, "category": {"$ne": TRASH_SHELF}},
        {"_id": 0, "book_id": 1, "category": 1},
    )
    moved = 0
    async for b in cursor:
        await db.books.update_one(
            {"book_id": b["book_id"], "user_id": user.user_id},
            {
                "$set": {
                    "category": TRASH_SHELF,
                    "trash_expires_at": expires_at,
                    "dupe_action_meta": {
                        "action": "discard",
                        "prev_category_new": b.get("category"),
                        "applied_at": now_iso,
                    },
                },
            },
        )
        moved += 1
    return {"deleted": moved, "trashed": moved, "trash_expires_at": expires_at}


class ResetStateBody(BaseModel):
    reset_progress: bool = False  # progress_fraction, last_opened_at, reading_minutes, reading_activity
    reset_tags: bool = False       # clear book.tags
    reset_smart_shelves: bool = False  # drop user's smart_shelves
    reset_versions: bool = False   # collapse "Old stories"/"Updated stories YYYY-MM-DD" back into a single category


@api_router.post("/books/reset-state")
async def reset_state(body: ResetStateBody, user: User = Depends(get_current_user)):
    """Selectively wipe non-book metadata while keeping every EPUB intact.

    Each flag is independent — pass `true` only on the dimensions you want to clear.
    All books and their files stay on disk; only DB metadata is touched.
    """
    if not any([body.reset_progress, body.reset_tags, body.reset_smart_shelves, body.reset_versions]):
        raise HTTPException(status_code=400, detail="Pick at least one thing to reset.")

    summary: Dict[str, int] = {}

    if body.reset_progress:
        r = await db.books.update_many(
            {"user_id": user.user_id},
            {"$unset": {
                "progress_fraction": "",
                "last_opened_at": "",
                "reading_minutes": "",
                "manually_uploaded_at": "",
            }},
        )
        ra = await db.reading_activity.delete_many({"user_id": user.user_id})
        summary["books_progress_cleared"] = r.modified_count
        summary["activity_rows_deleted"] = ra.deleted_count

    if body.reset_tags:
        r = await db.books.update_many(
            {"user_id": user.user_id},
            {"$set": {"tags": []}},
        )
        summary["books_tags_cleared"] = r.modified_count

    if body.reset_smart_shelves:
        ss = await db.smart_shelves.delete_many({"user_id": user.user_id})
        summary["smart_shelves_deleted"] = ss.deleted_count

    if body.reset_versions:
        # Collapse old/updated shelves back to their best-guess category.
        # If a book has a fandom we send it to "Fanfiction", else to "Unclassified".
        cursor = db.books.find(
            {
                "user_id": user.user_id,
                "$or": [
                    {"category": OLD_STORIES_SHELF},
                    {"category": {"$regex": r"^Updated stories \d{4}-\d{2}-\d{2}$"}},
                ],
            },
            {"_id": 0, "book_id": 1, "fandom": 1},
        )
        count = 0
        async for b in cursor:
            target = "Fanfiction" if b.get("fandom") else "Unclassified"
            await db.books.update_one(
                {"book_id": b["book_id"], "user_id": user.user_id},
                {
                    "$set": {"category": target},
                    "$unset": {"replaced_by": "", "replaces": "", "replaced_at": "", "refresh_summary": "", "update_seen": "", "manually_uploaded_at": ""},
                },
            )
            count += 1
        # And remove any auto-created dated-shelf entries in `categories`
        await db.categories.delete_many({
            "user_id": user.user_id,
            "auto_created": True,
            "name": {"$regex": r"^Updated stories \d{4}-\d{2}-\d{2}$"},
        })
        summary["versions_collapsed"] = count

    return {"ok": True, **summary}


class WipeLibraryBody(BaseModel):
    confirm: str  # must equal "DELETE_EVERYTHING"


@api_router.post("/books/wipe-library")
async def wipe_library(body: WipeLibraryBody, user: User = Depends(get_current_user)):
    """Delete every book the user owns — DB rows, EPUBs, covers, link sidecars.
    Requires `confirm == "DELETE_EVERYTHING"` so an accidental POST won't nuke a library.

    Note: also clears reading_activity, smart_shelves, and the templated-onboarding
    flag so the user effectively starts fresh.
    """
    if body.confirm != "DELETE_EVERYTHING":
        raise HTTPException(
            status_code=400,
            detail='Confirmation required. Pass {"confirm": "DELETE_EVERYTHING"} to proceed.',
        )

    # Drop every on-disk file under the user's storage dir
    user_dir = STORAGE_DIR / user.user_id
    files_removed = 0
    if user_dir.exists():
        for p in user_dir.iterdir():
            try:
                if p.is_file():
                    p.unlink()
                    files_removed += 1
            except Exception as e:
                logger.warning("wipe_library couldn't delete %s: %s", p, e)

    # Drop collections scoped to this user
    deletes = {
        "books": (await db.books.delete_many({"user_id": user.user_id})).deleted_count,
        "reading_activity": (await db.reading_activity.delete_many({"user_id": user.user_id})).deleted_count,
        "smart_shelves": (await db.smart_shelves.delete_many({"user_id": user.user_id})).deleted_count,
        "categories": (await db.categories.delete_many({"user_id": user.user_id})).deleted_count,
    }
    # Reset onboarding so the user-prompt can show again on fresh re-upload
    await db.users.update_one(
        {"user_id": user.user_id},
        {"$unset": {
            "template_prompt_dismissed": "",
            "template_prompt_accepted": "",
            "template_prompt_dismissed_at": "",
        }},
    )

    return {
        "ok": True,
        "files_removed": files_removed,
        **deletes,
        "message": f"Library wiped: {deletes['books']} books and {files_removed} files removed.",
    }




@api_router.post("/books/bulk/move")
async def bulk_move(body: BulkMoveBody, user: User = Depends(get_current_user)):
    if not body.book_ids:
        return {"updated": 0}
    update: Dict[str, Any] = {"classifier": "manual", "confidence": 1.0}
    if body.category is not None:
        update["category"] = body.category
    if body.fandom is not None:
        update["fandom"] = _canonicalize_fandom(body.fandom) if body.fandom else None
    if len(update) == 2:  # only classifier+confidence — nothing to move to
        raise HTTPException(status_code=400, detail="No category or fandom provided")
    result = await db.books.update_many(
        {"book_id": {"$in": body.book_ids}, "user_id": user.user_id},
        {"$set": update},
    )
    return {"updated": result.modified_count}


class BulkMetadataBody(BaseModel):
    book_ids: List[str]
    author: Optional[str] = None      # if provided & non-empty, sets author on all
    fandom: Optional[str] = None      # "" => clear fandom (None); None => leave as-is
    category: Optional[str] = None    # category to set on all
    series_name: Optional[str] = None # "" => clear series; None => leave as-is
    series_start_index: Optional[float] = None  # if set, assigns series_index sequentially starting at this value
    title_prefix_strip: Optional[str] = None    # if provided & non-empty, strips this prefix from each book's title
    add_tags: Optional[List[str]] = None         # tags to ADD to every selected book
    remove_tags: Optional[List[str]] = None      # tags to REMOVE from every selected book


@api_router.post("/books/bulk/metadata")
async def bulk_metadata(body: BulkMetadataBody, user: User = Depends(get_current_user)):
    """Edit metadata across many books at once.

    Use cases: fix a misspelled author across a series, drop everything into a
    new fandom shelf, group books into a series and number them in upload order,
    or strip a common prefix from titles (e.g. "[OLD] ").
    """
    if not body.book_ids:
        return {"updated": 0}

    # Fields that apply identically to every selected book
    set_common: Dict[str, Any] = {}
    unset_common: Dict[str, Any] = {}
    if body.author and body.author.strip():
        set_common["author"] = body.author.strip()
    if body.category is not None:
        set_common["category"] = body.category
        set_common["classifier"] = "manual"
        set_common["confidence"] = 1.0
    if body.fandom is not None:
        if body.fandom.strip():
            set_common["fandom"] = _canonicalize_fandom(body.fandom.strip())
        else:
            unset_common["fandom"] = ""
    if body.series_name is not None and body.series_start_index is None:
        if body.series_name.strip():
            set_common["series_name"] = body.series_name.strip()
        else:
            unset_common["series_name"] = ""
            unset_common["series_index"] = ""

    updated = 0

    if set_common or unset_common:
        ops: Dict[str, Any] = {}
        if set_common:
            ops["$set"] = set_common
        if unset_common:
            ops["$unset"] = unset_common
        result = await db.books.update_many(
            {"book_id": {"$in": body.book_ids}, "user_id": user.user_id},
            ops,
        )
        updated = max(updated, result.modified_count)

    # Series numbering: assign sequentially in the order book_ids was provided
    if body.series_name is not None and body.series_start_index is not None and body.series_name.strip():
        idx = float(body.series_start_index)
        for bid in body.book_ids:
            await db.books.update_one(
                {"book_id": bid, "user_id": user.user_id},
                {"$set": {"series_name": body.series_name.strip(), "series_index": idx}},
            )
            idx += 1
        updated = max(updated, len(body.book_ids))

    # Title prefix strip (per-book, since each title is different)
    if body.title_prefix_strip and body.title_prefix_strip.strip():
        prefix = body.title_prefix_strip
        books = await db.books.find(
            {"book_id": {"$in": body.book_ids}, "user_id": user.user_id},
            {"_id": 0, "book_id": 1, "title": 1},
        ).to_list(5000)
        for b in books:
            t = b.get("title") or ""
            if t.startswith(prefix):
                new_t = t[len(prefix):].lstrip()
                if new_t and new_t != t:
                    await db.books.update_one(
                        {"book_id": b["book_id"], "user_id": user.user_id},
                        {"$set": {"title": new_t}},
                    )
                    updated += 1

    # Bulk tag add/remove
    add = _normalize_tags(body.add_tags or [])
    rm = _normalize_tags(body.remove_tags or [])
    if add:
        result = await db.books.update_many(
            {"book_id": {"$in": body.book_ids}, "user_id": user.user_id},
            {"$addToSet": {"tags": {"$each": add}}},
        )
        updated = max(updated, result.modified_count)
    if rm:
        result = await db.books.update_many(
            {"book_id": {"$in": body.book_ids}, "user_id": user.user_id},
            {"$pull": {"tags": {"$in": rm}}},
        )
        updated = max(updated, result.modified_count)

    return {"updated": updated}


# ============================================================
# TAGS ROUTES — extracted to routes/tags.py in books.py Phase 2 refactor.
# See ``backend/routes/tags.py`` for the 7 endpoints under /api/tags/* and
# /api/books/{book_id}/tags*. They still register on the same shared
# api_router so URLs are unchanged.
# ============================================================


# ============================================================
# AUTHOR ROUTES — extracted to routes/authors.py in Phase 2 refactor.
# See ``backend/routes/authors.py`` for /authors, /library/authors,
# and /library/by-author.
# ============================================================
@api_router.get("/fandoms")
async def list_fandoms(user: User = Depends(get_current_user)):
    """Distinct fandoms in the user's library with book counts.

    Used by the Download page so all fandoms appear (not just the top 8 that
    /stats/overview returns for the dashboard). Each row is annotated with
    `is_crossover` + `parts` so the UI can render the crossover treatment
    without re-parsing strings.
    """
    pipeline = [
        {"$match": {"user_id": user.user_id, "fandom": {"$ne": None, "$exists": True}}},
        {"$group": {"_id": "$fandom", "count": {"$sum": 1}}},
        {"$sort": {"count": -1, "_id": 1}},
    ]
    rows = await db.books.aggregate(pipeline).to_list(5000)
    fandoms: List[Dict[str, Any]] = []
    crossover_count = 0
    for r in rows:
        name = r.get("_id")
        if not name or not str(name).strip():
            continue
        parts = [p.strip() for p in str(name).split(" / ") if p.strip()]
        is_x = len(parts) >= 2
        if is_x:
            crossover_count += 1
        fandoms.append({
            "name": name,
            "count": r["count"],
            "is_crossover": is_x,
            "parts": parts if is_x else [],
        })
    return {"fandoms": fandoms, "crossover_count": crossover_count}


# ============================================================
# COMPLETION-STATUS SHELVES (complete / ongoing)
# Detection runs once at upload time and persists to `books.status`.
# User overrides land in `books.manual_status`; `effective_status()`
# picks the override when set. Counts on the dashboard come from
# ============================================================
# COMPLETE / ONGOING STATUS — list endpoints, status counts,
# and the `_status_query` / `_list_status_shelf` helpers moved
# to ``routes/library_views.py`` in the Phase 5E refactor.
# Only the mutator `PATCH /books/{book_id}/status` stays here
# (it's a write path, not a view).
# ============================================================
class SetStatusBody(BaseModel):
    """Body for `PATCH /books/{book_id}/status`. `status=None` clears the
    manual override and falls back to the auto-detected value."""
    status: Optional[str] = None


@api_router.patch("/books/{book_id}/status")
async def set_book_status(
    book_id: str,
    body: SetStatusBody,
    user: User = Depends(get_current_user),
):
    """Override the auto-detected completion status for a single book.

    Persists to `manual_status` so a future re-detection (or refresh)
    can't blow the user's override away — choice 4b. Passing `status:
    null` clears the override and reverts to the auto-detected value.

    Accepts only `"complete"` / `"ongoing"` / `null`.
    """
    raw = (body.status or "").strip().lower()
    if raw and raw not in (STATUS_COMPLETE, STATUS_ONGOING):
        raise HTTPException(
            status_code=400,
            detail=f"status must be one of: {STATUS_COMPLETE}, {STATUS_ONGOING}, null",
        )
    update = (
        {"$set": {"manual_status": raw}}
        if raw else
        {"$unset": {"manual_status": ""}}
    )
    res = await db.books.update_one(
        {"book_id": book_id, "user_id": user.user_id},
        update,
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Book not found")
    book = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "book_id": 1, "status": 1, "manual_status": 1},
    )
    return {
        "ok": True,
        "book_id": book_id,
        "status": book.get("status"),
        "manual_status": book.get("manual_status"),
        "effective_status": effective_status(book),
    }


# ============================================================
# AUTHOR SHELVES — extracted to routes/authors.py in Phase 2 refactor.
# See ``backend/routes/authors.py`` for /library/authors + /library/by-author.
# ============================================================


# ============================================================
# PAIRINGS / SHIP BROWSER — extracted to routes/pairings.py in Phase 2.
# See ``backend/routes/pairings.py`` for /library/pairings + /library/by-pairing.
# ============================================================



# Library backup, restore, backup-reminder and backup-history
# endpoints live in routes/library_backup.py (extracted 2026-06-13).



# NOTE: `GET /library/linkless` was moved to routes/library_views.py
# in the Phase 5E refactor (2026-06-14).


@api_router.get("/admin/unknown-sources")
async def list_unknown_sources(
    since: Optional[str] = None,
    user: User = Depends(require_admin),
):
    """Return every story-shaped URL host that's NOT on the accepted-sources
    list but has been pasted/uploaded by ANY user. Sorted by `last_seen`
    descending so newly-spotted hosts surface first.

    Used by the Shelfsort dev (the agent reviewing this codebase) to
    decide which hosts to add to `utils/url_canonical`. Returns:
      * `host` — the de-subdomain'd root host (e.g. `scribblehub.com`)
      * `hit_count` — total times we saw a URL on this host
      * `contexts` — dict of {upload|paste|claim → count}
      * `samples` — up to 5 sample full URLs (most recent)
      * `first_seen` / `last_seen`
      * `last_book_title` / `last_book_author` / `last_book_id` (upload-only)

    Optional `?since=<iso8601>` filters to hosts seen on/after the
    timestamp so the agent can poll for "what's new this session".
    Authentication is required; data is global (not per-user) because
    the accepted-list lives at the codebase level.
    """
    query: Dict[str, Any] = {}
    if since:
        try:
            cutoff = datetime.fromisoformat(since.replace("Z", "+00:00"))
            query["last_seen"] = {"$gte": cutoff}
        except ValueError:
            pass  # silently ignore malformed cutoff
    cursor = db.unknown_sources.find(query, {"_id": 0}).sort("last_seen", -1)
    rows = await cursor.to_list(500)
    # ISO-serialize datetimes so the response is JSON-safe.
    for r in rows:
        for k in ("first_seen", "last_seen"):
            v = r.get(k)
            if isinstance(v, datetime):
                r[k] = v.isoformat()
    return {"count": len(rows), "hosts": rows}


@api_router.delete("/admin/unknown-sources/{host}")
async def dismiss_unknown_source(host: str, user: User = Depends(require_admin)):
    """Drop a host record after it's been actioned (either added to the
    accepted-sources list or confirmed-not-fanfic). Idempotent — returns
    `{ok: True, removed: 0|1}`."""
    res = await db.unknown_sources.delete_one({"host": host.lower()})
    if res.deleted_count:
        await record_admin_action(user, "unknown_source.dismiss", target=host.lower())
    return {"ok": True, "removed": res.deleted_count}


class MarkAcceptedBody(BaseModel):
    accepted: bool = True


class AddUnknownSourceBody(BaseModel):
    """Body for `POST /api/admin/unknown-sources` — manual queue add.

    The user is vouching for the URL (probably saw it on a new archive a
    friend mentioned), so we bypass the story-shape heuristic. Accepted
    sources are still skipped — no point logging a host the canonicalizer
    already knows about. `note` is a free-form comment that lives on the
    host record so the dev knows why it was queued.
    """
    url: str
    note: Optional[str] = None


@api_router.post("/admin/unknown-sources")
async def add_unknown_source_manual(
    body: AddUnknownSourceBody,
    user: User = Depends(require_admin),
):
    """Manually queue a host for review without an EPUB upload trigger.

    Returns `{ok, host, already_accepted}` — `already_accepted=True`
    means the URL canonicalizes to a known source (no record created
    because we already support it). When `host=None` the URL parsed but
    we couldn't extract a hostname (e.g. user pasted just a path).
    """
    raw = (body.url or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="URL is empty")

    from utils.unknown_sources import (
        record_unknown_sources, _host_of, normalize_fanfic_url,
    )

    # Already-accepted shortcut so the UI can tell the user "no need to
    # queue, we already support this" instead of silently doing nothing.
    if normalize_fanfic_url(raw) or classify_ao3_non_work(raw):
        return {
            "ok": True,
            "already_accepted": True,
            "host": _host_of(raw),
        }

    hosts = await record_unknown_sources(
        db, [raw], context="manual",
        user_id=user.user_id,
        note=body.note,
        skip_heuristic=True,
    )
    if not hosts:
        # Either the URL was un-parseable or had no hostname.
        raise HTTPException(
            status_code=400,
            detail="Couldn't extract a hostname from that URL — please paste a full http(s):// URL.",
        )
    return {"ok": True, "already_accepted": False, "host": hosts[0]}


@api_router.patch("/admin/unknown-sources/{host}/mark-accepted")
async def mark_unknown_source_accepted(
    host: str,
    body: MarkAcceptedBody,
    user: User = Depends(require_admin),
):
    """Flag (or un-flag) an unknown-source host as "user wants this added to
    the accepted-sources list." The flag is purely a signal for the next
    Shelfsort dev session — the host stays in `unknown_sources` until it's
    either dismissed (DELETE) or the regex is actually added to
    `utils/url_canonical.py` and the host record explicitly dismissed.

    Idempotent; returns the updated host doc.
    """
    host_norm = host.lower()
    now = datetime.now(timezone.utc)
    update = (
        {"$set": {"marked_accepted": True, "marked_accepted_at": now,
                  "marked_accepted_by": user.user_id}}
        if body.accepted else
        {"$unset": {"marked_accepted": "", "marked_accepted_at": "",
                    "marked_accepted_by": ""}}
    )
    res = await db.unknown_sources.update_one({"host": host_norm}, update)
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Host not found")
    doc = await db.unknown_sources.find_one({"host": host_norm}, {"_id": 0})
    for k in ("first_seen", "last_seen", "marked_accepted_at"):
        v = doc.get(k) if doc else None
        if isinstance(v, datetime):
            doc[k] = v.isoformat()
    return {"ok": True, "host": doc}


# NOTE: `GET /library/unreadable` was moved to routes/library_views.py
# in the Phase 5E refactor (2026-06-14).


@api_router.get("/books/{book_id}/download-original")
async def download_original_file(book_id: str, user: User = Depends(get_current_user)):
    """Serve the user's original (pre-conversion) source file.

    Used by the Unreadable shelf when an upload was a PDF/Kindle/DOCX that
    Calibre couldn't convert — the EPUB target was never written, but the
    original bytes still live at `{book_id}.{original_format}`. Falls back
    to whichever `.{format}` file actually exists on disk so this also
    works for an `Originals` book the user wants the source for.
    """
    book = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "book_id": 1, "title": 1, "author": 1,
         "original_format": 1, "filename": 1},
    )
    if not book:
        raise HTTPException(status_code=404, detail="Not found")
    ext = (book.get("original_format") or "").lstrip(".")
    user_dir = STORAGE_DIR / user.user_id
    candidate = user_dir / f"{book_id}.{ext}" if ext else None
    fp = None
    if candidate and candidate.exists():
        fp = candidate
    else:
        # Last-ditch fallback: scan the user dir for any file starting with
        # the book id. Covers the case where `original_format` was lost or
        # stored without an extension.
        for p in user_dir.glob(f"{book_id}.*"):
            if p.suffix.lower() not in (".cover", ".links.txt"):
                fp = p
                ext = p.suffix.lstrip(".")
                break
    if not fp or not fp.exists():
        raise HTTPException(status_code=404, detail="Original file missing on disk")
    download_name = _templated_filename(
        book.get("title"), book.get("author"), book_id, ext=f".{ext or 'bin'}",
    )
    return FileResponse(str(fp), filename=download_name)



class ClaimSourceUrlBody(BaseModel):
    """Body for `PATCH /books/{book_id}/source-url`.

    Accepts either field name — `url` (newer Linkless-shelf clients) or
    `source_url` (older "manual correction" clients / tests) — so we
    don't break either caller while we have just one endpoint.
    """
    url: Optional[str] = None
    source_url: Optional[str] = None


@api_router.patch("/books/{book_id}/source-url")
async def claim_source_url(
    book_id: str,
    body: ClaimSourceUrlBody,
    user: User = Depends(get_current_user),
):
    """Attach (or correct) the fanfic source URL on an existing book.

    Used by:
      * the Linkless library shelf — paste the URL the book "actually"
        came from to drop it out of `/library/linkless`;
      * the "Can't find online" flow — manually correct the URL after
        FanFicFare failed to identify it.

    The URL is normalized to canonical form (per source site) and
    written to BOTH `source_url` and `fanfic_urls` so future URL-list
    dedupe matches it. Also clears the `unavailable` / `last_fetch_error`
    flags so the next refresh tries the new URL.

    Rejects URLs that don't match any known fanfic source.
    """
    raw = (body.url or body.source_url or "").strip()
    if not raw:
        raise HTTPException(status_code=400, detail="Source URL is empty")
    canon = _canonical_fanfic_url(raw)
    if not canon:
        # User pasted something they THOUGHT was a fanfic URL but the host
        # isn't on the accepted list. Log it for review before rejecting.
        try:
            from utils.unknown_sources import record_unknown_sources
            await record_unknown_sources(
                db, [raw], context="claim", user_id=user.user_id, book_id=book_id,
            )
        except Exception as _e:
            logger.warning("unknown_sources record failed for claim_source_url: %s", _e)
        raise HTTPException(
            status_code=400,
            detail="Not a recognized fanfic source URL. We support AO3, FFnet, FictionPress, RoyalRoad, SpaceBattles, SufficientVelocity, QQ, AFF, Potions & Snitches, and Twilighted.",
        )
    book = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "book_id": 1, "fanfic_urls": 1},
    )
    if not book:
        raise HTTPException(status_code=404, detail="Book not found.")
    # Edge case: another book in the user's library already owns this URL.
    # If we silently overwrite we end up with two books bearing the same
    # source_url and future URL-list dedupe collapses into a coin-toss. Surface
    # the collision via 409 so the frontend can offer "open the other book
    # instead" rather than leaving the user with a hidden duplicate. The trash
    # shelf is excluded from the collision check — restoring a trashed book
    # via its source URL is a legitimate workflow.
    conflict = await db.books.find_one(
        {
            "user_id": user.user_id,
            "book_id": {"$ne": book_id},
            "category": {"$ne": TRASH_SHELF},
            "$or": [
                {"source_url": canon},
                {"fanfic_urls": canon},
            ],
        },
        {"_id": 0, "book_id": 1, "title": 1, "author": 1, "fandom": 1},
    )
    if conflict:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "url_already_claimed",
                "message": "Another book in your library already has this URL.",
                "conflict_book": {
                    "book_id": conflict.get("book_id"),
                    "title": conflict.get("title") or "Untitled",
                    "author": conflict.get("author") or "Unknown author",
                    "fandom": conflict.get("fandom"),
                },
            },
        )
    existing_urls = book.get("fanfic_urls") or []
    if canon not in existing_urls:
        existing_urls = [canon, *existing_urls]
    await db.books.update_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"$set": {
            "source_url": canon,
            "fanfic_urls": existing_urls,
            "unavailable": False,
            "last_fetch_error": None,
        }},
    )
    return {
        "ok": True,
        "book_id": book_id,
        "source_url": canon,
        "fanfic_urls": existing_urls,
    }


# NOTE: /api/fandoms/* routes were moved to routes/fandoms.py in the
# Phase 5 refactor (2026-06-14).





@api_router.get("/authors/{name}")
async def get_author(name: str, user: User = Depends(get_current_user)):
    """All books by this author, newest first."""
    books = await db.books.find(
        {"user_id": user.user_id, "author": name},
        {"_id": 0},
    ).sort("created_at", -1).to_list(2000)
    return {"name": name, "books": books}


@api_router.patch("/books/{book_id}")
async def update_book(book_id: str, body: UpdateBookBody, user: User = Depends(get_current_user)):
    book = await db.books.find_one({"book_id": book_id, "user_id": user.user_id}, {"_id": 0})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")
    update: Dict[str, Any] = {}
    classifier_touched = False
    if body.category is not None:
        update['category'] = body.category
        classifier_touched = True
    if body.fandom is not None:
        update['fandom'] = _canonicalize_fandom(body.fandom) if body.fandom else None
        classifier_touched = True
    # Length-limited so a paste-bomb description can't blow up our docs.
    if body.title is not None:
        update['title'] = body.title.strip()[:500]
    if body.author is not None:
        update['author'] = body.author.strip()[:500]
    if body.description is not None:
        update['description'] = body.description.strip()[:5000]
    if classifier_touched:
        update['classifier'] = 'manual'
        update['confidence'] = 1.0

    if not update:
        return {"ok": True, "noop": True}

    await db.books.update_one({"book_id": book_id, "user_id": user.user_id}, {"$set": update})

    # If any user-visible metadata field changed AND we still have the EPUB on
    # disk (i.e. not a link-only ingest), rewrite the EPUB so downloads stay
    # consistent. DB always wins — EPUB failures don't fail the request.
    epub_written = None
    file_only_fields = any(v is not None for v in (body.title, body.author, body.description))
    if file_only_fields:
        fp = STORAGE_DIR / user.user_id / f"{book_id}.epub"
        if fp.exists():
            result = update_epub_metadata(
                fp,
                title=body.title,
                author=body.author,
                description=body.description,
            )
            epub_written = bool(result.get("ok"))
            if not epub_written:
                logger.info(f"In-place EPUB metadata write skipped for {book_id}: {result.get('error')}")
    return {"ok": True, "epub_updated": epub_written}


# NOTE: /api/books/export/zip + _safe_folder helper were moved to
# routes/exports.py in the Phase 5 refactor (2026-06-14).



@api_router.post("/books/detect-series-all")
async def detect_series_all(user: User = Depends(get_current_user)):
    """Re-scan every book without a series_name and try to detect one from the title."""
    books = await db.books.find(
        {"user_id": user.user_id},
        {"_id": 0, "book_id": 1, "title": 1, "series_name": 1},
    ).to_list(5000)
    user_dir = STORAGE_DIR / user.user_id
    found = 0
    for b in books:
        if b.get("series_name"):
            continue
        # Try filesystem EPUB metadata first
        sn = None
        si = None
        fp = user_dir / f"{b['book_id']}.epub"
        if fp.exists():
            try:
                m = extract_epub_metadata(fp)
                sn = m.get("series_name")
                si = m.get("series_index")
            except Exception:
                pass
        if not sn:
            sn, si = detect_series_from_title(b.get("title") or "")
        if sn:
            await db.books.update_one(
                {"book_id": b["book_id"], "user_id": user.user_id},
                {"$set": {"series_name": sn, "series_index": si}},
            )
            found += 1
    return {"scanned": len(books), "found": found}


class SetSourceBody(BaseModel):
    source_url: str

class SetSeriesBody(BaseModel):
    series_name: Optional[str] = None
    series_index: Optional[float] = None


@api_router.patch("/books/{book_id}/series")
async def set_series(book_id: str, body: SetSeriesBody, user: User = Depends(get_current_user)):
    update: Dict[str, Any] = {
        "series_name": (body.series_name.strip() if body.series_name else None),
        "series_index": body.series_index,
    }
    result = await db.books.update_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"$set": update},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}



@api_router.post("/books/{book_id}/upload-new-version")
async def upload_new_version(
    book_id: str,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
):
    """Upload a freshly-downloaded EPUB as a NEW version of the given book.

    Mirrors the refresh flow exactly:
      * new book record on a date-stamped 'Updated stories YYYY-MM-DD' shelf
      * old book archived to 'Old stories' (with replaced_by back-pointer)
      * tags / source_url / fandom / series / classifier carried over
      * refresh_summary computed for the bell badge + email digest
      * house template applied (if enabled)
      * existing "Old stories" / archived books are NOT re-versioned

    Use case: when FanFicFare can't fetch from the source (bot protection,
    Cloudflare, locked work) you grab the EPUB locally and drop it here.
    """
    book = await db.books.find_one({"book_id": book_id, "user_id": user.user_id})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")
    # Block uploads onto already-archived books — pick the current copy instead
    if book.get("category") == OLD_STORIES_SHELF or book.get("replaced_by"):
        raise HTTPException(
            status_code=400,
            detail="This is already an archived copy. Open the current version and upload there.",
        )

    # Validate
    name = (file.filename or "").lower()
    if not name.endswith(".epub"):
        raise HTTPException(status_code=400, detail="Please upload an .epub file")
    raw = await file.read()
    if not raw or len(raw) < 256:
        raise HTTPException(status_code=400, detail="That file is empty or too small to be an EPUB")
    if not raw.startswith(b"PK\x03\x04"):
        raise HTTPException(status_code=400, detail="That doesn't look like a valid EPUB (zip header missing)")

    # Antivirus pre-scan — same policy as /books/upload (2026-06-18).
    from utils.antivirus import scan_bytes, record_quarantine
    import asyncio as _asyncio
    _av = await _asyncio.to_thread(scan_bytes, raw, hint_name=file.filename or "version.epub")
    if _av.get("infected"):
        await record_quarantine(
            user_id=user.user_id,
            filename=file.filename or "",
            scan=_av,
            source="upload",
            extra={"book_id": book_id, "size_bytes": len(raw)},
        )
        raise HTTPException(
            status_code=400,
            detail=f"\"{file.filename or 'this file'}\" appears unsafe ({_av.get('signature') or 'flagged by antivirus'}). Upload blocked.",
        )

    # Apply the house template (idempotent — noop if already templated)
    user_dir = STORAGE_DIR / user.user_id
    user_dir.mkdir(parents=True, exist_ok=True)
    old_book_id = book_id

    fff_options = (
        (await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "fff_options": 1}) or {})
        .get("fff_options") or {}
    )
    loop = asyncio.get_event_loop()
    if fff_options.get("apply_template", True):
        meta_for_template = {
            "title": book.get("title") or "",
            "author": book.get("author") or "",
            "description": book.get("description") or "",
            "chapters": book.get("chapters") or 0,
            "rawExtendedMeta": (book.get("source_meta") or {}).get("rawExtendedMeta") or {},
        }
        raw = await loop.run_in_executor(
            None,
            apply_template_to_epub,
            raw,
            meta_for_template,
            book.get("source_url") or "",
        )

    # Allocate new book_id and persist the bytes
    new_book_id = f"book_{uuid.uuid4().hex[:12]}"
    new_epub_path = user_dir / f"{new_book_id}.epub"
    await _write_local_and_mirror_to_r2(
        new_epub_path, raw, user.user_id, new_book_id, ".epub",
    )

    # Try to extract fresh metadata (chapters/words) — non-fatal
    new_meta: Dict[str, Any] = {
        "title": book.get("title") or "Untitled",
        "author": book.get("author") or "Unknown",
        "description": book.get("description") or "",
        "language": book.get("language") or "en",
        "publisher": book.get("publisher") or "",
    }
    extracted_extra: Dict[str, Any] = {}
    try:
        ex = extract_epub_metadata(new_epub_path) or {}
        if ex.get("title"):
            new_meta["title"] = ex["title"]
        if ex.get("author"):
            new_meta["author"] = ex["author"]
        if ex.get("description"):
            new_meta["description"] = ex["description"]
        if ex.get("chapters"):
            extracted_extra["chapters"] = int(ex["chapters"])
        if ex.get("words"):
            extracted_extra["words"] = int(ex["words"])
    except Exception as e:
        logger.warning("upload_new_version metadata extract failed: %s", e)

    # Re-extract embedded URLs
    try:
        new_links = extract_urls_from_epub(new_epub_path) or []
        links_path = user_dir / f"{new_book_id}.links.txt"
        links_path.write_text(
            format_links_txt(new_meta["title"], new_meta["author"], new_links),
            encoding="utf-8",
        )
        links_count = len(new_links)
    except Exception as e:
        logger.warning("upload_new_version link extract failed: %s", e)
        links_count = 0

    now_dt = datetime.now(timezone.utc)
    now_iso = now_dt.isoformat()
    updated_shelf = _updated_shelf_name(now_dt)

    # 1) Insert the new book on the dated shelf
    new_doc = {
        "book_id": new_book_id,
        "user_id": user.user_id,
        "filename": _templated_filename(new_meta["title"], new_meta["author"], new_book_id),
        "title": new_meta["title"],
        "author": new_meta["author"],
        "description": new_meta["description"],
        "language": new_meta["language"],
        "publisher": new_meta["publisher"],
        "has_cover": book.get("has_cover", False),  # cover preserved separately if needed
        "category": updated_shelf,
        "fandom": book.get("fandom"),
        "series_name": book.get("series_name"),
        "series_index": book.get("series_index"),
        "tags": book.get("tags") or [],
        "confidence": book.get("confidence", 0.0),
        "classifier": "manual_upload",
        "size_bytes": len(raw),
        "links_count": links_count,
        "source_url": book.get("source_url"),
        "last_refreshed_at": now_iso,
        "manually_uploaded_at": now_iso,
        "replaces": old_book_id,
        "created_at": now_iso,
        **extracted_extra,
    }
    await db.books.insert_one(new_doc)

    # Register the dated shelf as a custom category
    await db.categories.update_one(
        {"user_id": user.user_id, "name": updated_shelf},
        {"$setOnInsert": {
            "user_id": user.user_id,
            "name": updated_shelf,
            "created_at": now_iso,
            "auto_created": True,
        }},
        upsert=True,
    )

    # 2) Archive the old book
    await db.books.update_one(
        {"book_id": old_book_id, "user_id": user.user_id},
        {"$set": {
            "category": OLD_STORIES_SHELF,
            "replaced_by": new_book_id,
            "replaced_at": now_iso,
        }},
    )

    # 3) Compute refresh_summary for the bell badge / email digest
    refresh_summary: Optional[Dict[str, Any]] = None
    try:
        old_epub_path = user_dir / f"{old_book_id}.epub"
        if old_epub_path.exists():
            old_chapters = await loop.run_in_executor(None, extract_chapters, old_epub_path)
            new_chapters = await loop.run_in_executor(None, extract_chapters, new_epub_path)
            d = diff_chapters(old_chapters, new_chapters)
            refresh_summary = {
                "chapters_added": d["summary"]["chapters_added"],
                "chapters_changed": d["summary"]["chapters_changed"],
                "chapters_removed": d["summary"]["chapters_removed"],
                "words_delta": d["summary"]["words_delta"],
                "first_changed_href": (d.get("first_changed_chapter") or {}).get("new_href", ""),
                "first_changed_title": (d.get("first_changed_chapter") or {}).get("title", ""),
                "first_changed_kind": (d.get("first_changed_chapter") or {}).get("kind", ""),
            }
    except Exception as e:
        logger.warning("upload_new_version diff failed for %s -> %s: %s", old_book_id, new_book_id, e)

    await db.books.update_one(
        {"book_id": new_book_id, "user_id": user.user_id},
        {"$set": {"refresh_summary": refresh_summary, "update_seen": False}},
    )

    return {
        "ok": True,
        "new_book_id": new_book_id,
        "old_book_id": old_book_id,
        "title": new_meta["title"],
        "updated_shelf": updated_shelf,
        "message": f'Saved as a new version in "{updated_shelf}". The previous copy moved to Old stories.',
    }



# ----------------------------------------------------------------------
# DUPLICATE RESOLUTION
# `POST /books/{id}/resolve-duplicate`, `POST /books/resolve-group`,
# `GET /library/duplicates`, and `GET /library/duplicates/count` were
# moved to `routes/duplicate_resolution.py` in the Phase 5D refactor.
# `OLD_STORIES_SHELF`, `_updated_shelf_name`, `_normalize_title_for_match`,
# `extract_chapters`, `diff_chapters`, and `extract_epub_metadata` stay
# here because the upload + refresh paths still use them; the new module
# imports them from this file.
# ----------------------------------------------------------------------
# ----------------------------------------------------------------------
# TRASH SHELF — extracted to routes/trash.py in Phase 2 refactor.
# See ``backend/routes/trash.py`` for /trash, /trash/restore/*,
# /trash/restore-all, /trash/empty, and the ``sweep_expired_trash``
# background helper (now imported by digest.py from routes.trash).
# ----------------------------------------------------------------------



# ----------------------------------------------------------------------
# RELATIONSHIPS / PAIRINGS — first-class browsable dimension
# ----------------------------------------------------------------------

@api_router.get("/relationships")
async def list_relationships(user: User = Depends(get_current_user)):
    """Every distinct relationship across the user's library, with counts."""
    pipeline = [
        {"$match": {"user_id": user.user_id, "category": {"$ne": TRASH_SHELF}, "relationships": {"$exists": True, "$ne": []}}},
        {"$unwind": "$relationships"},
        {"$group": {"_id": "$relationships", "count": {"$sum": 1}, "fandoms": {"$addToSet": "$fandom"}}},
        {"$sort": {"count": -1}},
    ]
    out = []
    async for r in db.books.aggregate(pipeline):
        out.append({
            "name": r["_id"],
            "count": r["count"],
            "fandoms": [f for f in (r.get("fandoms") or []) if f],
        })
    return {"relationships": out, "count": len(out)}


@api_router.post("/relationships/backfill")
async def backfill_relationships(user: User = Depends(get_current_user)):
    """Walk every book in the library and re-extract relationships from the
    EPUB metadata. Useful for libraries seeded before this feature shipped."""
    user_dir = STORAGE_DIR / user.user_id
    cursor = db.books.find(
        {"user_id": user.user_id, "category": {"$ne": TRASH_SHELF}},
        {"_id": 0, "book_id": 1, "description": 1, "relationships": 1},
    )
    updated = 0
    skipped = 0
    async for b in cursor:
        epub_path = user_dir / f"{b['book_id']}.epub"
        if not epub_path.exists():
            skipped += 1
            continue
        try:
            loop = asyncio.get_event_loop()
            meta = await loop.run_in_executor(None, extract_epub_metadata, epub_path)
            new_rels = meta.get("relationships") or []
            old_rels = b.get("relationships") or []
            if sorted(new_rels) != sorted(old_rels):
                await db.books.update_one(
                    {"book_id": b["book_id"], "user_id": user.user_id},
                    {"$set": {"relationships": new_rels}},
                )
                updated += 1
        except Exception as e:
            logger.warning("backfill_relationships failed for %s: %s", b.get("book_id"), e)
            skipped += 1
    return {"updated": updated, "skipped": skipped}


# ---------------------------------------------------------------------------
# "Polish my library" — bulk metadata cleanup (added 2026-06-16)
# ---------------------------------------------------------------------------
#
# Ride on the in-place EPUB rewrite we shipped earlier today. Scan a user's
# library for books with messy title / author fields ("Unknown", a raw
# filename, the book_id itself, …) and offer one-click corrections inferred
# from the file or the source URL. Always shows a preview first — the user
# accepts per-book before anything writes.

from utils.polish import suggest_polish, polishable_mongo_filter


class PolishApplyItem(BaseModel):
    book_id: str
    apply_title: bool = False
    apply_author: bool = False


class PolishApplyBody(BaseModel):
    items: List[PolishApplyItem]


@api_router.get("/books/polish/preview")
async def polish_preview(
    limit: int = 200,
    user: User = Depends(get_current_user),
):
    """Scan the user's library for polishable books and return a preview.

    ``limit`` caps the number of *suggestions* returned (not the number of
    books scanned) so the UI can paginate later if needed. Default 200 is
    plenty for the typical "click through and accept" workflow.
    """
    limit = max(1, min(limit, 500))
    candidate_cursor = db.books.find(
        {"user_id": user.user_id, **polishable_mongo_filter()},
        {
            "_id": 0,
            "book_id": 1,
            "title": 1,
            "author": 1,
            "source_url": 1,
            "fandom": 1,
        },
    ).limit(limit * 3)  # over-fetch a bit; suggest_polish filters again.

    suggestions: List[Dict[str, Any]] = []
    async for b in candidate_cursor:
        s = suggest_polish(b)
        if s:
            s["fandom"] = b.get("fandom")
            suggestions.append(s)
            if len(suggestions) >= limit:
                break

    # Total polishable count (best-effort upper bound — same Mongo filter).
    total_candidates = await db.books.count_documents(
        {"user_id": user.user_id, **polishable_mongo_filter()}
    )
    return {
        "suggestions": suggestions,
        "returned": len(suggestions),
        "candidates_scanned": total_candidates,
        "limit": limit,
    }


@api_router.post("/books/polish/apply")
async def polish_apply(
    body: PolishApplyBody,
    user: User = Depends(get_current_user),
):
    """Apply the user-accepted polish suggestions.

    For each item: re-runs `suggest_polish` server-side (defence in depth
    against tampered payloads), then updates DB + rewrites EPUB in place
    for whichever of title / author the caller opted in on.

    Reuses the same write path as ``PATCH /books/{book_id}`` so the EPUB
    metadata is always kept in sync with the DB.
    """
    if not body.items:
        return {"updated": 0, "skipped": 0, "details": []}

    book_ids = [it.book_id for it in body.items if (it.apply_title or it.apply_author)]
    if not book_ids:
        return {"updated": 0, "skipped": 0, "details": []}

    rows = await db.books.find(
        {"user_id": user.user_id, "book_id": {"$in": book_ids}},
        {"_id": 0, "book_id": 1, "title": 1, "author": 1, "source_url": 1},
    ).to_list(length=len(book_ids))
    by_id = {r["book_id"]: r for r in rows}

    items_by_id = {it.book_id: it for it in body.items}

    updated = 0
    skipped = 0
    details: List[Dict[str, Any]] = []
    for bid in book_ids:
        row = by_id.get(bid)
        opt = items_by_id.get(bid)
        if not row or not opt:
            skipped += 1
            details.append({"book_id": bid, "ok": False, "error": "not_found"})
            continue
        sug = suggest_polish(row)
        if not sug:
            # The book has been cleaned up in the meantime (or never qualified).
            skipped += 1
            details.append({"book_id": bid, "ok": False, "error": "no_longer_polishable"})
            continue
        update: Dict[str, Any] = {}
        if opt.apply_title and sug.get("suggested_title"):
            update["title"] = sug["suggested_title"][:500]
        if opt.apply_author and sug.get("suggested_author"):
            update["author"] = sug["suggested_author"][:500]
        if not update:
            skipped += 1
            details.append({"book_id": bid, "ok": False, "error": "nothing_to_apply"})
            continue

        await db.books.update_one(
            {"book_id": bid, "user_id": user.user_id},
            {"$set": update},
        )
        # Mirror into the EPUB file if it's on disk. DB always wins —
        # an EPUB write failure here is logged but doesn't fail the request.
        fp = STORAGE_DIR / user.user_id / f"{bid}.epub"
        epub_written: Optional[bool] = None
        if fp.exists():
            result = update_epub_metadata(
                fp,
                title=update.get("title"),
                author=update.get("author"),
            )
            epub_written = bool(result.get("ok"))
            if not epub_written:
                logger.info(
                    "polish_apply: EPUB write skipped for %s — %s",
                    bid, result.get("error"),
                )
        updated += 1
        details.append({"book_id": bid, "ok": True, "epub_written": epub_written, "applied": update})

    return {"updated": updated, "skipped": skipped, "details": details}



# ---------------------------------------------------------------------
# AI cover regeneration (2026-06-17)
# ---------------------------------------------------------------------
# Two-phase flow so the user previews before committing:
#   1. POST /api/books/{book_id}/preview-cover    → returns base64 PNG
#      + a short-lived preview_id stored in memory
#   2. POST /api/books/{book_id}/apply-cover      → persists the
#      previewed image to disk + flips has_cover on the doc
# Keeping the preview in memory (vs. on disk) means a rejected
# regeneration leaves zero artifacts.

from utils.cover_gen import generate_cover as _generate_cover

# Per-user preview cache.  Bounded to ~1 hr by storing the timestamp.
# Trimmed lazily on each apply / on backend restart.  In a multi-worker
# deploy we'd need Redis — single-worker is fine for now.
_COVER_PREVIEW_CACHE: Dict[str, Dict[str, Any]] = {}
_COVER_PREVIEW_TTL_SECONDS = 60 * 60  # 1 hour


class CoverPreviewBody(BaseModel):
    nudge: Optional[str] = None
    style_id: Optional[str] = None


@api_router.post("/books/{book_id}/preview-cover")
async def preview_book_cover(
    book_id: str,
    body: CoverPreviewBody,
    user: User = Depends(get_current_user),
):
    """Generate a cover image via nano-banana and return it as base64 PNG.
    Does NOT persist anything — the user must call ``apply-cover`` to
    keep it.  Holds the bytes in memory keyed by a preview_id so the
    apply step doesn't re-bill another generation."""
    doc = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "book_id": 1, "title": 1, "author": 1, "fandom": 1,
         "tags": 1, "description": 1, "summary": 1},
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Book not found")

    # Resolve style: built-in slug → prompt string, or custom user style.
    from utils.cover_styles import get_style_prompt as _get_built_in_style
    style_prompt = ""
    if body.style_id:
        style_prompt = _get_built_in_style(body.style_id)
        if not style_prompt and body.style_id.startswith("custom:"):
            custom_id = body.style_id.split(":", 1)[1]
            custom = await db.user_cover_styles.find_one(
                {"style_id": custom_id, "user_id": user.user_id},
                {"_id": 0, "prompt": 1},
            )
            if custom:
                style_prompt = custom.get("prompt", "")

    try:
        png_bytes, prompt = await _generate_cover(
            doc, nudge=body.nudge, style_prompt=style_prompt or None,
        )
    except Exception as e:  # noqa: BLE001 — surface as 502 with detail
        logger.exception("cover_gen failed for %s", book_id)
        raise HTTPException(status_code=502, detail=f"Cover generation failed: {e}")
    preview_id = uuid.uuid4().hex
    _COVER_PREVIEW_CACHE[preview_id] = {
        "book_id": book_id,
        "user_id": user.user_id,
        "png_bytes": png_bytes,
        "prompt": prompt,
        "created_at": datetime.now(timezone.utc),
    }
    # Trim expired entries opportunistically so the cache doesn't grow
    # unbounded if users abandon previews.
    now = datetime.now(timezone.utc)
    for pid, entry in list(_COVER_PREVIEW_CACHE.items()):
        if (now - entry["created_at"]).total_seconds() > _COVER_PREVIEW_TTL_SECONDS:
            _COVER_PREVIEW_CACHE.pop(pid, None)
    return {
        "preview_id": preview_id,
        "image_base64": base64.b64encode(png_bytes).decode("ascii"),
        "mime_type": "image/png",
    }


class CoverApplyBody(BaseModel):
    preview_id: str


@api_router.post("/books/{book_id}/apply-cover")
async def apply_book_cover(
    book_id: str,
    body: CoverApplyBody,
    user: User = Depends(get_current_user),
):
    """Persist the previewed cover bytes to disk + DB.  Stores up to
    ``_COVER_VARIANT_CAP`` historic variants per book so the user can
    switch back without re-paying for generation.

    Layout on disk:
        {user_dir}/{book_id}.cover             ← active cover (served)
        {user_dir}/{book_id}.cover-v-{nonce}   ← inactive variants

    Original EPUB is never touched.
    """
    entry = _COVER_PREVIEW_CACHE.get(body.preview_id)
    if not entry or entry["book_id"] != book_id or entry["user_id"] != user.user_id:
        raise HTTPException(status_code=404, detail="Preview expired or unknown")
    user_dir = STORAGE_DIR / user.user_id
    user_dir.mkdir(parents=True, exist_ok=True)

    # 1. Fetch the current variant list so we can append and FIFO-trim.
    book_doc = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "cover_variants": 1, "has_cover": 1},
    )
    if book_doc is None:
        raise HTTPException(status_code=404, detail="Book not found")
    variants = list(book_doc.get("cover_variants") or [])

    # 2. Mint a nonce for the new variant file.  Keeping a stable
    # filename on disk so the variant can be switched-to without
    # re-encoding.
    variant_id = uuid.uuid4().hex[:12]
    variant_filename = f"{book_id}.cover-v-{variant_id}"
    await _write_local_and_mirror_to_r2(
        user_dir / variant_filename,
        entry["png_bytes"],
        user.user_id, book_id, f".cover-v-{variant_id}",
    )

    now_iso = datetime.now(timezone.utc).isoformat()

    # 3. Mark every existing variant as inactive, append the new one,
    # FIFO-drop the oldest if we'd exceed the cap.
    for v in variants:
        v["active"] = False
    variants.append({
        "variant_id": variant_id,
        "file": variant_filename,
        "generated_at": now_iso,
        "active": True,
        "prompt": entry.get("prompt", "")[:500],  # truncate for storage
    })
    if len(variants) > _COVER_VARIANT_CAP:
        # FIFO-drop the oldest INACTIVE variant — the active one we just
        # added is at the end so trimming from the front is safe.
        dropped = variants[: len(variants) - _COVER_VARIANT_CAP]
        variants = variants[len(variants) - _COVER_VARIANT_CAP:]
        for d in dropped:
            try:
                (user_dir / d["file"]).unlink(missing_ok=True)
            except Exception:
                logger.exception("failed to unlink dropped cover variant %s", d.get("file"))

    # 4. Refresh the active "served" file to point at the new variant
    # bytes (we just write them again — cheap, ~1 MB, avoids symlink
    # quirks on some filesystems).
    await _write_local_and_mirror_to_r2(
        user_dir / f"{book_id}.cover",
        entry["png_bytes"],
        user.user_id, book_id, ".cover",
    )

    # 5. Persist.
    await db.books.update_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"$set": {
            "has_cover": True,
            "cover_source": "ai_generated",
            "cover_generated_at": now_iso,
            "cover_active_variant": variant_id,
            "cover_variants": variants,
        }},
    )
    _COVER_PREVIEW_CACHE.pop(body.preview_id, None)
    return {
        "ok": True,
        "book_id": book_id,
        "has_cover": True,
        "active_variant": variant_id,
        "variant_count": len(variants),
    }


# Per-book cap on stored cover variants.  Each variant is ~1 MB so 20
# variants × ~50 actively-iterated books = ~1 GB max disk footprint per
# heavy user.  20 is the right ceiling: enough to keep iterations + a
# handful of community-imported covers without growing unboundedly.
# Bumped from 5 → 20 on 2026-06-17.
_COVER_VARIANT_CAP = 20


@api_router.get("/books/{book_id}/cover-variants")
async def list_cover_variants(
    book_id: str,
    user: User = Depends(get_current_user),
):
    """Return every stored variant for this book + the active one's id.
    The thumbnails are base64'd inline because the page only needs them
    when the user explicitly opens the variants drawer — no need for a
    separate `/cover-variants/{id}.png` endpoint just for previews."""
    doc = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "cover_variants": 1, "cover_active_variant": 1},
    )
    if doc is None:
        raise HTTPException(status_code=404, detail="Book not found")
    variants = list(doc.get("cover_variants") or [])
    active_id = doc.get("cover_active_variant")
    user_dir = STORAGE_DIR / user.user_id
    out: List[Dict[str, Any]] = []
    for v in variants:
        path = user_dir / v["file"]
        if not path.exists():
            continue  # variant file was deleted on disk — skip silently
        out.append({
            "variant_id":   v["variant_id"],
            "generated_at": v["generated_at"],
            "active":       v["variant_id"] == active_id,
            "image_base64": base64.b64encode(path.read_bytes()).decode("ascii"),
            "mime_type":    "image/png",
        })
    return {"variants": out, "active_variant_id": active_id}


@api_router.post("/books/{book_id}/cover-variants/{variant_id}/activate")
async def activate_cover_variant(
    book_id: str,
    variant_id: str,
    user: User = Depends(get_current_user),
):
    """Switch the active cover to a previously generated variant."""
    doc = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "cover_variants": 1},
    )
    if doc is None:
        raise HTTPException(status_code=404, detail="Book not found")
    variants = list(doc.get("cover_variants") or [])
    target = next((v for v in variants if v["variant_id"] == variant_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Variant not found")
    user_dir = STORAGE_DIR / user.user_id
    src = user_dir / target["file"]
    if not src.exists():
        raise HTTPException(status_code=410, detail="Variant file missing on disk")
    await _write_local_and_mirror_to_r2(
        user_dir / f"{book_id}.cover",
        src.read_bytes(),
        user.user_id, book_id, ".cover",
    )
    for v in variants:
        v["active"] = (v["variant_id"] == variant_id)
    await db.books.update_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"$set": {
            "cover_active_variant": variant_id,
            "cover_variants": variants,
            "cover_generated_at": target["generated_at"],
        }},
    )
    return {"ok": True, "active_variant": variant_id}


@api_router.delete("/books/{book_id}/cover-variants/{variant_id}")
async def delete_cover_variant(
    book_id: str,
    variant_id: str,
    user: User = Depends(get_current_user),
):
    """Drop a stored variant.  Refuses to delete the currently active
    variant (the user must activate a different one first) so the book
    never ends up with `has_cover: True` and no file on disk."""
    doc = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "cover_variants": 1, "cover_active_variant": 1},
    )
    if doc is None:
        raise HTTPException(status_code=404, detail="Book not found")
    if doc.get("cover_active_variant") == variant_id:
        raise HTTPException(status_code=400, detail="Activate another variant first")
    variants = list(doc.get("cover_variants") or [])
    target = next((v for v in variants if v["variant_id"] == variant_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Variant not found")
    user_dir = STORAGE_DIR / user.user_id
    try:
        (user_dir / target["file"]).unlink(missing_ok=True)
    except Exception:
        logger.exception("failed to delete cover variant %s", target.get("file"))
    new_variants = [v for v in variants if v["variant_id"] != variant_id]
    await db.books.update_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"$set": {"cover_variants": new_variants}},
    )
    return {"ok": True, "remaining": len(new_variants)}




# ---------------------------------------------------------------------
# Community cover pool (2026-06-17)
# ---------------------------------------------------------------------
# Opt-in sharing: a user publishes one of their variants to a public
# pool keyed by normalized (title, author, fandom).  Other users
# importing the same fic can browse and adopt without re-paying for
# AI generation.  Bytes live in /app/uploads/community_covers/{id} so
# they're decoupled from per-user storage and survive user deletion.

_COMMUNITY_COVERS_DIR = STORAGE_DIR.parent / "community_covers"
_COMMUNITY_COVERS_DIR.mkdir(parents=True, exist_ok=True)


def _norm_book_key(title: str, author: str, fandom: str) -> Dict[str, str]:
    """Normalize the book identity for community pool lookups: lower,
    strip, collapse whitespace.  Two books that match on this triple
    are treated as the same work for cover-sharing purposes."""
    def clean(s):
        return " ".join((s or "").strip().lower().split())
    return {
        "title_key":  clean(title),
        "author_key": clean(author),
        "fandom_key": clean(fandom),
    }



# ---------------------------------------------------------------------
# Cover style catalog (2026-06-17 Tier 2)
# Built-in styles ship via utils/cover_styles.py; user-defined ones
# live in the `user_cover_styles` collection.
# ---------------------------------------------------------------------

from utils.cover_styles import built_in_list, style_exists


class CustomStyleBody(BaseModel):
    name: str
    prompt: str


@api_router.get("/cover-styles")
async def list_cover_styles(user: User = Depends(get_current_user)):
    """List every style available to the caller — built-in + custom.
    Custom styles return with ``id`` already prefixed ``custom:`` so the
    frontend can pass it back into ``preview-cover`` without
    transformation."""
    custom_rows = await db.user_cover_styles.find(
        {"user_id": user.user_id},
        {"_id": 0, "style_id": 1, "name": 1, "description": 1, "prompt": 1},
    ).sort("created_at", -1).to_list(length=50)
    customs = [
        {
            "id": f"custom:{r['style_id']}",
            "name": r["name"],
            "description": r.get("description") or r.get("prompt", "")[:80],
            "kind": "custom",
        }
        for r in custom_rows
    ]
    return {"styles": built_in_list() + customs}


@api_router.post("/cover-styles/custom")
async def create_custom_style(
    body: CustomStyleBody,
    user: User = Depends(get_current_user),
):
    """Save a named user-defined style.  Prompt is appended verbatim to
    the generation prompt so users can express anything they like.
    Capped at 20 customs per user so a runaway user can't bloat Mongo."""
    name = (body.name or "").strip()
    prompt = (body.prompt or "").strip()
    if not name or not prompt:
        raise HTTPException(status_code=400, detail="name and prompt are required")
    if len(prompt) > 1000:
        raise HTTPException(status_code=400, detail="prompt too long (max 1000 chars)")
    existing = await db.user_cover_styles.count_documents({"user_id": user.user_id})
    if existing >= 20:
        raise HTTPException(status_code=400, detail="Custom style cap reached (20).  Delete one first.")
    style_id = uuid.uuid4().hex[:12]
    await db.user_cover_styles.insert_one({
        "style_id":   style_id,
        "user_id":    user.user_id,
        "name":       name[:60],
        "prompt":     prompt,
        "description": prompt[:80],
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"ok": True, "id": f"custom:{style_id}", "name": name[:60]}


@api_router.delete("/cover-styles/custom/{style_id}")
async def delete_custom_style(
    style_id: str,
    user: User = Depends(get_current_user),
):
    """Remove a saved custom style.  No effect on covers already
    generated with it — those bytes are independent."""
    r = await db.user_cover_styles.delete_one(
        {"style_id": style_id, "user_id": user.user_id},
    )
    if r.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Custom style not found")
    return {"ok": True}



@api_router.post("/books/{book_id}/cover-variants/{variant_id}/share")
async def share_cover_to_community(
    book_id: str,
    variant_id: str,
    user: User = Depends(get_current_user),
):
    """Publish a variant to the public community cover pool.  Idempotent:
    re-sharing the same variant returns the existing community_cover_id."""
    doc = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "title": 1, "author": 1, "fandom": 1, "cover_variants": 1},
    )
    if doc is None:
        raise HTTPException(status_code=404, detail="Book not found")
    target = next(
        (v for v in (doc.get("cover_variants") or []) if v["variant_id"] == variant_id),
        None,
    )
    if not target:
        raise HTTPException(status_code=404, detail="Variant not found")
    src = STORAGE_DIR / user.user_id / target["file"]
    if not src.exists():
        raise HTTPException(status_code=410, detail="Variant file missing on disk")

    # Dedupe on (variant_id, sharer) so re-clicking is a no-op.
    existing = await db.community_covers.find_one(
        {"source_variant_id": variant_id, "shared_by_user_id": user.user_id},
        {"_id": 0, "cover_id": 1},
    )
    if existing:
        return {"ok": True, "community_cover_id": existing["cover_id"], "deduped": True}

    cover_id = uuid.uuid4().hex[:14]
    dest = _COMMUNITY_COVERS_DIR / cover_id
    dest.write_bytes(src.read_bytes())

    # Pull the sharer's username for attribution.  Fallback to email
    # local-part if they don't have a public handle yet.
    user_doc = await db.users.find_one(
        {"user_id": user.user_id},
        {"_id": 0, "username": 1, "email": 1, "name": 1},
    ) or {}
    shared_by = (
        user_doc.get("username")
        or (user_doc.get("email") or "").split("@", 1)[0]
        or "anon"
    )

    keys = _norm_book_key(doc.get("title", ""), doc.get("author", ""), doc.get("fandom", ""))
    # If this variant was originally imported from another community
    # cover, remember the lineage so the remix tree is visible.
    parent_cover_id = ""
    src_tag = (target.get("source") or "") if isinstance(target, dict) else ""
    if src_tag.startswith("community:"):
        parent_cover_id = src_tag.split(":", 1)[1]
    record = {
        "cover_id": cover_id,
        **keys,
        "title":             doc.get("title", ""),
        "author":            doc.get("author", ""),
        "fandom":            doc.get("fandom", ""),
        "file":              cover_id,                       # bare filename in community_covers/
        "source_book_id":    book_id,
        "source_variant_id": variant_id,
        "parent_cover_id":   parent_cover_id,
        "shared_by_user_id": user.user_id,
        "shared_by_username": shared_by,
        "shared_at":         datetime.now(timezone.utc).isoformat(),
        "import_count":      0,
    }
    await db.community_covers.insert_one(record)
    # Fan-out to friends — single in-app ping per friend, quiet failure.
    try:
        await notify_friends_of_new_share(
            cover_id=cover_id,
            sharer_user_id=user.user_id,
            title=doc.get("title", ""),
        )
    except Exception as e:
        logger.exception("notify_friends_of_new_share fan-out errored: %s", e)
    return {"ok": True, "community_cover_id": cover_id, "shared_by": shared_by}


@api_router.get("/community-covers")
async def browse_community_covers(
    title: str,
    author: str = "",
    fandom: str = "",
    limit: int = 24,
    user: User = Depends(get_current_user),
):
    """Browse community-shared covers for a given (title, author, fandom).
    Matching is case-insensitive on title (required); author + fandom are
    refinements that narrow when supplied.  Returns thumbnails inline
    so the frontend can render a grid in one round-trip."""
    limit = max(1, min(int(limit), 60))
    keys = _norm_book_key(title, author, fandom)
    if not keys["title_key"]:
        raise HTTPException(status_code=400, detail="title query is required")
    query: Dict[str, Any] = {"title_key": keys["title_key"]}
    if keys["author_key"]:
        query["author_key"] = keys["author_key"]
    if keys["fandom_key"]:
        query["fandom_key"] = keys["fandom_key"]
    cursor = (
        db.community_covers.find(query, {"_id": 0})
        .sort([("votes", -1), ("import_count", -1), ("shared_at", -1)])
        .limit(limit)
    )
    rows: List[Dict[str, Any]] = []
    async for r in cursor:
        path = _COMMUNITY_COVERS_DIR / r["file"]
        if not path.exists():
            continue
        rows.append({
            "cover_id":     r["cover_id"],
            "shared_by":    r.get("shared_by_username", "anon"),
            "shared_at":    r.get("shared_at"),
            "import_count": int(r.get("import_count", 0)),
            "votes":        int(r.get("votes", 0)),
            "voted_by_me":  user.user_id in (r.get("voters") or []),
            "image_base64": base64.b64encode(path.read_bytes()).decode("ascii"),
            "mime_type":    "image/png",
        })
    return {"covers": rows, "count": len(rows)}


@api_router.post("/books/{book_id}/import-community-cover/{cover_id}")
async def import_community_cover(
    book_id: str,
    cover_id: str,
    user: User = Depends(get_current_user),
):
    """Adopt a community cover as a new variant for the caller's book.
    Doesn't re-pay any LLM cost — bytes copied directly.  Increments
    the community cover's import_count for popularity sorting."""
    src = _COMMUNITY_COVERS_DIR / cover_id
    if not src.exists():
        raise HTTPException(status_code=404, detail="Community cover not found")
    book = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "cover_variants": 1},
    )
    if book is None:
        raise HTTPException(status_code=404, detail="Book not found")
    user_dir = STORAGE_DIR / user.user_id
    user_dir.mkdir(parents=True, exist_ok=True)
    variant_id = uuid.uuid4().hex[:12]
    variant_filename = f"{book_id}.cover-v-{variant_id}"
    bytes_ = src.read_bytes()
    await _write_local_and_mirror_to_r2(
        user_dir / variant_filename,
        bytes_,
        user.user_id, book_id, f".cover-v-{variant_id}",
    )
    await _write_local_and_mirror_to_r2(
        user_dir / f"{book_id}.cover",
        bytes_,
        user.user_id, book_id, ".cover",
    )
    now_iso = datetime.now(timezone.utc).isoformat()

    variants = list(book.get("cover_variants") or [])
    for v in variants:
        v["active"] = False
    variants.append({
        "variant_id":   variant_id,
        "file":         variant_filename,
        "generated_at": now_iso,
        "active":       True,
        "source":       f"community:{cover_id}",
    })
    # FIFO cap (same logic as apply-cover).
    if len(variants) > _COVER_VARIANT_CAP:
        dropped = variants[: len(variants) - _COVER_VARIANT_CAP]
        variants = variants[len(variants) - _COVER_VARIANT_CAP:]
        for d in dropped:
            try:
                (user_dir / d["file"]).unlink(missing_ok=True)
            except Exception:
                logger.exception("failed to unlink dropped variant %s", d.get("file"))

    await db.books.update_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"$set": {
            "has_cover": True,
            "cover_source": "community_imported",
            "cover_generated_at": now_iso,
            "cover_active_variant": variant_id,
            "cover_variants": variants,
        }},
    )
    await db.community_covers.update_one(
        {"cover_id": cover_id},
        {"$inc": {"import_count": 1}},
    )
    # Read back the new import count so we can fire a milestone ping
    # if this import crossed a 1/5/10/25 boundary.
    try:
        post = await db.community_covers.find_one(
            {"cover_id": cover_id},
            {"_id": 0, "import_count": 1},
        )
        if post:
            await notify_import_milestone(
                cover_id=cover_id,
                new_import_count=int(post.get("import_count", 0)),
            )
    except Exception as e:
        logger.exception("notify_import_milestone errored: %s", e)
    return {"ok": True, "variant_id": variant_id, "active_variant": variant_id}


@api_router.delete("/community-covers/{cover_id}")
async def unshare_community_cover(
    cover_id: str,
    user: User = Depends(get_current_user),
):
    """Remove a cover from the community pool — only the original sharer
    (or an admin / moderator) may unshare.  Doesn't touch any users who
    already imported the cover; those copies remain in their libraries."""
    record = await db.community_covers.find_one(
        {"cover_id": cover_id},
        {"_id": 0, "shared_by_user_id": 1, "file": 1},
    )
    if not record:
        raise HTTPException(status_code=404, detail="Community cover not found")
    if (
        record["shared_by_user_id"] != user.user_id
        and not getattr(user, "is_admin", False)
        and not getattr(user, "is_moderator", False)
    ):
        raise HTTPException(status_code=403, detail="Not your cover to unshare")
    try:
        (_COMMUNITY_COVERS_DIR / record["file"]).unlink(missing_ok=True)
    except Exception:
        logger.exception("failed to unlink community cover %s", record["file"])
    await db.community_covers.delete_one({"cover_id": cover_id})
    return {"ok": True}


# ---------------------------------------------------------------------
# Community covers — voting + featured (Tier 3, 2026-06-17)
# ---------------------------------------------------------------------

@api_router.post("/community-covers/{cover_id}/vote")
async def vote_community_cover(
    cover_id: str,
    user: User = Depends(get_current_user),
):
    """Toggle the caller's heart on a community cover.  One vote per
    user — re-voting un-hearts it.  Voters are tracked in a per-cover
    list so we can show the user their current state on browse."""
    record = await db.community_covers.find_one(
        {"cover_id": cover_id},
        {"_id": 0, "voters": 1, "votes": 1},
    )
    if record is None:
        raise HTTPException(status_code=404, detail="Community cover not found")
    voters = set(record.get("voters") or [])
    if user.user_id in voters:
        voters.discard(user.user_id)
        action = "unvoted"
    else:
        voters.add(user.user_id)
        action = "voted"
    new_count = len(voters)
    await db.community_covers.update_one(
        {"cover_id": cover_id},
        {"$set": {"voters": list(voters), "votes": new_count}},
    )
    # Milestone notification — only fires on exact thresholds so the
    # toggle path doesn't re-fire.
    if action == "voted":
        try:
            await notify_vote_milestone(
                cover_id=cover_id,
                new_vote_count=new_count,
                voter_user_id=user.user_id,
            )
        except Exception as e:
            logger.exception("notify_vote_milestone errored: %s", e)
    return {"ok": True, "votes": new_count, "voted_by_me": action == "voted"}


@api_router.get("/community-covers/featured")
async def featured_community_covers(
    limit: int = 12,
    days: int = 7,
    user: Optional[User] = Depends(get_current_user_or_none),
):
    """Top-voted community covers across the whole pool, lifetime
    (``days=0``) or scoped to the recent ``days`` window.  Used by a
    "Featured this week" homepage strip + the dedicated discovery page.

    Each row carries a ``trending`` flag — covers that have collected
    ≥3 hearts AND were shared within the last 48 h.  These cards get a
    distinct pill on the homepage strip so high-velocity covers are
    visible even when a long-running cover holds the cumulative #1."""
    limit = max(1, min(int(limit), 60))
    days = max(0, min(int(days), 365))
    query: Dict[str, Any] = {}
    if days > 0:
        cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
        query["shared_at"] = {"$gte": cutoff}
    cursor = (
        db.community_covers.find(query, {"_id": 0})
        .sort([("votes", -1), ("import_count", -1), ("shared_at", -1)])
        .limit(limit)
    )
    trending_cutoff = (datetime.now(timezone.utc) - timedelta(hours=48)).isoformat()
    rows: List[Dict[str, Any]] = []
    async for r in cursor:
        path = _COMMUNITY_COVERS_DIR / r["file"]
        if not path.exists():
            continue
        votes = int(r.get("votes", 0))
        shared_at = r.get("shared_at") or ""
        rows.append({
            "cover_id":     r["cover_id"],
            "title":        r.get("title", ""),
            "author":       r.get("author", ""),
            "fandom":       r.get("fandom", ""),
            "shared_by":    r.get("shared_by_username", "anon"),
            "shared_by_user_id": r.get("shared_by_user_id", ""),
            "votes":        votes,
            "import_count": int(r.get("import_count", 0)),
            "voted_by_me":  bool(user) and user.user_id in (r.get("voters") or []),
            "trending":     votes >= 3 and shared_at >= trending_cutoff,
            "parent_cover_id": r.get("parent_cover_id", ""),
            "image_base64": base64.b64encode(path.read_bytes()).decode("ascii"),
            "mime_type":    "image/png",
        })
    return {"covers": rows, "window_days": days}


# ---------------------------------------------------------------------
# Cover ecosystem Tier 4 — public profile, achievements, lineage
# (2026-06-18)
# ---------------------------------------------------------------------

@api_router.get("/users/{username}/cover-profile")
async def public_cover_profile(
    username: str,
    user: Optional[User] = Depends(get_current_user_or_none),  # noqa: ARG001 — public
):
    """Public-but-auth-gated profile page powering `/u/{username}`.
    Surfaces lifetime cover stats, the trophies the user has unlocked
    (e.g. ``top_of_week``), and a feed of their best community
    covers.  Returns 404 for unknown / username-less accounts so the
    URL doesn't leak existence of private users."""
    target = await db.users.find_one(
        {"username": username},
        {
            "_id": 0, "user_id": 1, "username": 1, "name": 1,
            "cover_achievements": 1, "picture": 1, "created_at": 1,
        },
    )
    if target is None:
        raise HTTPException(status_code=404, detail="No such user")
    cursor = (
        db.community_covers.find(
            {"shared_by_user_id": target["user_id"]},
            {"_id": 0},
        )
        .sort([("votes", -1), ("import_count", -1), ("shared_at", -1)])
        .limit(48)
    )
    covers: List[Dict[str, Any]] = []
    lifetime_votes = 0
    lifetime_imports = 0
    total_shared = 0
    async for r in cursor:
        path = _COMMUNITY_COVERS_DIR / r["file"]
        if not path.exists():
            continue
        votes = int(r.get("votes", 0))
        imports = int(r.get("import_count", 0))
        lifetime_votes += votes
        lifetime_imports += imports
        total_shared += 1
        covers.append({
            "cover_id":     r["cover_id"],
            "title":        r.get("title", ""),
            "author":       r.get("author", ""),
            "fandom":       r.get("fandom", ""),
            "votes":        votes,
            "import_count": imports,
            "shared_at":    r.get("shared_at"),
            "image_base64": base64.b64encode(path.read_bytes()).decode("ascii"),
            "mime_type":    "image/png",
        })
    achievements = target.get("cover_achievements") or []
    return {
        "username":       target.get("username"),
        "display_name":   target.get("name") or target.get("username"),
        "picture":        target.get("picture", ""),
        "joined_at":      target.get("created_at"),
        "covers":         covers,
        "achievements":   achievements,
        "totals": {
            "shared":  total_shared,
            "votes":   lifetime_votes,
            "imports": lifetime_imports,
        },
    }


@api_router.get("/community-covers/{cover_id}/lineage")
async def community_cover_lineage(
    cover_id: str,
    user: Optional[User] = Depends(get_current_user_or_none),  # noqa: ARG001 — public
):
    """Returns the parent (if this cover was imported + re-shared) and
    the direct children (covers downstream that name this one as
    parent).  Used by the cover-card UI to surface "Remixed from @x"
    and "Remixed 3 times" badges so the community can see lineage."""
    cur = await db.community_covers.find_one(
        {"cover_id": cover_id},
        {"_id": 0, "cover_id": 1, "title": 1, "parent_cover_id": 1, "shared_by_username": 1},
    )
    if cur is None:
        raise HTTPException(status_code=404, detail="Community cover not found")
    parent = None
    pid = cur.get("parent_cover_id")
    if pid:
        parent = await db.community_covers.find_one(
            {"cover_id": pid},
            {"_id": 0, "cover_id": 1, "title": 1, "shared_by_username": 1},
        )
    children_cursor = db.community_covers.find(
        {"parent_cover_id": cover_id},
        {"_id": 0, "cover_id": 1, "title": 1, "shared_by_username": 1, "votes": 1},
    ).sort("shared_at", -1).limit(20)
    children = [c async for c in children_cursor]
    return {
        "cover_id": cover_id,
        "parent":   parent,
        "children": children,
        "remix_count": len(children),
    }



# ---------------------------------------------------------------------
# "Polish my covers" — list cover-less books so the frontend can run
# the existing preview-cover / apply-cover flow against each one
# without polling every book in the library client-side.
# ---------------------------------------------------------------------

@api_router.get("/books/cover-less")
async def list_cover_less_books(
    limit: int = 100,
    user: User = Depends(get_current_user),
):
    """Return books owned by the caller that don't have a cover yet.

    Result is intentionally small (id + title + author + fandom + tags)
    — the bulk page renders them as placeholder tiles and lets the user
    kick off cover generation per-book.  Cap at 200 because generating
    a cover is a paid LLM call; we don't want the UI promising "all
    300" if it'll cost the user too much.
    """
    limit = max(1, min(int(limit), 200))
    cursor = db.books.find(
        {
            "user_id": user.user_id,
            "$or": [
                {"has_cover": {"$exists": False}},
                {"has_cover": False},
                {"has_cover": None},
            ],
        },
        {"_id": 0, "book_id": 1, "title": 1, "author": 1, "fandom": 1,
         "tags": 1, "category": 1},
    ).sort("title", 1).limit(limit)
    rows = await cursor.to_list(length=limit)
    total = await db.books.count_documents({
        "user_id": user.user_id,
        "$or": [
            {"has_cover": {"$exists": False}},
            {"has_cover": False},
            {"has_cover": None},
        ],
    })
    return {"books": rows, "total": total, "limit": limit}
