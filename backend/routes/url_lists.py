"""URL-list helpers — paste a blob of fanfic URLs, dedupe against the
user's library, and either return the trimmed list, export as XLSX, or pull
every URL into the library sequentially.  Extracted from ``routes/books.py``
as part of the Phase 5 refactor (2026-06-14).

Routes:
    POST /api/books/url-list/dedupe
    POST /api/books/url-list/export-xlsx
    POST /api/books/url-list/pull
"""
import asyncio
import io
import os
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import Depends, HTTPException, Response
from pydantic import BaseModel

from deps import db, api_router, logger
from models import User
from auth_dep import get_current_user
from utils.constants import TRASH_SHELF
from routes.duplicate_resolution import _parse_urls_from_sidecar  # noqa: E402
from routes.books import (
    OLD_STORIES_SHELF,
    STORAGE_DIR,
    FanficNotFoundError,
    apply_template_to_epub,
    classify_book,
    extract_chapters,
    extract_epub_metadata,
    extract_fanfic_urls,
    extract_urls_from_epub,
    fetch_fanfic_with_fallback,
    find_duplicate_candidates,
    find_source_url,
    format_links_txt,
    normalize_fanfic_url,
)


# ---------------------------------------------------------------------------
# URL-list dedupe — treats a .txt of fanfic URLs as a wishlist and strips out
# URLs that already correspond to books in the user's library.
# ---------------------------------------------------------------------------

from utils.url_canonical import (  # noqa: E402
    _URL_RE,
    _canonical_fanfic_url,
    _is_ao3_host,
    _looks_like_url_list,
    classify_ao3_non_work,
)


class UrlListBody(BaseModel):
    text: str


@api_router.post("/books/url-list/dedupe")
async def dedupe_url_list_endpoint(body: UrlListBody, user: User = Depends(get_current_user)):
    """Dedupe a list of URLs pasted/typed in by the user. Same logic as the
    upload-time path but without a file in between."""
    if not body.text or not body.text.strip():
        raise HTTPException(status_code=400, detail="No URL text provided")
    return await _dedupe_url_list(body.text, user.user_id)


async def _backfill_user_fanfic_urls(user_id: str, limit: int = 2000) -> int:
    """Repopulate `fanfic_urls` and `source_url` for any of the user's books
    that are missing the fields, using the on-disk `.links.txt` sidecar.

    Idempotent — books that already have a non-empty `fanfic_urls` are
    skipped. Returns the count of records actually updated.
    """
    user_dir = STORAGE_DIR / user_id
    if not user_dir.exists():
        return 0
    cursor = db.books.find(
        {
            "user_id": user_id,
            "$or": [
                {"fanfic_urls": {"$exists": False}},
                {"fanfic_urls": []},
                {"source_url": {"$exists": False}},
                {"source_url": None},
            ],
        },
        {"_id": 0, "book_id": 1, "fanfic_urls": 1, "source_url": 1},
    )
    updated = 0
    async for b in cursor:
        if updated >= limit:
            break
        sidecar = user_dir / f"{b['book_id']}.links.txt"
        raw_urls = _parse_urls_from_sidecar(sidecar)
        if not raw_urls:
            # Mark with an empty list so the next sweep skips this book.
            if b.get("fanfic_urls") is None or "fanfic_urls" not in b:
                await db.books.update_one(
                    {"book_id": b["book_id"], "user_id": user_id},
                    {"$set": {"fanfic_urls": []}},
                )
            continue
        canon: List[str] = []
        seen: set = set()
        for u in raw_urls:
            c = normalize_fanfic_url(u)
            if c and c not in seen:
                seen.add(c)
                canon.append(c)
        patch: Dict[str, Any] = {}
        existing = b.get("fanfic_urls") or []
        if canon and canon != existing:
            patch["fanfic_urls"] = canon
        if canon and not b.get("source_url"):
            patch["source_url"] = canon[0]
        if "fanfic_urls" not in b and not canon:
            patch["fanfic_urls"] = []
        if patch:
            await db.books.update_one(
                {"book_id": b["book_id"], "user_id": user_id},
                {"$set": patch},
            )
            updated += 1
    return updated


async def _dedupe_url_list(text: str, user_id: str) -> Dict[str, Any]:
    """Walk every URL in `text`, dedupe against the user's library.

    Returns `{total, already_owned, new_urls, unrecognized, by_source,
    ao3_non_work, duplicate_in_list}`.
    """
    raw_urls: List[str] = []
    seen: set = set()
    for m in _URL_RE.finditer(text):
        url = m.group(0).rstrip('.,);]>')
        if url in seen:
            continue
        seen.add(url)
        raw_urls.append(url)

    # Bucket each URL:
    #   * fanfic permalink → check ownership
    #   * AO3 non-work link (series / collection / user) → separate bucket
    #     (the user pasted it; we want to surface that we saw it but it's
    #     not a story to dedupe against)
    #   * everything else → "unrecognized"
    canonical_pairs: List[Dict[str, str]] = []
    canonical_first_seen: Dict[str, int] = {}
    duplicate_in_list: List[Dict[str, str]] = []
    unrecognized: List[str] = []
    ao3_non_work: List[Dict[str, str]] = []
    for url in raw_urls:
        canonical = _canonical_fanfic_url(url)
        if canonical:
            if canonical in canonical_first_seen:
                duplicate_in_list.append({"url": url, "canonical": canonical})
            else:
                canonical_first_seen[canonical] = len(canonical_pairs)
                canonical_pairs.append({"url": url, "canonical": canonical})
            continue
        non_work = classify_ao3_non_work(url)
        if non_work:
            ao3_non_work.append({"url": url, "kind": non_work})
            continue
        unrecognized.append(url)

    canonicals = sorted({p["canonical"] for p in canonical_pairs})
    # Look up everything in one Mongo round-trip
    owned_map: Dict[str, Dict[str, Any]] = {}
    if canonicals:
        # Opportunistic backfill: many books (especially older uploads) were
        # saved without `fanfic_urls` / `source_url` populated. Before the
        # match query, walk the user's books that are missing the field and
        # repopulate from the on-disk `.links.txt` sidecar so the match
        # actually works. Capped at 2000 sidecars per request — well above
        # any realistic library size for a single dedupe pass.
        await _backfill_user_fanfic_urls(user_id, limit=2000)

        cursor = db.books.find(
            {
                "user_id": user_id,
                "category": {"$ne": TRASH_SHELF},
                "$or": [
                    {"fanfic_urls": {"$in": canonicals}},
                    {"source_url": {"$in": canonicals}},
                ],
            },
            {"_id": 0, "book_id": 1, "title": 1, "author": 1, "fanfic_urls": 1, "source_url": 1, "category": 1, "fandom": 1},
        )
        async for b in cursor:
            for u in (b.get("fanfic_urls") or []):
                if u in canonicals and u not in owned_map:
                    owned_map[u] = b
            if b.get("source_url") in canonicals and b["source_url"] not in owned_map:
                owned_map[b["source_url"]] = b

    already_owned: List[Dict[str, Any]] = []
    new_urls: List[Dict[str, str]] = []
    for p in canonical_pairs:
        match = owned_map.get(p["canonical"])
        if match:
            already_owned.append({
                "url": p["url"],
                "canonical": p["canonical"],
                "book_id": match["book_id"],
                "title": match.get("title") or "",
                "author": match.get("author") or "",
                "fandom": match.get("fandom") or "",
                "category": match.get("category") or "",
            })
        else:
            new_urls.append({"url": p["url"], "canonical": p["canonical"]})

    # Per-source breakdown across the whole list (including unrecognized).
    by_source: Dict[str, int] = {}
    for p in canonical_pairs:
        src = _source_for(p["canonical"]) or "Other"
        by_source[src] = by_source.get(src, 0) + 1
    for d in duplicate_in_list:
        src = _source_for(d["canonical"]) or "Other"
        by_source[src] = by_source.get(src, 0) + 1
    if ao3_non_work:
        by_source["AO3 (not a story)"] = len(ao3_non_work)
    if unrecognized:
        by_source["Unrecognized"] = len(unrecognized)

    # AO3 mirror detection — surface a friendly heads-up when the user
    # pasted URLs from a non-`.org` AO3 hostname (archiveofourown.com/net/gay,
    # ao3.org, archive.transformativeworks.org, insecure.archiveofourown.org).
    # All of them serve the same archive but the heads-up is useful when a
    # user wonders why every mirror dedupes to the same work.
    ao3_mirror_hosts: Dict[str, int] = {}
    for url_str in raw_urls:
        lower = url_str.lower()
        if not _is_ao3_host(lower):
            continue
        # Skip the canonical host — only surface the alt mirrors.
        if "archiveofourown.org" in lower and "insecure." not in lower:
            continue
        # Pull out just the hostname for the chip label.
        try:
            host = lower.split("://", 1)[1].split("/", 1)[0]
        except IndexError:
            continue
        ao3_mirror_hosts[host] = ao3_mirror_hosts.get(host, 0) + 1

    # Unknown-source recorder: log unrecognized URLs that LOOK story-shaped
    # (eFiction-style query string, forum thread, `/works/N`, `/s/N` etc.)
    # so we can surface new fic archives the accepted-list should be
    # extended to cover. Returns the de-duplicated list of hosts captured.
    from utils.unknown_sources import record_unknown_sources
    unknown_sources_found = await record_unknown_sources(
        db, unrecognized, context="paste", user_id=user_id,
    )

    return {
        "total": len(raw_urls),
        "already_owned": already_owned,
        "new_urls": new_urls,
        "unrecognized": unrecognized,
        "ao3_non_work": ao3_non_work,
        "duplicate_in_list": duplicate_in_list,
        "by_source": by_source,
        "ao3_mirrors": ao3_mirror_hosts,
        "unknown_sources_found": unknown_sources_found,
    }


class UrlListExportBody(BaseModel):
    urls: List[str]  # net-new URLs (not already owned)
    # Optional: when the frontend has the already-owned list, ship it so the
    # workbook contains BOTH sheets in one file. Each row is a {url, title?,
    # author?, book_id?} dict so we can show metadata for owned items.
    owned: Optional[List[Dict[str, Any]]] = None
    # Optional: surface forms of canonical URLs that appeared more than once
    # in the pasted text — e.g. the user pasted both
    # `/works/12345` and `/works/12345/chapters/9`. Each row is
    # {url, canonical}. Useful when auditing where AO3 link variants are
    # coming from.
    duplicates: Optional[List[Dict[str, Any]]] = None


from utils.url_canonical import _source_for  # noqa: E402


@api_router.post("/books/url-list/export-xlsx")
async def export_url_list_xlsx(body: UrlListExportBody, user: User = Depends(get_current_user)):
    """Build an Excel summarizing a URL-list dedupe pass.

    Sheet 1 — "New URLs": the net-new fanfic URLs (not in the library)
    Sheet 2 — "Already owned": URLs already in the library, with the matched
              book's title / author / id so the user can review what they
              already have without leaving the spreadsheet.
    Sheet 3 — "Duplicate pastes" (when present): surface forms of canonical
              URLs that appeared more than once in the pasted text — e.g.
              `/works/12345` and `/works/12345/chapters/9` both pointing to
              the same AO3 work. Helps audit messy bookmark dumps.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="6B46C1")

    # Sheet 1 — net-new
    ws_new = wb.active
    ws_new.title = "New URLs"
    ws_new.append(["URL", "Canonical", "Source"])
    for cell in ws_new[1]:
        cell.font = head_font
        cell.fill = head_fill
    for u in body.urls:
        c = _canonical_fanfic_url(u) or ""
        ws_new.append([u, c, _source_for(u)])
    for col, width in (("A", 60), ("B", 60), ("C", 18)):
        ws_new.column_dimensions[col].width = width
    ws_new.freeze_panes = "A2"
    if ws_new.max_row > 1:
        ws_new.auto_filter.ref = ws_new.dimensions

    # Sheet 2 — already owned
    ws_owned = wb.create_sheet("Already owned")
    ws_owned.append(["URL", "Title", "Author", "Book ID", "Source"])
    for cell in ws_owned[1]:
        cell.font = head_font
        cell.fill = head_fill
    for item in (body.owned or []):
        url = item.get("url") or ""
        ws_owned.append([
            url,
            item.get("title") or "",
            item.get("author") or "",
            item.get("book_id") or "",
            _source_for(url),
        ])
    for col, width in (("A", 60), ("B", 40), ("C", 24), ("D", 18), ("E", 18)):
        ws_owned.column_dimensions[col].width = width
    ws_owned.freeze_panes = "A2"
    if ws_owned.max_row > 1:
        ws_owned.auto_filter.ref = ws_owned.dimensions

    # Sheet 3 — duplicate pastes (only emitted when present so existing
    # workflows that don't pass `duplicates` see no change).
    if body.duplicates:
        ws_dups = wb.create_sheet("Duplicate pastes")
        ws_dups.append(["URL pasted", "Canonical", "Source"])
        for cell in ws_dups[1]:
            cell.font = head_font
            cell.fill = head_fill
        for item in body.duplicates:
            url = item.get("url") or ""
            ws_dups.append([
                url,
                item.get("canonical") or "",
                _source_for(url),
            ])
        for col, width in (("A", 60), ("B", 60), ("C", 18)):
            ws_dups.column_dimensions[col].width = width
        ws_dups.freeze_panes = "A2"
        if ws_dups.max_row > 1:
            ws_dups.auto_filter.ref = ws_dups.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    payload = buf.getvalue()
    return Response(
        content=payload,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="shelfsort_url_list.xlsx"',
            "Content-Length": str(len(payload)),
        },
    )


class UrlListPullBody(BaseModel):
    """Body for the `/api/books/url-list/pull` endpoint. Either pass `urls`
    directly (post-dedupe) or `text` to re-extract URLs from a pasted blob.
    """
    urls: Optional[List[str]] = None
    text: Optional[str] = None


@api_router.post("/books/url-list/pull")
async def pull_url_list(body: UrlListPullBody, user: User = Depends(get_current_user)):
    """Fetch every URL in the list sequentially and add the resulting
    EPUBs to the user's library.

    Strictly one URL at a time — both the FanFicFare path and the FicHub
    fallback path go through their own single-flight locks/sleeps, so
    even if the user pastes 200 URLs we never blast any source site.
    Already-owned URLs (matched against `fanfic_urls`/`source_url`) are
    skipped without a network hit.
    """
    # 1) Pull the URL list. Either explicit `urls` or extract from `text`.
    raw_urls: List[str] = []
    seen: set = set()
    if body.urls:
        for u in body.urls:
            if u and u not in seen:
                seen.add(u)
                raw_urls.append(u)
    elif body.text:
        for m in _URL_RE.finditer(body.text):
            url = m.group(0).rstrip('.,);]>')
            if url and url not in seen:
                seen.add(url)
                raw_urls.append(url)

    # 2) Canonicalize each and skip anything that isn't a recognized fanfic
    #    permalink (also drops AO3 series / collection / user pages).
    canonicals: List[str] = []
    canon_seen: set = set()
    unrecognized: List[str] = []
    for u in raw_urls:
        c = _canonical_fanfic_url(u)
        if c:
            if c not in canon_seen:
                canon_seen.add(c)
                canonicals.append(c)
        else:
            unrecognized.append(u)

    if not canonicals:
        return {
            "queued": 0,
            "added": [],
            "already_owned": [],
            "failed": [],
            "unrecognized": unrecognized,
        }

    # 3) Skip canonical URLs the user already has on a shelf (not Trash).
    owned_cursor = db.books.find(
        {
            "user_id": user.user_id,
            "category": {"$ne": TRASH_SHELF},
            "$or": [
                {"fanfic_urls": {"$in": canonicals}},
                {"source_url": {"$in": canonicals}},
            ],
        },
        {"_id": 0, "book_id": 1, "title": 1, "fanfic_urls": 1, "source_url": 1},
    )
    owned_canon: set = set()
    already_owned: List[Dict[str, Any]] = []
    async for b in owned_cursor:
        matched = None
        for c in canonicals:
            if c == b.get("source_url") or c in (b.get("fanfic_urls") or []):
                matched = c
                break
        if matched and matched not in owned_canon:
            owned_canon.add(matched)
            already_owned.append({
                "canonical": matched,
                "book_id": b["book_id"],
                "title": b.get("title") or "",
            })
    to_fetch = [c for c in canonicals if c not in owned_canon]

    # 4) Walk the list serially.
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0, "fff_options": 1, "fandom_aliases": 1})
    fff_options = (user_doc or {}).get("fff_options") or {}
    user_aliases = (user_doc or {}).get("fandom_aliases") or {}
    from routes.admin import get_global_fandom_aliases_dict
    fandom_aliases = {**(await get_global_fandom_aliases_dict()), **user_aliases}
    user_dir = STORAGE_DIR / user.user_id
    user_dir.mkdir(parents=True, exist_ok=True)

    added: List[Dict[str, Any]] = []
    failed: List[Dict[str, Any]] = []
    for canon in to_fetch:
        try:
            epub_bytes, source_meta = await fetch_fanfic_with_fallback(canon, options=fff_options)
        except FanficNotFoundError as e:
            failed.append({"canonical": canon, "error": str(e)})
            await asyncio.sleep(0.5)
            continue
        except Exception as e:  # pragma: no cover — network errors etc.
            logger.warning("url-list/pull unexpected error for %s: %s", canon, e)
            failed.append({"canonical": canon, "error": str(e)})
            await asyncio.sleep(0.5)
            continue

        # Apply the FicHub-style template if the user wants it.
        if fff_options.get("apply_template", True):
            try:
                epub_bytes = await asyncio.get_event_loop().run_in_executor(
                    None, apply_template_to_epub, epub_bytes, source_meta, canon
                )
            except Exception as e:
                logger.warning("template apply failed for %s: %s", canon, e)

        # Write the EPUB and create the book record using the same pipeline
        # the upload endpoint uses.
        book_id = f"book_{uuid.uuid4().hex[:12]}"
        epub_path = user_dir / f"{book_id}.epub"
        epub_path.write_bytes(epub_bytes)
        try:
            meta = extract_epub_metadata(epub_path)
            links = extract_urls_from_epub(epub_path) or []
            classification = await classify_book(meta)
            series_name = meta.get("series_name")
            series_index = meta.get("series_index")
            if not series_name:
                sn, si = detect_series_from_title(meta["title"])
                if sn:
                    series_name = sn
                    series_index = si if si is not None else series_index
            doc = {
                "book_id": book_id,
                "user_id": user.user_id,
                "filename": f"{(meta.get('title') or 'untitled').strip()}.epub",
                "title": meta["title"],
                "author": meta["author"],
                "description": meta["description"],
                "language": meta["language"],
                "publisher": meta["publisher"],
                "has_cover": bool(meta.get("cover_bytes")),
                "category": classification["category"],
                "fandom": _canonicalize_fandom(classification.get("fandom"), fandom_aliases),
                "confidence": classification.get("confidence"),
                "classifier": classification.get("classifier"),
                "size_bytes": len(epub_bytes),
                "links_count": len(links),
                "source_url": canon,
                "fanfic_urls": extract_fanfic_urls(links),
                "last_refreshed_at": None,
                "series_name": series_name,
                "series_index": series_index,
                "relationships": meta.get("relationships") or [],
                "rating": meta.get("rating"),
                "warnings": meta.get("warnings") or [],
                "categories": meta.get("categories") or [],
                "ao3_freeform_tags": meta.get("ao3_freeform_tags") or [],
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            if meta.get("cover_bytes"):
                (user_dir / f"{book_id}.cover").write_bytes(meta["cover_bytes"])
            (user_dir / f"{book_id}.links.txt").write_text(
                format_links_txt(meta["title"], meta["author"], links),
                encoding="utf-8",
            )
            await db.books.insert_one(doc)
            added.append({
                "canonical": canon,
                "book_id": book_id,
                "title": doc["title"],
                "fandom": doc["fandom"],
            })
        except Exception as e:
            logger.warning("url-list/pull post-fetch processing failed for %s: %s", canon, e)
            try:
                epub_path.unlink(missing_ok=True)
            except Exception:
                pass
            failed.append({"canonical": canon, "error": f"Could not process EPUB: {e}"})

    return {
        "queued": len(to_fetch),
        "added": added,
        "already_owned": already_owned,
        "failed": failed,
        "unrecognized": unrecognized,
    }

