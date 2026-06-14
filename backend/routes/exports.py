"""EPUB-bundle export (the "Download ZIP" button).  Streams a tar/zip of
the user's library through Starlette's StreamingResponse.  Extracted from
``routes/books.py`` as part of the Phase 5 refactor (2026-06-14).

Route:
    GET /api/books/export/zip
"""
import asyncio
import io
import os
import re
import zipfile
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse

from deps import db, api_router, logger
from models import User
from auth_dep import get_current_user
from routes.books import (
    OLD_STORIES_SHELF,
    STORAGE_DIR,
    _safe_filename,
    _templated_filename,
    apply_template_to_epub,
    format_links_txt,
)


def _safe_folder(name: str) -> str:
    name = re.sub(r'[^\w\s-]', '', name or 'Uncategorized').strip()
    name = re.sub(r'\s+', '_', name)
    return name or 'Uncategorized'


@api_router.get("/books/export/zip")
async def export_zip(
    request: Request,
    category: Optional[List[str]] = Query(None),
    fandom: Optional[List[str]] = Query(None),
    relationship: Optional[List[str]] = Query(None),
    author: Optional[List[str]] = Query(None),
    user: User = Depends(get_current_user),
):
    query: Dict[str, Any] = {"user_id": user.user_id}
    if category:
        query["category"] = {"$in": category} if len(category) > 1 else category[0]
    if fandom:
        query["fandom"] = {"$in": fandom} if len(fandom) > 1 else fandom[0]
    if relationship:
        # Match books listing ANY of the chosen pairings.
        query["relationships"] = {"$in": relationship} if len(relationship) > 1 else relationship[0]
    if author:
        query["author"] = {"$in": author} if len(author) > 1 else author[0]
    books = await db.books.find(query, {"_id": 0}).to_list(5000)
    if not books:
        raise HTTPException(status_code=404, detail="No books")

    # True streaming zip: bytes start flowing to the client within ~1 second,
    # even for libraries with thousands of books. We don't pre-build a temp
    # file — the archive is generated and shipped on the fly via stream-zip,
    # which means no proxy can time out waiting for the first byte.
    from stream_zip import stream_zip, ZIP_64
    from stat import S_IFREG
    from datetime import datetime as _dt

    modified_at = _dt.now()
    mode = S_IFREG | 0o600

    def _file_chunks(path: Path):
        with open(path, "rb") as f:
            while True:
                chunk = f.read(65536)
                if not chunk:
                    return
                yield chunk

    def _bytes_chunks(data: bytes):
        yield data

    # Pre-bucket books by folder path so we can sort within each folder and
    # emit a clean README. Fanfiction is grouped Fanfiction/<Fandom>/<Pairing>/,
    # with books that have no pairing landing in Fanfiction/<Fandom>/_No_pairing/.
    # Books with multiple relationships file under their FIRST one (alphabetical),
    # which matches how the dashboard relationship chips already work.
    def _folder_for(b: Dict[str, Any]) -> str:
        cat = _safe_folder(b.get('category') or 'Uncategorized')
        fnd = b.get('fandom')
        if cat == 'Fanfiction' and fnd:
            rels = b.get('relationships') or []
            if rels:
                rel = sorted([r for r in rels if r])[0]
                return f"Fanfiction/{_safe_folder(fnd)}/{_safe_folder(rel)}"
            return f"Fanfiction/{_safe_folder(fnd)}/_No_pairing"
        if cat == 'Fanfiction':
            # Fanfic with no fandom — rare but keep it visible.
            return "Fanfiction/_Unsorted"
        return cat

    buckets: Dict[str, List[Dict[str, Any]]] = {}
    for b in books:
        fp = STORAGE_DIR / user.user_id / f"{b['book_id']}.epub"
        if not fp.exists():
            continue
        buckets.setdefault(_folder_for(b), []).append(b)
    # Sort each bucket alphabetically by title (case-insensitive), tiebreak
    # on author then book_id so the ordering is deterministic.
    for folder in buckets:
        buckets[folder].sort(key=lambda x: (
            (x.get('title') or '').lower(),
            (x.get('author') or '').lower(),
            x.get('book_id') or '',
        ))

    # Build a friendly README that explains the layout + lists every folder
    # with a per-bucket book count. This goes in first so it's the first
    # thing the user sees when they open the zip.
    def _build_readme() -> bytes:
        now_str = _dt.utcnow().strftime('%Y-%m-%d %H:%M UTC')
        total_books = sum(len(v) for v in buckets.values())
        scope_lines: List[str] = []
        if fandom:
            scope_lines.append(f"Filter: fandom = {', '.join(fandom)}")
        if relationship:
            scope_lines.append(f"Filter: pairing = {', '.join(relationship)}")
        if author:
            scope_lines.append(f"Filter: author = {', '.join(author)}")
        if category:
            scope_lines.append(f"Filter: category = {', '.join(category)}")
        lines: List[str] = [
            "Shelfsort library export",
            f"Generated: {now_str}",
            f"Books: {total_books}",
            f"Folders: {len(buckets)}",
        ]
        if scope_lines:
            lines.append("")
            lines.extend(scope_lines)
        lines.extend([
            "",
            "Folder layout",
            "-------------",
            "Fanfiction/<Fandom>/<Pairing>/Title_by_Author-<id>.epub",
            "  — fanfic, grouped first by fandom (Harry Potter, Twilight, ...)",
            "    then by ship/pairing (Harry-Severus, Edward-Bella, ...).",
            "  — books with multiple pairings file under the first one (alphabetical).",
            "  — fanfic with no pairing landing under '_No_pairing/' in that fandom.",
            "",
            "<Category>/Title_by_Author-<id>.epub",
            "  — Original Fiction, Non-fiction, custom shelves, etc.",
            "",
            "Filenames mirror the user's preferred 'Title_by_Author-<short-id>' format.",
            "Books are sorted alphabetically by title within each folder.",
            "",
            "Also included: library_index.xlsx — one row per book with Folder,",
            "Fandom, Pairing, Title, Author, Source URL, Words. Paste-friendly",
            "for spreadsheets.",
            "",
            "Index",
            "-----",
        ])
        for folder, items in sorted(buckets.items()):
            lines.append(f"{folder}/  ({len(items)} book{'s' if len(items) != 1 else ''})")
        lines.append("")
        return ("\n".join(lines)).encode("utf-8")

    readme_bytes = _build_readme()

    # Excel index: one row per book with the columns the user asked for.
    # Built in-memory via openpyxl. Lives at the top of the zip alongside
    # the README so the user has a paste-friendly inventory.
    def _build_index_xlsx() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO as _BIO
        wb = Workbook()
        ws = wb.active
        ws.title = "Library"
        headers = ["Folder", "Fandom", "Pairing", "Title", "Author", "Source URL", "Words"]
        ws.append(headers)
        # Header styling — matches the existing xlsx export's look.
        head_font = Font(bold=True, color="FFFFFF")
        head_fill = PatternFill("solid", fgColor="6B46C1")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        for folder in sorted(buckets.keys()):
            for b in buckets[folder]:
                # Parse fandom + pairing from the folder path so the index
                # matches the on-disk layout exactly.
                if folder.startswith("Fanfiction/"):
                    parts = folder.split("/")
                    fnd_cell = parts[1].replace("_", " ") if len(parts) > 1 else ""
                    pair_cell = parts[2].replace("_", " ") if len(parts) > 2 else ""
                else:
                    fnd_cell = b.get("fandom") or ""
                    pair_cell = ""
                rels = b.get("relationships") or []
                if not pair_cell and rels:
                    pair_cell = sorted([r for r in rels if r])[0]
                ws.append([
                    folder,
                    fnd_cell,
                    pair_cell,
                    b.get("title") or "Untitled",
                    b.get("author") or "Unknown",
                    b.get("source_url") or "",
                    b.get("words") or "",
                ])
        # Sensible column widths + freeze + autofilter
        widths = [32, 22, 22, 38, 22, 48, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        bio = _BIO()
        wb.save(bio)
        return bio.getvalue()

    index_xlsx_bytes = _build_index_xlsx()

    def _members():
        # README + Excel index first so they're visible at the top of the archive.
        yield ("README.txt", modified_at, mode, ZIP_64, _bytes_chunks(readme_bytes))
        yield ("library_index.xlsx", modified_at, mode, ZIP_64, _bytes_chunks(index_xlsx_bytes))
        for folder in sorted(buckets.keys()):
            for b in buckets[folder]:
                fp = STORAGE_DIR / user.user_id / f"{b['book_id']}.epub"
                arcname = f"{folder}/{_templated_filename(b.get('title'), b.get('author'), b['book_id'])}"
                yield (arcname, modified_at, mode, ZIP_64, _file_chunks(fp))

    # Build a friendly zip filename. For a single-value filter we use it
    # directly; multi-value filters collapse to "filtered" to keep things short.
    def _single(v): return v[0] if v and len(v) == 1 else None
    sf, sr, sa, sc = _single(fandom), _single(relationship), _single(author), _single(category)
    zip_name = "shelfsort_library.zip"
    if sf and sr:
        zip_name = f"shelfsort_{_safe_folder(sf)}_{_safe_folder(sr)}.zip"
    elif sf:
        zip_name = f"shelfsort_{_safe_folder(sf)}.zip"
    elif sr:
        zip_name = f"shelfsort_{_safe_folder(sr)}.zip"
    elif sa:
        zip_name = f"shelfsort_{_safe_folder(sa)}.zip"
    elif sc:
        zip_name = f"shelfsort_{_safe_folder(sc)}.zip"
    elif any([fandom, relationship, author, category]):
        zip_name = "shelfsort_filtered.zip"

    logger.info(
        "export-zip streaming start: user=%s books=%d", user.user_id, len(books),
    )
    # StreamingResponse with a sync generator → Starlette runs it in a thread
    # pool, so file I/O doesn't block the event loop. No Content-Length is set
    # (chunked transfer); browsers handle this fine for downloads.
    return StreamingResponse(
        stream_zip(_members()),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{zip_name}"',
            # Tell upstreams (nginx, cloudflare) not to buffer — we want the
            # client to get bytes as fast as we produce them.
            "X-Accel-Buffering": "no",
            "Cache-Control": "no-store",
        },
    )

