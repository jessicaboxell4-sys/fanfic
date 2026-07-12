"""Links + originals download routes — extracted from ``routes/books.py`` in
the Phase 6C-A refactor (2026-07-XX).

Routes:
    GET /api/books/export/links
    GET /api/books/{book_id}/links
    GET /api/books/{book_id}/download-original

Shared helpers (``_safe_folder``, ``_safe_filename``, ``_templated_filename``,
``extract_urls_from_epub``, ``format_links_txt``) still live in
``routes/books.py`` — imported by name here.
"""
import re
import zipfile
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import Depends, HTTPException, Query, Response
from fastapi.responses import FileResponse

from deps import db, api_router, STORAGE_DIR
from models import User
from auth_dep import get_current_user

from routes.books import (
    _safe_folder,
    _safe_filename,
    _templated_filename,
    extract_urls_from_epub,
    format_links_txt,
)


@api_router.get("/books/export/links")
async def export_all_links(
    category: Optional[List[str]] = Query(None),
    fandom: Optional[List[str]] = Query(None),
    relationship: Optional[List[str]] = Query(None),
    author: Optional[List[str]] = Query(None),
    format: str = "txt",
    user: User = Depends(get_current_user),
):
    """Download every URL across the user's library (or a filter).

    ``format=txt`` (default): one combined .txt file.
    ``format=zip``: a .zip with one .txt per fandom (grouped by like fanfiction).
    ``format=xlsx``: a single .xlsx workbook with one sheet per fandom, each
        row containing the book's full metadata + extracted URL count.
    """
    query: Dict[str, Any] = {"user_id": user.user_id}
    if category:
        query["category"] = {"$in": category} if len(category) > 1 else category[0]
    if fandom:
        query["fandom"] = {"$in": fandom} if len(fandom) > 1 else fandom[0]
    if relationship:
        query["relationships"] = {"$in": relationship} if len(relationship) > 1 else relationship[0]
    if author:
        query["author"] = {"$in": author} if len(author) > 1 else author[0]
    books = await db.books.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=None)
    if not books:
        raise HTTPException(status_code=404, detail="No books")

    user_dir = STORAGE_DIR / user.user_id

    # XLSX format — single workbook, one sheet per fandom, full metadata per row
    if format == "xlsx":
        import io as _io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        # Group books by fandom (or category for non-fanfic)
        buckets: Dict[str, List[Dict[str, Any]]] = {}
        for b in books:
            cat = b.get('category') or 'Uncategorized'
            bucket = b.get('fandom') if cat == 'Fanfiction' and b.get('fandom') else cat
            buckets.setdefault(bucket, []).append(b)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="6B46C1")
        header_align = Alignment(horizontal="left", vertical="center")
        columns = [
            ("Filename", "filename", 32),
            ("Title", "title", 36),
            ("Author", "author", 22),
            ("Fandom", "fandom", 22),
            ("Rating", "rating", 14),
            ("Categories", "categories", 16),
            ("Archive Warnings", "warnings", 26),
            ("Relationships", "relationships", 30),
            ("AO3 Tags", "ao3_freeform_tags", 28),
            ("User Tags", "tags", 22),
            ("Source URL", "source_url", 60),
        ]

        # Summary sheet first
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary["A1"] = "Shelfsort library export"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        ws_summary["A3"] = f"Books total: {len(books)}"
        ws_summary["A4"] = f"Fandoms / categories: {len(buckets)}"
        ws_summary["A6"] = "Fandom / Category"
        ws_summary["B6"] = "Books"
        ws_summary["A6"].font = header_font
        ws_summary["B6"].font = header_font
        ws_summary["A6"].fill = header_fill
        ws_summary["B6"].fill = header_fill
        for i, (bk, lst) in enumerate(sorted(buckets.items()), start=7):
            ws_summary[f"A{i}"] = bk
            ws_summary[f"B{i}"] = len(lst)
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 10

        def _sheet_name(name: str) -> str:
            # Excel limits: ≤31 chars, no : \ / ? * [ ]
            cleaned = re.sub(r'[:\\/?*\[\]]', '-', name)[:31] or "Sheet"
            return cleaned

        used_names: set = {"Summary"}
        for bucket_name, bucket_books in sorted(buckets.items()):
            base = _sheet_name(bucket_name)
            name = base
            suffix = 2
            while name in used_names:
                name = (base[:28] + f"_{suffix}")[:31]
                suffix += 1
            used_names.add(name)
            ws = wb.create_sheet(title=name)
            # Header row
            for col_idx, (label, _key, width) in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=label)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                ws.column_dimensions[cell.column_letter].width = width
            ws.freeze_panes = "A2"
            # Data rows
            for r_idx, b in enumerate(bucket_books, start=2):
                for c_idx, (label, key, _w) in enumerate(columns, start=1):
                    raw = b.get(key)
                    if isinstance(raw, list):
                        value = ", ".join(str(x) for x in raw if x)
                    else:
                        value = raw or ""
                    ws.cell(row=r_idx, column=c_idx, value=value)
            ws.auto_filter.ref = ws.dimensions

        buf = _io.BytesIO()
        wb.save(buf)
        payload = buf.getvalue()
        xlsx_name = "shelfsort_library.xlsx"
        if fandom and len(fandom) == 1:
            xlsx_name = f"shelfsort_{_safe_folder(fandom[0])}.xlsx"
        elif category and len(category) == 1:
            xlsx_name = f"shelfsort_{_safe_folder(category[0])}.xlsx"
        elif any([fandom, category, relationship, author]):
            xlsx_name = "shelfsort_filtered.xlsx"
        return Response(
            content=payload,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{xlsx_name}"',
                "Content-Length": str(len(payload)),
            },
        )

    # ZIP format — one .txt per fandom (or category for non-fanfiction)
    if format == "zip":
        import io as _io
        buckets: Dict[str, List[Dict[str, Any]]] = {}
        for b in books:
            category_val = b.get('category') or 'Uncategorized'
            if category_val == 'Fanfiction':
                bucket = b.get('fandom') or 'Unsorted Fanfiction'
            else:
                bucket = category_val
            buckets.setdefault(bucket, []).append(b)

        buf = _io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            now_str = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')
            summary_lines = [
                "Shelfsort — links grouped by fandom",
                f"Generated: {now_str}",
                f"Books: {len(books)}",
                f"Fandoms / shelves: {len(buckets)}",
                "",
                "Each .txt file groups every fanfic from one fandom (or",
                "category, for non-fanfiction books). Stories are separated",
                "by blank lines so you can scan a whole fandom at a glance.",
                "",
            ]
            zf.writestr("README.txt", "\n".join(summary_lines))

            for bucket_name, bucket_books in sorted(buckets.items()):
                bucket_lines: List[str] = []
                bucket_lines.append(f"=== {bucket_name} ===")
                bucket_lines.append(
                    f"{len(bucket_books)} book{'s' if len(bucket_books) != 1 else ''} · generated {now_str}"
                )
                bucket_lines.append("")
                bucket_total = 0
                for b in bucket_books:
                    epub_path = user_dir / f"{b['book_id']}.epub"
                    bucket_lines.append(
                        f"{b.get('title','Untitled')} — {b.get('author','Unknown')}"
                    )
                    if not epub_path.exists():
                        from utils.storage_hydration import hydrate_epub_if_missing
                        await hydrate_epub_if_missing(user.user_id, b['book_id'])
                    if not epub_path.exists():
                        bucket_lines.append("  (EPUB missing on disk)")
                        bucket_lines.append("")
                        continue
                    links = extract_urls_from_epub(epub_path)
                    bucket_total += len(links)
                    if not links:
                        bucket_lines.append("  (no URLs)")
                    else:
                        for item in links:
                            anchor = item.get('anchor')
                            if anchor:
                                bucket_lines.append(f"  {item['url']}  —  {anchor}")
                            else:
                                bucket_lines.append(f"  {item['url']}")
                    bucket_lines.append("")
                bucket_lines.insert(2, f"Total URLs: {bucket_total}")
                arcname = f"{_safe_folder(bucket_name)}.txt"
                zf.writestr(arcname, "\n".join(bucket_lines) + "\n")

        payload = buf.getvalue()
        zip_name = "shelfsort_links_by_fandom.zip"
        if fandom and len(fandom) == 1:
            zip_name = f"shelfsort_{_safe_folder(fandom[0])}_links.zip"
        elif category and len(category) == 1:
            zip_name = f"shelfsort_{_safe_folder(category[0])}_links.zip"
        elif any([fandom, category, relationship, author]):
            zip_name = "shelfsort_filtered_links.zip"
        return Response(
            content=payload,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{zip_name}"',
                "Content-Length": str(len(payload)),
            },
        )

    # TXT format — combined single file (default, backward-compatible)
    scope = "your library"
    if fandom and len(fandom) == 1:
        scope = f"the {fandom[0]} shelf"
    elif category and len(category) == 1:
        scope = f"the {category[0]} shelf"
    elif any([fandom, category, relationship, author]):
        scope = "the filtered selection"

    lines: List[str] = []
    lines.append(f"Shelfsort — links extracted from {scope}")
    lines.append(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    lines.append(f"Books scanned: {len(books)}")
    lines.append("=" * 70)
    lines.append("")

    total_links = 0
    for b in books:
        epub_path = user_dir / f"{b['book_id']}.epub"
        if not epub_path.exists():
            from utils.storage_hydration import hydrate_epub_if_missing
            await hydrate_epub_if_missing(user.user_id, b['book_id'])
        if not epub_path.exists():
            continue
        links = extract_urls_from_epub(epub_path)
        total_links += len(links)

        shelf = b.get('category') or 'Uncategorized'
        if shelf == 'Fanfiction' and b.get('fandom'):
            shelf = f"Fanfiction / {b['fandom']}"

        lines.append(f"[{shelf}] {b.get('title','')} — {b.get('author','')}")
        if not links:
            lines.append("  (no URLs)")
        else:
            for item in links:
                if item.get('anchor'):
                    lines.append(f"  {item['url']}  —  {item['anchor']}")
                else:
                    lines.append(f"  {item['url']}")
        lines.append("")

    lines.insert(3, f"Total URLs:    {total_links}")
    body = "\n".join(lines) + "\n"

    fname = "shelfsort_all_links.txt"
    if fandom and len(fandom) == 1:
        fname = f"shelfsort_{_safe_folder(fandom[0])}_links.txt"
    elif category and len(category) == 1:
        fname = f"shelfsort_{_safe_folder(category[0])}_links.txt"
    elif any([fandom, category, relationship, author]):
        fname = "shelfsort_filtered_links.txt"
    headers = {"Content-Disposition": f"attachment; filename={fname}"}
    return Response(content=body, media_type="text/plain; charset=utf-8", headers=headers)


@api_router.get("/books/{book_id}/links")
async def get_book_links(book_id: str, user: User = Depends(get_current_user)):
    """Download the extracted URLs for a single book as a .txt file."""
    book = await db.books.find_one({"book_id": book_id, "user_id": user.user_id}, {"_id": 0})
    if not book:
        raise HTTPException(status_code=404, detail="Not found")

    user_dir = STORAGE_DIR / user.user_id
    links_path = user_dir / f"{book_id}.links.txt"

    # Regenerate if missing (e.g., older book uploaded before this feature)
    if not links_path.exists():
        epub_path = user_dir / f"{book_id}.epub"
        if not epub_path.exists():
            from utils.storage_hydration import hydrate_epub_if_missing
            await hydrate_epub_if_missing(user.user_id, book_id)
        if not epub_path.exists():
            raise HTTPException(status_code=404, detail="File missing")
        links = extract_urls_from_epub(epub_path)
        links_path.write_text(
            format_links_txt(book['title'], book['author'], links),
            encoding='utf-8',
        )
        await db.books.update_one(
            {"book_id": book_id, "user_id": user.user_id},
            {"$set": {"links_count": len(links)}},
        )

    filename = _safe_filename(book.get('title') or book_id, '.links.txt')
    return FileResponse(str(links_path), media_type="text/plain; charset=utf-8", filename=filename)


@api_router.get("/books/{book_id}/download-original")
async def download_original_file(book_id: str, user: User = Depends(get_current_user)):
    """Serve the user's original (pre-conversion) source file.

    Used by the Unreadable shelf when an upload was a PDF/Kindle/DOCX that
    Calibre couldn't convert — the EPUB target was never written, but the
    original bytes still live at ``{book_id}.{original_format}``. Falls
    back to whichever ``.{format}`` file actually exists on disk so this
    also works for an ``Originals`` book the user wants the source for.
    """
    book = await db.books.find_one(
        {"book_id": book_id, "user_id": user.user_id},
        {"_id": 0, "book_id": 1, "title": 1, "author": 1,
         "original_format": 1, "filename": 1},
    )
    if not book:
        raise HTTPException(status_code=404, detail="Not found")
    ext = (book.get("original_format") or "").lstrip(".")
    user_dir = STORAGE_DIR / user.user_id
    candidate = user_dir / f"{book_id}.{ext}" if ext else None
    fp = None
    if candidate and candidate.exists():
        fp = candidate
    else:
        # Last-ditch fallback: scan the user dir for any file starting with
        # the book id. Covers the case where ``original_format`` was lost or
        # stored without an extension.
        for p in user_dir.glob(f"{book_id}.*"):
            if p.suffix.lower() not in (".cover", ".links.txt"):
                fp = p
                ext = p.suffix.lstrip(".")
                break
    if not fp or not fp.exists():
        raise HTTPException(status_code=404, detail="Original file missing on disk")
    download_name = _templated_filename(
        book.get("title"), book.get("author"), book_id, ext=f".{ext or 'bin'}",
    )
    return FileResponse(str(fp), filename=download_name)
